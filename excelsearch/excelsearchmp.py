from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk

import os, sys, re, time, multiprocessing

from pathlib import Path
from threading import Thread
from multiprocessing import Process, Queue, current_process, Event, Pipe,  active_children

import openpyxl


from tkintertable.Tables import TableCanvas
from tkintertable.TableModels import TableModel

global root
global scrnwparam
global scrnhparam
scrnwparam = 300
scrnhparam = 300


def main():
    global root
    global bckgcolor
    bckgcolor = '#f7f7f7'

    root = Tk()
    root.resizable(False, False)
    
    datafile = "compare.ico"
    if not hasattr(sys, "frozen"):
        datafile = os.path.join(os.path.dirname(__file__), datafile)
    else:
        datafile = os.path.join(sys.prefix, datafile)
    root.iconbitmap(default=resource_path(datafile))
    
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('955x420+{}+{}'.format(scrnw, scrnh))
    
    app = GUI(root)
    root.mainloop()


class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background=bckgcolor)   
        self.parent = parent
        self.parent.title("Add Data v6.5")
        self.pack(fill=BOTH, expand=1)
        self.initUI()

    
    def initUI(self):
    
        global AdittCond1
        global AdittCond2
        global AdittCond3
        
        global Res2
        global Res3
        global Res4

        AdittCond1 = IntVar()
        AdittCond1.set(0)
        AdittCond2 = IntVar()
        AdittCond2.set(0)
        AdittCond3 = IntVar()
        AdittCond3.set(0)
        
        Res2 = IntVar()
        Res2.set(0)
        Res3 = IntVar()
        Res3.set(0)
        Res4 = IntVar()
        Res4.set(0)
    
        global MaxRowsSearch
        global MaxRowsData
        global PreviewRows
        global PreviewColumns
        PreviewColumns = 50
        
        global RunningCheckMaxRows
        RunningCheckMaxRows = False
        
        global RunningPrevWindowThread
        RunningPrevWindowThread = False
        
        global RunningCompareTrhead
        RunningCompareTrhead = False
        
        global RunningDuplicatesThread
        RunningDuplicatesThread = False
        
        global AbortAllTasks
        AbortAllTasks = False
        
        global prevwindevent # Declare before 
        prevwindevent = Event()
        global compareevent # Declare before 
        compareevent = Event()
        global duplicatesevent # Declare before 
        duplicatesevent = Event()
        
        global WorkingMode
        WorkingMode = "AddData"
        
        
    
    
        global LblColmName1
        LblColmName1 = Label(text="Настройка столбцов поиска", background=bckgcolor, font=("Arial", 14))
        
        global LblSearch
        LblSearch = Label(text="Искомых значений:", background=bckgcolor, font=("Arial", 11))
        
        global TBxSearch
        TBxSearch = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        global LblSearchData
        LblSearchData = Label(text="Диапазона поиска:", background=bckgcolor, font=("Arial", 11))
        
        global TBxSearchData
        TBxSearchData = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        
        
        
        global ChBxAdittCond
        ChBxAdittCond = Checkbutton(text="Дополнительное условие сопоставления 1", background=bckgcolor, variable=AdittCond1, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)

        global LblSearchAd1
        LblSearchAd1 = Label(text="Искомых значений:", background=bckgcolor, font=("Arial", 11))
        
        global TBxSearchAd1
        TBxSearchAd1 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        global LblSearchDataAd1
        LblSearchDataAd1 = Label(text="Диапазона поиска:", background=bckgcolor, font=("Arial", 11))
        
        global TBxSearchDataAd1
        TBxSearchDataAd1 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        
        global ChBxAdittCond2
        ChBxAdittCond2 = Checkbutton(text="Дополнительное условие сопоставления 2", background=bckgcolor, variable=AdittCond2, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)
        
        global LblSearchAd2
        LblSearchAd2 = Label(text="Искомых значений:", background=bckgcolor, font=("Arial", 11))
        
        global TBxSearchAd2
        TBxSearchAd2 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        global LblSearchDataAd2
        LblSearchDataAd2 = Label(text="Диапазона поиска:", background=bckgcolor, font=("Arial", 11))
        
        global TBxSearchDataAd2
        TBxSearchDataAd2 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        
        global ChBxAdittCond3
        ChBxAdittCond3 = Checkbutton(text="Дополнительное условие сопоставления 3", background=bckgcolor, variable=AdittCond3, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)
        
        global LblSearchAd3
        LblSearchAd3 = Label(text="Искомых значений:", background=bckgcolor, font=("Arial", 11))
        
        global TBxSearchAd3
        TBxSearchAd3 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        global LblSearchDataAd3
        LblSearchDataAd3 = Label(text="Диапазона поиска:", background=bckgcolor, font=("Arial", 11))
        
        global TBxSearchDataAd3
        TBxSearchDataAd3 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        
        
        
        global LblColmName2
        LblColmName2 = Label(text="Настройка столбцов результатов", background=bckgcolor, font=("Arial", 14))
        
        global LblRes1
        LblRes1 = Label(text="Результат 1 :", background=bckgcolor, font=("Arial", 11))
        
        global TbxRes1
        TbxRes1 = Entry(textvariable = "9", fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        global LblResTarget1
        LblResTarget1 = Label(text="Скопировать в столбец:", background=bckgcolor, font=("Arial", 11))
        
        global TbxResTarget1
        TbxResTarget1 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        
        global ChBxRes2
        ChBxRes2 = Checkbutton(text="Включить результат 2", background=bckgcolor, variable=Res2, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)
        
        global LblRes2
        LblRes2 = Label(text="Результат 2 :", background=bckgcolor, font=("Arial", 11))
        
        global TbxRes2
        TbxRes2 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        global LblResTarget2
        LblResTarget2 = Label(text="Скопировать в столбец:", background=bckgcolor, font=("Arial", 11))
        
        global TbxResTarget2
        TbxResTarget2 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        
        global ChBxRes3
        ChBxRes3 = Checkbutton(text="Включить результат 3", background=bckgcolor, variable=Res3, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)
        
        global LblRes3
        LblRes3 = Label(text="Результат 3 :", background=bckgcolor, font=("Arial", 11))
        
        global TbxRes3
        TbxRes3 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        global LblResTarget3
        LblResTarget3 = Label(text="Скопировать в столбец:", background=bckgcolor, font=("Arial", 11))
        
        global TbxResTarget3
        TbxResTarget3 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        
        global ChBxRes4
        ChBxRes4 = Checkbutton(text="Включить результат 4", background=bckgcolor, variable=Res4, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)
        
        global LblRes4
        LblRes4 = Label(text="Результат 4 :", background=bckgcolor, font=("Arial", 11))
        
        global TbxRes4
        TbxRes4 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        global LblResTarget4
        LblResTarget4 = Label(text="Скопировать в столбец:", background=bckgcolor, font=("Arial", 11))
        
        global TbxResTarget4
        TbxResTarget4 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        
        
        
        global InputFileLbl
        InputFileLbl = Label(text="Выберите файл таблицы:", background=bckgcolor, font=("Arial", 14))
        InputFileLbl.place(x=650, y=70)
        
        global InputFileEntry
        InputFileEntry = Entry(fg="black", bg=bckgcolor, width=27, font=("Arial", 13))
        InputFileEntry.place(x=655, y=105)
        
        global InputFileBtn
        InputFileBtn = Button(text='⮌', command=SelectFile, font=("Arial", 13))
        InputFileBtn.place(x=905, y=105, height=23)
        
        
        
        global OpenPrevTableBtn
        OpenPrevTableBtn = Button(text='Просмотр', command=RunPrevWindowThread, font=("Arial", 13))
        OpenPrevTableBtn.place(x=829, y=135, width=105, height=30)
        OpenPrevTableBtn.configure(state = DISABLED)
        
        global AddDataModeBtn
        AddDataModeBtn = Button(text='Сопоставление', command=ToggleAddDataMode, font=("Arial", 8, "bold"), bg='#9AEBB7')
        AddDataModeBtn.place(x=705, y=368, width=105, height=28)
        AddDataModeBtn.configure(state = NORMAL)
        
        global DuplicateFinderModeBtn
        DuplicateFinderModeBtn = Button(text='Поиск дубликатов', command=ToggleDuplicateFinderMode, font=("Arial", 8, "bold"))
        DuplicateFinderModeBtn.place(x=820, y=368, width=115, height=28)
        DuplicateFinderModeBtn.configure(state = NORMAL)
        
        
        
        global LblStatusAnim
        LblStatusAnim = Label(text="", background=bckgcolor, font=("Arial", 11), fg='#157d00')
        LblStatusAnim.place(x=651, y=135)
        
        global LblStatus
        LblStatus = Label(text="", background=bckgcolor, font=("Arial", 11))
        LblStatus.place(x=651, y=170)
        
        
        global LblStatus2
        LblStatus2 = Label(text="", background=bckgcolor, font=("Arial", 11))
        LblStatus2.place(x=651, y=205)
        
        global LblStatus3
        LblStatus3 = Label(text="", background=bckgcolor, font=("Arial", 11))
        LblStatus3.place(x=651, y=230)
        
        global LblProgressBar
        LblProgressBar = Label(text="", background=bckgcolor, font=("Arial", 11), justify=RIGHT)
        LblProgressBar.place(x=885, y=230, width=50)
        
        global ProgressBar
        ProgressBar = ttk.Progressbar(orient='horizontal', mode='determinate', length=278)
        #ProgressBar.place(x=654, y=260)
        

        
        
        global RunBtn
        RunBtn = Button(text='Запуск', command=RunMainThreads, font=("Arial", 13, "bold"))
        #RunBtn.place(x=735, y=285+5, width=198, height=40)
        
        global CancelBtn
        CancelBtn = Button(text='🞭', command=KillProcc, font=("Arial", 25, "bold"), fg='#ff0000')
        #CancelBtn.place(x=694, y=285+5, width=40, height=40)
        
        global ExcBtn
        ExcBtn = Button(text='🕮', command=OpenWithSystem, font=("Arial", 25, "bold"), bg='#9dfab1')
        #ExcBtn.place(x=653, y=285+5, width=40, height=40)
        
        
        
        
        global EmptyRowBtn1
        EmptyRowBtn1 = Button(text='🡿', command=PrintRow, font=("Arial", 13))
        #EmptyRowBtn1.place(x=905, y=17, width=27, height=23)
        
        global EmptyRowBtn2
        EmptyRowBtn2 = Button(text='🟆', command=PrintProccList, font=("Arial", 13))
        #EmptyRowBtn2.place(x=850, y=17, width=27, height=23)
        
        global EmptyRowBtn3
        EmptyRowBtn3 = Button(text='🞭', command=KillProcc, font=("Arial", 13))
        #EmptyRowBtn3.place(x=800, y=17, width=27, height=23)
        
        
        global LblColmNameDup
        LblColmNameDup = Label(text="Поиск дубликатов", background=bckgcolor, font=("Arial", 14))
        
        global LblSearchDup
        LblSearchDup = Label(text="Колонка поиска:", background=bckgcolor, font=("Arial", 11))
        
        global TBxSearchDup
        TBxSearchDup = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        
        global LblResultDup
        LblResultDup = Label(text="Колонка результата:", background=bckgcolor, font=("Arial", 11))
        
        global TBxResultDup
        TBxResultDup = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))

        
        
        #ToggleDuplicateFinderMode()
        ToggleAddDataMode()
        InterfaceTrigger()

  


def SelectFile():
    print('**** SF: Started')
    global InputFile
    global OpenPreview

    InputFile = ""
    InputFile = filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx")])
    if InputFile:
        InputFileEntry.configure(state = NORMAL)
        InputFileEntry.delete(0,END)
        InputFileEntry.insert(0,str(Path(InputFile).name))
        InputFileEntry.configure(state = DISABLED)
        
        OpenPrevTableBtn.configure(state = NORMAL)
        OpenPreview = True
        
        RunBtn.place_forget()
        CancelBtn.place_forget()
        ExcBtn.place_forget()
        
        maxrowsthread = Thread(target=CheckMaxRows)
        maxrowsthread.start()
        maxrowsthread = ""
        
        taskswatchdogthread = Thread(target=TasksWatchdog)
        taskswatchdogthread.start()
        taskswatchdogthread = ""
        
        print('**** SF: InputFile : {0}'.format(InputFile))
    else:
        print('**** SF: InputFile not selected !')


    

def CheckMaxRows():
    print('***** CMR: Started')
    global PreviewRows
    global PreviewColumns
    
    global RunningCheckMaxRows
    RunningCheckMaxRows = True

    book = openpyxl.load_workbook(InputFile)
    sheet = book.active
    
    EmptyCellCounter = 0
    EmptyRowCounter = 0
    
    for t in range(500000):
        EmptyCellCounter = 0
        for y in range(PreviewColumns):
            if sheet.cell(row=t+1, column=y+1).value is None:
                EmptyCellCounter = EmptyCellCounter + 1

        if EmptyCellCounter == PreviewColumns:
            EmptyRowCounter = EmptyRowCounter + 1
        else: 
            EmptyRowCounter = 0
        if EmptyRowCounter >= 10:
            PreviewRows = t - 9
            break
    book.close()
    
    print('***** CMR: PreviewRows = ', PreviewRows)

    RunningCheckMaxRows = False
    
    global prevwindevent
    prevwindevent = Event()
    prevwindowthread = Thread(target=PrevWindowThread)
    prevwindowthread.start()
    
    RunBtn.place(x=735, y=285+5, width=198, height=40)
    CancelBtn.place(x=694, y=285+5, width=40, height=40)
    ExcBtn.place(x=653, y=285+5, width=40, height=40)
    
    print('***** CMR: Ended')
    



def RunDeprecated():
    global InputFile
    global PreviewRows
    MaxRowsSearch = PreviewRows
    MaxRowsData = PreviewRows
    
    RunBtn.configure(state = DISABLED)
    ExcBtn.configure(state = DISABLED)
    
    book = openpyxl.load_workbook(InputFile)
    sheet = book.active
    
    
    # Main data
    ColmSearch = int(TBxSearch.get())
    ColmData = int(TBxSearchData.get())
    
    ColmResult1 = int(TbxRes1.get())
    ColmResultTarget1 = int(TbxResTarget1.get())
    
    ArrSearch = []
    ArrData = []
    
    
    # Main conditions array appending
    for i in range(MaxRowsSearch):
        ArrSearch.append(sheet.cell(row=i+1, column=ColmSearch).value)   
    for k in range(MaxRowsData):
        ArrData.append(sheet.cell(row=k+1, column=ColmData).value)
    
    
    # Additional conditions arrays appending
    if AdittCond1.get() == 1:
        ColmAddSearch1 = int(TBxSearchAd1.get())
        ColmAddData1 = int(TBxSearchDataAd1.get())
        ArrAddSearch1 = []
        ArrAddData1 = []
        for i in range(MaxRowsSearch):
            ArrAddSearch1.append(sheet.cell(row=i+1, column=ColmAddSearch1).value)
            ArrAddData1.append(sheet.cell(row=i+1, column=ColmAddData1).value)
            
    if AdittCond2.get() == 1:
        ColmAddSearch2 = int(TBxSearchAd2.get())
        ColmAddData2 = int(TBxSearchDataAd2.get())
        ArrAddSearch2 = []
        ArrAddData2 = []
        for i in range(MaxRowsSearch):
            ArrAddSearch2.append(sheet.cell(row=i+1, column=ColmAddSearch2).value)
            ArrAddData2.append(sheet.cell(row=i+1, column=ColmAddData2).value)
            
    if AdittCond3.get() == 1:
        ColmAddSearch3 = int(TBxSearchAd3.get())
        ColmAddData3 = int(TBxSearchDataAd3.get())
        ArrAddSearch3 = []
        ArrAddData3 = []
        for i in range(MaxRowsSearch):
            ArrAddSearch3.append(sheet.cell(row=i+1, column=ColmAddSearch3).value)
            ArrAddData3.append(sheet.cell(row=i+1, column=ColmAddData3).value)
    
    
    # Additional results
    if Res2.get() == 1:
        ColmResult2 = int(TbxRes2.get())
        ColmResultTarget2 = int(TbxResTarget2.get())
    if Res3.get() == 1:
        ColmResult3 = int(TbxRes3.get())
        ColmResultTarget3 = int(TbxResTarget3.get())
    if Res4.get() == 1:
        ColmResult4 = int(TbxRes4.get())
        ColmResultTarget4 = int(TbxResTarget4.get())



    # Starting search
    SearchedRowsCount = 0
    FindedRowsCount = 0
    for m in range(MaxRowsSearch):
        SearchedRowsCount = SearchedRowsCount + 1
        if ArrSearch[m]!=None:
            for n in range(MaxRowsData):
                Finded = False
                if ArrSearch[m] == ArrData[n]:
                    Finded = True
                if AdittCond1.get() == 1 and ArrAddSearch1[m] != ArrAddData1[n]:
                    Finded = False
                if AdittCond2.get() == 1 and ArrAddSearch2[m] != ArrAddData2[n]:
                    Finded = False
                if AdittCond3.get() == 1 and ArrAddSearch3[m] != ArrAddData3[n]:
                    Finded = False
                    
                if Finded == True:
                    sheet.cell(row=m+1, column=ColmResultTarget1).value = sheet.cell(row=n+1, column=ColmResult1).value
                    if Res2.get() == 1:
                        sheet.cell(row=m+1, column=ColmResultTarget2).value = sheet.cell(row=n+1, column=ColmResult2).value
                    if Res3.get() == 1:
                        sheet.cell(row=m+1, column=ColmResultTarget3).value = sheet.cell(row=n+1, column=ColmResult3).value
                    if Res4.get() == 1:
                        sheet.cell(row=m+1, column=ColmResultTarget4).value = sheet.cell(row=n+1, column=ColmResult4).value
                    print('m={0}, finded: {1} = {2}'.format(m, ArrSearch[m], ArrData[n]))
                    FindedRowsCount = FindedRowsCount + 1
                    break
                
    book.save(InputFile)
    
    
    RunBtn.configure(state = NORMAL)
    ExcBtn.configure(state = NORMAL)
    
    msgbxlbl = ['Обработка завершена !', 'Обработано: {0}'.format(SearchedRowsCount), 'Найдено: {0}'.format(FindedRowsCount)]
    messagebox.showinfo("", "\n".join(msgbxlbl))


def RunMainThreads():
    ProgressBar.place(x=699-45, y=255+5)
    
    if WorkingMode == "AddData":
        comparethread = Thread(target=CompareThread)
        comparethread.start()
        comparethread = ""
    elif WorkingMode == "DuplicateFinder":
        duplicatesthread = Thread(target=DuplicatesThread)
        duplicatesthread.start()
        duplicatesthread = ""
    
    taskswatchdogthread = Thread(target=TasksWatchdog)
    taskswatchdogthread.start()
    taskswatchdogthread = ""
    
    
def CompareThread(): # Thread for process start 
    print('====== CT Started')
    global InputFile
    global PreviewRows
    
    global compareprocess
    global compareevent
    conn1, conn2 = Pipe()
    
    global RunningCompareTrhead
    RunningCompareTrhead = True
    global AbortAllTasks
    
    # Main data get
    global TBxSearch  
    global TBxSearchData
    ColmSearch = int(TBxSearch.get())
    ColmData = int(TBxSearchData.get())
    
    # Adittional data get
    global ChBxAdittCond
    global TBxSearchAd1
    global TBxSearchDataAd1
    EnableAdittCond1 = bool(AdittCond1.get())
    if EnableAdittCond1 == True:
        ColmAddSearch1 = int(TBxSearchAd1.get())
        ColmAddData1 = int(TBxSearchDataAd1.get())
    else:
        ColmAddSearch1 = 0
        ColmAddData1 = 0
    
    global ChBxAdittCond2
    global TBxSearchAd2
    global TBxSearchDataAd2
    EnableAdittCond2 = bool(AdittCond2.get())
    if EnableAdittCond2 == True:
        ColmAddSearch2 = int(TBxSearchAd2.get())
        ColmAddData2 = int(TBxSearchDataAd2.get())
    else:
        ColmAddSearch2 = 0
        ColmAddData2 = 0
    
    global AdittCond3
    global TBxSearchAd3
    global TBxSearchDataAd3
    EnableAdittCond3 = bool(AdittCond3.get())
    if EnableAdittCond3 == True:
        ColmAddSearch3 = int(TBxSearchAd3.get())
        ColmAddData3 = int(TBxSearchDataAd3.get())
    else:
        ColmAddSearch3 = 0
        ColmAddData3 = 0
    
    
    # Results data get
    global TbxRes1
    global TbxResTarget1
    ColmResult1 = int(TbxRes1.get())
    ColmResultTarget1 = int(TbxResTarget1.get())
    
    global Res2
    global TbxRes2
    global TbxResTarget2
    EnableRes2 = bool(Res2.get())
    if EnableRes2 == True:
        ColmResult2 = int(TbxRes2.get())
        ColmResultTarget2 = int(TbxResTarget2.get())
    else:
        ColmResult2 = 0
        ColmResultTarget2 = 0
    
    global Res3
    global TbxRes3
    global TbxResTarget3
    EnableRes3 = bool(Res3.get())
    if EnableRes3 == True:
        ColmResult3 = int(TbxRes3.get())
        ColmResultTarget3 = int(TbxResTarget3.get())
    else:
        ColmResult3 = 0
        ColmResultTarget3 = 0
    
    global Res4
    global TbxRes4
    global TbxResTarget4
    EnableRes4 = bool(Res4.get())
    if EnableRes4 == True:
        ColmResult4 = int(TbxRes4.get())
        ColmResultTarget4 = int(TbxResTarget4.get())
    else:
        ColmResult4 = 0
        ColmResultTarget4 = 0
    
    
    compareprocess = Process(target=CompareProcess, args=(conn2, compareevent, InputFile, PreviewRows, ColmSearch, ColmData, EnableAdittCond1, ColmAddSearch1, ColmAddData1, EnableAdittCond2, ColmAddSearch2, ColmAddData2, EnableAdittCond3, ColmAddSearch3, ColmAddData3, ColmResult1, ColmResultTarget1, EnableRes2, ColmResult2, ColmResultTarget2, EnableRes3, ColmResult3, ColmResultTarget3, EnableRes4, ColmResult4, ColmResultTarget4))
    compareprocess.start()
    
    while RunningCompareTrhead:
        processresult = conn1.recv()
        
        LblStatus2.config(text = "Обработано: {0} / {1}".format(processresult[0], PreviewRows))
        LblStatus3.config(text = "Найдено: {0}".format(processresult[1]))
        percentage = int (processresult[0]) / (PreviewRows / 100)
        ProgressBar['value'] = percentage
        LblProgressBar.config(text = "{0}%".format(int(percentage)))
        
        #time.sleep(0.01)
    
    if not AbortAllTasks:
        RunPrevWindowThread()
    
    print('====== CT Ended')


def CompareProcess(conn2, compareevent, InputFile, PreviewRows, ColmSearch, ColmData, EnableAdittCond1, ColmAddSearch1, ColmAddData1, EnableAdittCond2, ColmAddSearch2, ColmAddData2, EnableAdittCond3, ColmAddSearch3, ColmAddData3, ColmResult1, ColmResultTarget1, EnableRes2, ColmResult2, ColmResultTarget2, EnableRes3, ColmResult3, ColmResultTarget3, EnableRes4, ColmResult4, ColmResultTarget4): 
    print('######## pCP Started')

    MaxRowsSearch = PreviewRows
    MaxRowsData = PreviewRows

    book = openpyxl.load_workbook(InputFile)
    sheet = book.active

    ArrSearch = []
    ArrData = []
    
    
    # Main conditions array appending
    for i in range(MaxRowsSearch):
        ArrSearch.append(sheet.cell(row=i+1, column=ColmSearch).value)   
    for k in range(MaxRowsData):
        ArrData.append(sheet.cell(row=k+1, column=ColmData).value)
    
    
    # Additional conditions arrays appending
    if EnableAdittCond1 == True:
        ArrAddSearch1 = []
        ArrAddData1 = []
        for i in range(MaxRowsSearch):
            ArrAddSearch1.append(sheet.cell(row=i+1, column=ColmAddSearch1).value)
            ArrAddData1.append(sheet.cell(row=i+1, column=ColmAddData1).value)
            
    if EnableAdittCond2 == True:
        ArrAddSearch2 = []
        ArrAddData2 = []
        for i in range(MaxRowsSearch):
            ArrAddSearch2.append(sheet.cell(row=i+1, column=ColmAddSearch2).value)
            ArrAddData2.append(sheet.cell(row=i+1, column=ColmAddData2).value)
            
    if EnableAdittCond3 == True:
        ArrAddSearch3 = []
        ArrAddData3 = []
        for i in range(MaxRowsSearch):
            ArrAddSearch3.append(sheet.cell(row=i+1, column=ColmAddSearch3).value)
            ArrAddData3.append(sheet.cell(row=i+1, column=ColmAddData3).value)


    # Starting search
    SearchedRowsCount = 0
    FindedRowsCount = 0
    for m in range(MaxRowsSearch):
        SearchedRowsCount = SearchedRowsCount + 1
        if ArrSearch[m]!=None:
            for n in range(MaxRowsData):
                Finded = False
                if ArrSearch[m] == ArrData[n]:
                    Finded = True
                if EnableAdittCond1 == True and ArrAddSearch1[m] != ArrAddData1[n]:
                    Finded = False
                if EnableAdittCond2 == True and ArrAddSearch2[m] != ArrAddData2[n]:
                    Finded = False
                if EnableAdittCond3 == True and ArrAddSearch3[m] != ArrAddData3[n]:
                    Finded = False
                    
                if Finded == True:
                    sheet.cell(row=m+1, column=ColmResultTarget1).value = sheet.cell(row=n+1, column=ColmResult1).value
                    if EnableRes2 == True:
                        sheet.cell(row=m+1, column=ColmResultTarget2).value = sheet.cell(row=n+1, column=ColmResult2).value
                    if EnableRes3 == True:
                        sheet.cell(row=m+1, column=ColmResultTarget3).value = sheet.cell(row=n+1, column=ColmResult3).value
                    if EnableRes4 == True:
                        sheet.cell(row=m+1, column=ColmResultTarget4).value = sheet.cell(row=n+1, column=ColmResult4).value
                    #print('m={0}, finded: {1} = {2}'.format(m, ArrSearch[m], ArrData[n]))
                    FindedRowsCount = FindedRowsCount + 1
                    break
        conn2.send([SearchedRowsCount, FindedRowsCount])
    book.save(InputFile)

    compareevent.set()
    
    msgbxlbl = ['Обработка завершена !', 'Обработано: {0}'.format(SearchedRowsCount), 'Найдено: {0}'.format(FindedRowsCount)]
    messagebox.showinfo("", "\n".join(msgbxlbl))
    
    print('######## pCP Ended')



    
def DuplicatesThread(): # Thread for process start 
    print('====== DT Started')
    global InputFile
    global PreviewRows
    
    global duplicatesprocess
    global duplicatesevent
    conn1, conn2 = Pipe()
    
    global RunningDuplicatesThread
    RunningDuplicatesThread = True
    global AbortAllTasks
    
    # Main data get
    global LblSearchDup  
    global TBxResultDup
    ColmSearchDup = int(TBxSearchDup.get())
    ColmResultDup = int(TBxResultDup.get())
    

    
    duplicatesprocess = Process(target=DuplicatesProcess, args=(conn2, duplicatesevent, InputFile, PreviewRows, ColmSearchDup, ColmResultDup))
    duplicatesprocess.start()
    
    while RunningDuplicatesThread:
        processresult = conn1.recv()
        
        LblStatus2.config(text = "Обработано: {0} / {1}".format(processresult[0], PreviewRows))
        LblStatus3.config(text = "Найдено: {0}".format(processresult[1]))
        percentage = int (processresult[0]) / (PreviewRows / 100)
        ProgressBar['value'] = percentage
        LblProgressBar.config(text = "{0}%".format(int(percentage)))
        
        #time.sleep(0.01)
    
    if not AbortAllTasks:
        RunPrevWindowThread()
    
    print('====== DT Ended')


def DuplicatesProcess(conn2, duplicatesevent, InputFile, PreviewRows, ColmSearchDup, ColmResultDup): 
    print('######## pDP Started')

    MaxRowsSearch = PreviewRows

    book = openpyxl.load_workbook(InputFile)
    sheet = book.active

    ArrSearch = []
    ArrData = []
    
    
    # Main conditions array appending
    for i in range(MaxRowsSearch):
        ArrSearch.append(sheet.cell(row=i+1, column=ColmSearchDup).value)   


    # Starting search
    SearchedRowsCount = 0
    FindedRowsCount = 0
    for m in range(MaxRowsSearch-1):
        SearchedRowsCount = SearchedRowsCount + 1

        if ArrSearch[m]!=None:
            for n in range(m+1, MaxRowsSearch):
                if ArrSearch[m] == ArrSearch[n]:
                    sheet.cell(row=m+1, column=ColmResultDup).value = 'Дубликат {0}'.format(m)
                    sheet.cell(row=n+1, column=ColmResultDup).value = 'Дубликат {0}'.format(m)
                    #print('m={0}, finded: {1} = {2}'.format(m, ArrSearch[m], ArrData[n]))
                    FindedRowsCount = FindedRowsCount + 1
        
        conn2.send([SearchedRowsCount, FindedRowsCount])
    book.save(InputFile)

    duplicatesevent.set()
    
    msgbxlbl = ['Обработка завершена !', 'Обработано: {0}'.format(SearchedRowsCount), 'Найдено: {0}'.format(FindedRowsCount)]
    messagebox.showinfo("", "\n".join(msgbxlbl))
    
    print('######## pDP Ended')




def RunPrevWindowThread(): # For button click
    prevwindowthread = Thread(target=PrevWindowThread)
    prevwindowthread.start()
    prevwindowthread = ""
    
    taskswatchdogthread = Thread(target=TasksWatchdog)
    taskswatchdogthread.start()
    taskswatchdogthread = ""
    

def PrevWindowThread(): # Thread for process start 
    print('====== PWT Started')
    global InputFile
    global PreviewRows
    
    global prevwindowprocess
    global prevwindevent
    
    global RunningPrevWindowThread
    RunningPrevWindowThread = True
    
    prevwindowprocess = Process(target=PrevWindowProcess, args=(prevwindevent, InputFile, PreviewRows))
    prevwindowprocess.start()
    print('====== PWT Ended')


def PrevWindowProcess(prevwindevent, InputFile, PreviewRows):
    print('######## pPW Started')
    global pInputFile
    pInputFile = InputFile
    
    global pPreviewRows
    pPreviewRows = PreviewRows
    
    global pPreviewWindowOpened
    pPreviewWindowOpened = False
    
    window=Tk()
    window.title(Path(InputFile).name)
    window.geometry("1200x550")
    window.title("Add Data v5.1")
    
    t_frame=Frame(window)
    t_frame.pack(fill='both', expand=True)
    
    model = TableModel()
    table = TableCanvas(t_frame, model=model, editable=False)
    
    data = table.model.data
    cols = table.model.columnNames
    
    book = openpyxl.load_workbook(pInputFile)
    sheet = book.active
    
    PreviewColumns = 50
    for i in range(PreviewColumns):
        table.model.addColumn()
    
    for k in range(pPreviewRows):
        table.model.addRow()
    
    for r in range(pPreviewRows):
        for c in range(PreviewColumns):
            if sheet.cell(row=r+1, column=c+1).value != None:
                table.model.setValueAt(str(sheet.cell(row=r+1, column=c+1).value),r,c)
    
    book.close()
    table.show()
    prevwindevent.set()
    window.mainloop()
    
    print('######## pPW Ended')




def TasksWatchdog():
    global RunningCheckMaxRows
    global RunningPrevWindowThread
    global RunningCompareTrhead
    global RunningDuplicatesThread
    global prevwindevent
    global compareevent
    global duplicatesevent
    
    InterfaceDisabler(0)
    
    time.sleep(0.5)
    animframe = 0

    finished = False
    while not finished:
    
        
        if RunningCheckMaxRows == True:
            LblStatus.config(text = 'Подсчет максимального числа строк ...')
            
        if RunningPrevWindowThread == True:
            LblStatus.config(text = 'Открытие таблицы данных ...')
            if prevwindevent.is_set():
                prevwindevent.clear()
                RunningPrevWindowThread = False
                finished = True
                break
            
        if RunningCompareTrhead == True:
            LblStatus.config(text = 'Сопоставление данных ...')
            if compareevent.is_set():
                compareevent.clear()
                RunningCompareTrhead = False
                finished = True
                break
                
        if RunningDuplicatesThread == True:
            LblStatus.config(text = 'Поиск дубликатов ...')
            if duplicatesevent.is_set():
                duplicatesevent.clear()
                RunningDuplicatesThread = False
                finished = True
                break


        LblStatusAnim.place(x=649, y=136)
        animframe = animframe + 1
        if animframe == 1:
            LblStatusAnim.config(text = '🏾🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻', font=("Arial", 14))
        elif animframe == 2:
            LblStatusAnim.config(text = '🏿🏾🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻')
        elif animframe == 3:
            LblStatusAnim.config(text = '🏾🏿🏾🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻')
        elif animframe == 4:
            LblStatusAnim.config(text = '🏻🏾🏿🏾🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻')
        elif animframe == 5:
            LblStatusAnim.config(text = '🏻🏻🏾🏿🏾🏻🏻🏻🏻🏻🏻🏻🏻🏻')
        elif animframe == 6:
            LblStatusAnim.config(text = '🏻🏻🏻🏾🏿🏾🏻🏻🏻🏻🏻🏻🏻🏻')
        elif animframe == 7:
            LblStatusAnim.config(text = '🏻🏻🏻🏻🏾🏿🏾🏻🏻🏻🏻🏻🏻🏻')
        elif animframe == 8:
            LblStatusAnim.config(text = '🏻🏻🏻🏻🏻🏾🏿🏾🏻🏻🏻🏻🏻🏻')
        elif animframe == 9:
            LblStatusAnim.config(text = '🏻🏻🏻🏻🏻🏻🏾🏿🏾🏻🏻🏻🏻🏻')
        elif animframe == 10:
            LblStatusAnim.config(text = '🏻🏻🏻🏻🏻🏻🏻🏾🏿🏾🏻🏻🏻🏻')
        elif animframe == 11:
            LblStatusAnim.config(text = '🏻🏻🏻🏻🏻🏻🏻🏻🏾🏿🏾🏻🏻🏻')
        elif animframe == 12:
            LblStatusAnim.config(text = '🏻🏻🏻🏻🏻🏻🏻🏻🏻🏾🏿🏾🏻🏻')
        elif animframe == 13:
            LblStatusAnim.config(text = '🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏾🏿🏾🏻')
        elif animframe == 14:
            LblStatusAnim.config(text = '🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏾🏿🏾')
        elif animframe == 15:
            LblStatusAnim.config(text = '🏾🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏾🏿')
        elif animframe == 16:
            LblStatusAnim.config(text = '🏿🏾🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏾')
        elif animframe == 17:
            animframe = 3
            LblStatusAnim.config(text = '🏾🏿🏾🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻🏻')
        time.sleep(0.1)
    
    InterfaceDisabler(1)
    
    LblStatusAnim.place(x=651, y=145)
    if not AbortAllTasks:
        LblStatusAnim.config(text = 'Выполнено !', font=("Arial", 11))
        LblStatus.config(text = 'Таблицы данных открыта !')
    else:
        LblStatusAnim.config(text = 'Выполнено !', font=("Arial", 11))
        LblStatus.config(text = 'Задачи отменены !')


def InterfaceDisabler(enable):
    if enable == 1:
        RunBtn.configure(state = NORMAL)
        RunBtn.configure(bg = "#e1eafa")
        CancelBtn.configure(state = DISABLED)
        #CancelBtn.configure(bg = "#f0f0f0")
        ExcBtn.configure(state = NORMAL)
        ExcBtn.configure(bg = "#9dfab1")
        OpenPrevTableBtn.place(x=829, y=135, width=105, height=30)
        InputFileBtn.configure(state = NORMAL)
        
        ChBxAdittCond.configure(state = NORMAL)
        ChBxAdittCond2.configure(state = NORMAL)
        ChBxAdittCond3.configure(state = NORMAL)
        ChBxRes2.configure(state = NORMAL)
        ChBxRes3.configure(state = NORMAL)
        ChBxRes4.configure(state = NORMAL)
        
        TBxSearch.configure(state = NORMAL)
        TBxSearchData.configure(state = NORMAL)
        
        LblSearchAd1.configure(state = NORMAL)
        LblSearchDataAd1.configure(state = NORMAL)
        TBxSearchAd1.configure(state = NORMAL)
        TBxSearchDataAd1.configure(state = NORMAL)
        
        LblSearchAd2.configure(state = NORMAL)
        LblSearchDataAd2.configure(state = NORMAL)
        TBxSearchAd2.configure(state = NORMAL)
        TBxSearchDataAd2.configure(state = NORMAL)
        
        LblSearchAd3.configure(state = NORMAL)
        LblSearchDataAd3.configure(state = NORMAL)
        TBxSearchAd3.configure(state = NORMAL)
        TBxSearchDataAd3.configure(state = NORMAL)
        
        TbxRes1.configure(state = NORMAL)
        TbxResTarget1.configure(state = NORMAL)
        
        LblRes2.configure(state = NORMAL)
        LblResTarget2.configure(state = NORMAL)
        TbxRes2.configure(state = NORMAL)
        TbxResTarget2.configure(state = NORMAL)
        
        LblRes3.configure(state = NORMAL)
        LblResTarget3.configure(state = NORMAL)
        TbxRes3.configure(state = NORMAL)
        TbxResTarget3.configure(state = NORMAL)
        
        LblRes4.configure(state = NORMAL)
        LblResTarget4.configure(state = NORMAL)
        TbxRes4.configure(state = NORMAL)
        TbxResTarget4.configure(state = NORMAL)
        
        
        AddDataModeBtn.configure(state = NORMAL)
        DuplicateFinderModeBtn.configure(state = NORMAL)

        TBxSearchDup.configure(state = NORMAL)
        TBxResultDup.configure(state = NORMAL)
        
    else:
        RunBtn.configure(state = DISABLED)
        RunBtn.configure(bg = "#f0f0f0")
        CancelBtn.configure(state = NORMAL)
        #CancelBtn.configure(bg='#f0f0f0')
        ExcBtn.configure(state = DISABLED)
        ExcBtn.configure(bg = "#f0f0f0")
        OpenPrevTableBtn.place_forget()
        InputFileBtn.configure(state = DISABLED)
        
        ChBxAdittCond.configure(state = DISABLED)
        ChBxAdittCond2.configure(state = DISABLED)
        ChBxAdittCond3.configure(state = DISABLED)
        ChBxRes2.configure(state = DISABLED)
        ChBxRes3.configure(state = DISABLED)
        ChBxRes4.configure(state = DISABLED)
        
        TBxSearch.configure(state = DISABLED)
        TBxSearchData.configure(state = DISABLED)
        
        LblSearchAd1.configure(state = DISABLED)
        LblSearchDataAd1.configure(state = DISABLED)
        TBxSearchAd1.configure(state = DISABLED)
        TBxSearchDataAd1.configure(state = DISABLED)
        
        LblSearchAd2.configure(state = DISABLED)
        LblSearchDataAd2.configure(state = DISABLED)
        TBxSearchAd2.configure(state = DISABLED)
        TBxSearchDataAd2.configure(state = DISABLED)
        
        LblSearchAd3.configure(state = DISABLED)
        LblSearchDataAd3.configure(state = DISABLED)
        TBxSearchAd3.configure(state = DISABLED)
        TBxSearchDataAd3.configure(state = DISABLED)
        
        TbxRes1.configure(state = DISABLED)
        TbxResTarget1.configure(state = DISABLED)
        
        LblRes2.configure(state = DISABLED)
        LblResTarget2.configure(state = DISABLED)
        TbxRes2.configure(state = DISABLED)
        TbxResTarget2.configure(state = DISABLED)
        
        LblRes3.configure(state = DISABLED)
        LblResTarget3.configure(state = DISABLED)
        TbxRes3.configure(state = DISABLED)
        TbxResTarget3.configure(state = DISABLED)
        
        LblRes4.configure(state = DISABLED)
        LblResTarget4.configure(state = DISABLED)
        TbxRes4.configure(state = DISABLED)
        TbxResTarget4.configure(state = DISABLED)
        
        
        AddDataModeBtn.configure(state = DISABLED)
        DuplicateFinderModeBtn.configure(state = DISABLED)

        TBxSearchDup.configure(state = DISABLED)
        TBxResultDup.configure(state = DISABLED)


def InterfaceTrigger():

    if AdittCond1.get() == 1:
        LblSearchAd1.configure(state = NORMAL)
        LblSearchDataAd1.configure(state = NORMAL)
        TBxSearchAd1.configure(state = NORMAL)
        TBxSearchDataAd1.configure(state = NORMAL)
    else:
        LblSearchAd1.configure(state = DISABLED)
        LblSearchDataAd1.configure(state = DISABLED)
        TBxSearchAd1.configure(state = DISABLED)
        TBxSearchDataAd1.configure(state = DISABLED)


    if AdittCond2.get() == 1:
        LblSearchAd2.configure(state = NORMAL)
        LblSearchDataAd2.configure(state = NORMAL)
        TBxSearchAd2.configure(state = NORMAL)
        TBxSearchDataAd2.configure(state = NORMAL)
    else:
        LblSearchAd2.configure(state = DISABLED)
        LblSearchDataAd2.configure(state = DISABLED)
        TBxSearchAd2.configure(state = DISABLED)
        TBxSearchDataAd2.configure(state = DISABLED)
    
    
    if AdittCond3.get() == 1:
        LblSearchAd3.configure(state = NORMAL)
        LblSearchDataAd3.configure(state = NORMAL)
        TBxSearchAd3.configure(state = NORMAL)
        TBxSearchDataAd3.configure(state = NORMAL)
    else:
        LblSearchAd3.configure(state = DISABLED)
        LblSearchDataAd3.configure(state = DISABLED)
        TBxSearchAd3.configure(state = DISABLED)
        TBxSearchDataAd3.configure(state = DISABLED)


    if Res2.get() == 1:
        LblRes2.configure(state = NORMAL)
        LblResTarget2.configure(state = NORMAL)
        TbxRes2.configure(state = NORMAL)
        TbxResTarget2.configure(state = NORMAL)
    else:
        LblRes2.configure(state = DISABLED)
        LblResTarget2.configure(state = DISABLED)
        TbxRes2.configure(state = DISABLED)
        TbxResTarget2.configure(state = DISABLED)


    if Res3.get() == 1:
        LblRes3.configure(state = NORMAL)
        LblResTarget3.configure(state = NORMAL)
        TbxRes3.configure(state = NORMAL)
        TbxResTarget3.configure(state = NORMAL)
    else:
        LblRes3.configure(state = DISABLED)
        LblResTarget3.configure(state = DISABLED)
        TbxRes3.configure(state = DISABLED)
        TbxResTarget3.configure(state = DISABLED)
    
    
    if Res4.get() == 1:
        LblRes4.configure(state = NORMAL)
        LblResTarget4.configure(state = NORMAL)
        TbxRes4.configure(state = NORMAL)
        TbxResTarget4.configure(state = NORMAL)
    else:
        LblRes4.configure(state = DISABLED)
        LblResTarget4.configure(state = DISABLED)
        TbxRes4.configure(state = DISABLED)
        TbxResTarget4.configure(state = DISABLED)




def OpenWithSystem():
    global InputFile
    os.system('"{0}"'.format(InputFile))


def KillProcc():
    global compareevent
    global prevwindevent
    global AbortAllTasks
    global RunningPrevWindowThread
    global RunningCompareTrhead
    global RunningDuplicatesThread 

    AbortAllTasks = True

    active = active_children()
    for child in active:
        child.kill()
    
    
    prevwindevent.set()
    compareevent.set()
    time.sleep(0.5)
    RunningPrevWindowThread = False
    RunningCompareTrhead = False
    RunningDuplicatesThread = False
    
    
    messagebox.showerror("", "Задачи отменены !")
    AbortAllTasks = False
    

def PrintProccList():
    children = active_children()
    print(children)


def PrintRow():
    print('  ')
    
 

 
def ToggleAddDataMode():
    global WorkingMode
    WorkingMode = "AddData"
    
    AddDataModeBtn.configure(bg='#9AEBB7')
    DuplicateFinderModeBtn.configure(bg=bckgcolor)

    LblColmName1.place(x=16, y=5)
    LblSearch.place(x=16, y=45)
    TBxSearch.place(x=160, y=45)
    LblSearchData.place(x=16, y=75)
    TBxSearchData.place(x=160, y=75)

    ChBxAdittCond.place(x=25, y=115)
    LblSearchAd1.place(x=40, y=145)
    TBxSearchAd1.place(x=185, y=145)
    LblSearchDataAd1.place(x=40, y=175)
    TBxSearchDataAd1.place(x=185, y=175)

    ChBxAdittCond2.place(x=25, y=215)
    LblSearchAd2.place(x=40, y=245)
    TBxSearchAd2.place(x=185, y=245)
    LblSearchDataAd2.place(x=40, y=275)
    TBxSearchDataAd2.place(x=185, y=275)

    ChBxAdittCond3.place(x=25, y=315)
    LblSearchAd3.place(x=40, y=345)
    TBxSearchAd3.place(x=185, y=345)
    LblSearchDataAd3.place(x=40, y=375)
    TBxSearchDataAd3.place(x=185, y=375)

    LblColmName2.place(x=375, y=5)
    LblRes1.place(x=375, y=45)
    TbxRes1.place(x=475, y=45)
    LblResTarget1.place(x=375, y=75)
    TbxResTarget1.place(x=550, y=75)

    ChBxRes2.place(x=385, y=115)
    LblRes2.place(x=400, y=145)
    TbxRes2.place(x=500, y=145)
    LblResTarget2.place(x=400, y=175)
    TbxResTarget2.place(x=575, y=175)

    ChBxRes3.place(x=385, y=215)
    LblRes3.place(x=400, y=245)
    TbxRes3.place(x=500, y=245)
    LblResTarget3.place(x=400, y=275)
    TbxResTarget3.place(x=575, y=275)

    ChBxRes4.place(x=385, y=315)
    LblRes4.place(x=400, y=345)
    TbxRes4.place(x=500, y=345)
    LblResTarget4.place(x=400, y=375)
    TbxResTarget4.place(x=575, y=375)
    
    
    
    LblColmNameDup.place_forget()
    LblSearchDup.place_forget()
    TBxSearchDup.place_forget()
    LblResultDup.place_forget()
    TBxResultDup.place_forget()
 
    
def ToggleDuplicateFinderMode():
    global WorkingMode
    WorkingMode = "DuplicateFinder"
    
    AddDataModeBtn.configure(bg=bckgcolor)
    DuplicateFinderModeBtn.configure(bg='#9AEBB7')

    LblColmName1.place_forget()
    LblSearch.place_forget()
    TBxSearch.place_forget()
    LblSearchData.place_forget()
    TBxSearchData.place_forget()

    ChBxAdittCond.place_forget()
    LblSearchAd1.place_forget()
    TBxSearchAd1.place_forget()
    LblSearchDataAd1.place_forget()
    TBxSearchDataAd1.place_forget()

    ChBxAdittCond2.place_forget()
    LblSearchAd2.place_forget()
    TBxSearchAd2.place_forget()
    LblSearchDataAd2.place_forget()
    TBxSearchDataAd2.place_forget()

    ChBxAdittCond3.place_forget()
    LblSearchAd3.place_forget()
    TBxSearchAd3.place_forget()
    LblSearchDataAd3.place_forget()
    TBxSearchDataAd3.place_forget()

    LblColmName2.place_forget()
    LblRes1.place_forget()
    TbxRes1.place_forget()
    LblResTarget1.place_forget()
    TbxResTarget1.place_forget()

    ChBxRes2.place_forget()
    LblRes2.place_forget()
    TbxRes2.place_forget()
    LblResTarget2.place_forget()
    TbxResTarget2.place_forget()

    ChBxRes3.place_forget()
    LblRes3.place_forget()
    TbxRes3.place_forget()
    LblResTarget3.place_forget()
    TbxResTarget3.place_forget()

    ChBxRes4.place_forget()
    LblRes4.place_forget()
    TbxRes4.place_forget()
    LblResTarget4.place_forget()
    TbxResTarget4.place_forget()
    
    
    
    LblColmNameDup.place(x=36, y=25)
    LblSearchDup.place(x=36, y=65)
    TBxSearchDup.place(x=185, y=65)
    LblResultDup.place(x=36, y=95)
    TBxResultDup.place(x=185, y=95)




def CheckSetup():
    # Main data get
    global TBxSearch  
    global TBxSearchData
    ColmSearch = int(TBxSearch.get())
    ColmData = int(TBxSearchData.get())
    
    # Adittional data get
    global ChBxAdittCond
    global TBxSearchAd1
    global TBxSearchDataAd1
    EnableAdittCond1 = bool(AdittCond1.get())
    if EnableAdittCond1 == True:
        ColmAddSearch1 = int(TBxSearchAd1.get())
        ColmAddData1 = int(TBxSearchDataAd1.get())
    else:
        ColmAddSearch1 = 0
        ColmAddData1 = 0
    
    global ChBxAdittCond2
    global TBxSearchAd2
    global TBxSearchDataAd2
    EnableAdittCond2 = bool(AdittCond2.get())
    if EnableAdittCond2 == True:
        ColmAddSearch2 = int(TBxSearchAd2.get())
        ColmAddData2 = int(TBxSearchDataAd2.get())
    else:
        ColmAddSearch2 = 0
        ColmAddData2 = 0
    
    global AdittCond3
    global TBxSearchAd3
    global TBxSearchDataAd3
    EnableAdittCond3 = bool(AdittCond3.get())
    if EnableAdittCond3 == True:
        ColmAddSearch3 = int(TBxSearchAd3.get())
        ColmAddData3 = int(TBxSearchDataAd3.get())
    else:
        ColmAddSearch3 = 0
        ColmAddData3 = 0
    
    
    # Results data get
    global TbxRes1
    global TbxResTarget1
    ColmResult1 = int(TbxRes1.get())
    ColmResultTarget1 = int(TbxResTarget1.get())
    
    global Res2
    global TbxRes2
    global TbxResTarget2
    EnableRes2 = bool(Res2.get())
    if EnableRes2 == True:
        ColmResult2 = int(TbxRes2.get())
        ColmResultTarget2 = int(TbxResTarget2.get())
    else:
        ColmResult2 = 0
        ColmResultTarget2 = 0
    
    global Res3
    global TbxRes3
    global TbxResTarget3
    EnableRes3 = bool(Res3.get())
    if EnableRes3 == True:
        ColmResult3 = int(TbxRes3.get())
        ColmResultTarget3 = int(TbxResTarget3.get())
    else:
        ColmResult3 = 0
        ColmResultTarget3 = 0
    
    global Res4
    global TbxRes4
    global TbxResTarget4
    EnableRes4 = bool(Res4.get())
    if EnableRes4 == True:
        ColmResult4 = int(TbxRes4.get())
        ColmResultTarget4 = int(TbxResTarget4.get())
    else:
        ColmResult4 = 0
        ColmResultTarget4 = 0


def resource_path(relative_path):    
    try:       
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


if __name__ == '__main__':
    multiprocessing.freeze_support()
    main()
