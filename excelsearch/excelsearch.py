from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

import os, sys, re

from pathlib import Path
from threading import Thread
from multiprocessing import Process, Queue, current_process

import openpyxl


from tkintertable.Tables import TableCanvas
from tkintertable.TableModels import TableModel

global root
global scrnwparam
global scrnhparam
scrnwparam = 300
scrnhparam = 300


def main():
    global root
    global bckgcolor
    bckgcolor = '#f7f7f7'

    root = Tk()
    root.resizable(False, False)
    
    datafile = "compare.ico"
    if not hasattr(sys, "frozen"):
        datafile = os.path.join(os.path.dirname(__file__), datafile)
    else:
        datafile = os.path.join(sys.prefix, datafile)
    root.iconbitmap(default=resource_path(datafile))
    
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    #root.geometry('755x500+{}+{}'.format(scrnw, scrnh))
    root.geometry('715x500+{}+{}'.format(scrnw, scrnh))
    
    app = GUI(root)
    root.mainloop()


class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background=bckgcolor)   
        self.parent = parent
        self.parent.title("Add Data v5.1")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
        

    
    def initUI(self):
    
        global AdittCond1
        global AdittCond2
        global AdittCond3
        
        global Res2
        global Res3
        global Res4

        AdittCond1 = IntVar()
        AdittCond1.set(0)
        AdittCond2 = IntVar()
        AdittCond2.set(0)
        AdittCond3 = IntVar()
        AdittCond3.set(0)
        
        Res2 = IntVar()
        Res2.set(0)
        Res3 = IntVar()
        Res3.set(0)
        Res4 = IntVar()
        Res4.set(0)
    
        global MaxRowsSearch
        global MaxRowsData
        global PreviewRows
        global PreviewColumns
        PreviewColumns = 50
    
    
        global LblColmName1
        LblColmName1 = Label(text="Настройка столбцов поиска", background=bckgcolor, font=("Arial", 14))
        LblColmName1.place(x=16, y=5)
        
        global LblSearch
        LblSearch = Label(text="Искомых значений:", background=bckgcolor, font=("Arial", 11))
        LblSearch.place(x=16, y=45)
        
        global TBxSearch
        TBxSearch = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TBxSearch.place(x=160, y=45)
        #TBxSearch.insert(0, "1")
        
        global LblSearchData
        LblSearchData = Label(text="Диапазона поиска:", background=bckgcolor, font=("Arial", 11))
        LblSearchData.place(x=16, y=75)
        
        global TBxSearchData
        TBxSearchData = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TBxSearchData.place(x=160, y=75)
        #TBxSearchData.insert(0, "9")
        
        
        
        global ChBxAdittCond
        ChBxAdittCond = Checkbutton(text="Дополнительное условие сопоставления 1", background=bckgcolor, variable=AdittCond1, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)
        ChBxAdittCond.place(x=25, y=115)
        
        
        global LblSearchAd1
        LblSearchAd1 = Label(text="Искомых значений:", background=bckgcolor, font=("Arial", 11))
        LblSearchAd1.place(x=40, y=145)
        
        global TBxSearchAd1
        TBxSearchAd1 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TBxSearchAd1.place(x=185, y=145)
        #TBxSearchAd1.insert(0,"2")
        
        global LblSearchDataAd1
        LblSearchDataAd1 = Label(text="Диапазона поиска:", background=bckgcolor, font=("Arial", 11))
        LblSearchDataAd1.place(x=40, y=175)
        
        global TBxSearchDataAd1
        TBxSearchDataAd1 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TBxSearchDataAd1.place(x=185, y=175)
        #TBxSearchDataAd1.insert(0,"10")
        
        global ChBxAdittCond2
        ChBxAdittCond2 = Checkbutton(text="Дополнительное условие сопоставления 2", background=bckgcolor, variable=AdittCond2, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)
        ChBxAdittCond2.place(x=25, y=215)
        
        
        global LblSearchAd2
        LblSearchAd2 = Label(text="Искомых значений:", background=bckgcolor, font=("Arial", 11))
        LblSearchAd2.place(x=40, y=245)
        
        global TBxSearchAd2
        TBxSearchAd2 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TBxSearchAd2.place(x=185, y=245)
        #TBxSearchAd2.insert(0,"3")
        
        global LblSearchDataAd2
        LblSearchDataAd2 = Label(text="Диапазона поиска:", background=bckgcolor, font=("Arial", 11))
        LblSearchDataAd2.place(x=40, y=275)
        
        global TBxSearchDataAd2
        TBxSearchDataAd2 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TBxSearchDataAd2.place(x=185, y=275)
        #TBxSearchDataAd2.insert(0,"11")
        
        global ChBxAdittCond3
        ChBxAdittCond3 = Checkbutton(text="Дополнительное условие сопоставления 3", background=bckgcolor, variable=AdittCond3, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)
        ChBxAdittCond3.place(x=25, y=315)
        
        
        global LblSearchAd3
        LblSearchAd3 = Label(text="Искомых значений:", background=bckgcolor, font=("Arial", 11))
        LblSearchAd3.place(x=40, y=345)
        
        global TBxSearchAd3
        TBxSearchAd3 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TBxSearchAd3.place(x=185, y=345)
        #TBxSearchAd3.insert(0,"4")
        
        global LblSearchDataAd3
        LblSearchDataAd3 = Label(text="Диапазона поиска:", background=bckgcolor, font=("Arial", 11))
        LblSearchDataAd3.place(x=40, y=375)
        
        global TBxSearchDataAd3
        TBxSearchDataAd3 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TBxSearchDataAd3.place(x=185, y=375)
        #TBxSearchDataAd3.insert(0,"12")
        
        
        Offset = -25
        
        global LblColmName2
        LblColmName2 = Label(text="Настройка столбцов результатов", background=bckgcolor, font=("Arial", 14))
        LblColmName2.place(x=420+Offset, y=5)
        
        global LblRes1
        LblRes1 = Label(text="Результат 1 :", background=bckgcolor, font=("Arial", 11))
        LblRes1.place(x=420+Offset, y=45)
        
        global TbxRes1
        TbxRes1 = Entry(textvariable = "9", fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TbxRes1.place(x=520+Offset, y=45)
        #TbxRes1.insert(0,"9")
        
        global LblResTarget1
        LblResTarget1 = Label(text="Скопировать в столбец:", background=bckgcolor, font=("Arial", 11))
        LblResTarget1.place(x=420+Offset, y=75)
        
        global TbxResTarget1
        TbxResTarget1 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TbxResTarget1.place(x=595+Offset, y=75)
        #TbxResTarget1.insert(0,"5")
        
        global ChBxRes2
        ChBxRes2 = Checkbutton(text="Включить результат 2", background=bckgcolor, variable=Res2, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)
        ChBxRes2.place(x=430+Offset, y=115)
        
        global LblRes2
        LblRes2 = Label(text="Результат 2 :", background=bckgcolor, font=("Arial", 11))
        LblRes2.place(x=445+Offset, y=145)
        
        global TbxRes2
        TbxRes2 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TbxRes2.place(x=545+Offset, y=145)
        
        global LblResTarget2
        LblResTarget2 = Label(text="Скопировать в столбец:", background=bckgcolor, font=("Arial", 11))
        LblResTarget2.place(x=445+Offset, y=175)
        
        global TbxResTarget2
        TbxResTarget2 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TbxResTarget2.place(x=620+Offset, y=175)
        
        global ChBxRes3
        ChBxRes3 = Checkbutton(text="Включить результат 3", background=bckgcolor, variable=Res3, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)
        ChBxRes3.place(x=430+Offset, y=215)
        
        global LblRes3
        LblRes3 = Label(text="Результат 3 :", background=bckgcolor, font=("Arial", 11))
        LblRes3.place(x=445+Offset, y=245)
        
        global TbxRes3
        TbxRes3 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TbxRes3.place(x=545+Offset, y=245)
        
        global LblResTarget3
        LblResTarget3 = Label(text="Скопировать в столбец:", background=bckgcolor, font=("Arial", 11))
        LblResTarget3.place(x=445+Offset, y=275)
        
        global TbxResTarget3
        TbxResTarget3 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TbxResTarget3.place(x=620+Offset, y=275)
        
        global ChBxRes4
        ChBxRes4 = Checkbutton(text="Включить результат 4", background=bckgcolor, variable=Res4, onvalue=1, offvalue=0, font=("Arial", 11), command=InterfaceTrigger)
        ChBxRes4.place(x=430+Offset, y=315)
        
        global LblRes4
        LblRes4 = Label(text="Результат 4 :", background=bckgcolor, font=("Arial", 11))
        LblRes4.place(x=445+Offset, y=345)
        
        global TbxRes4
        TbxRes4 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TbxRes4.place(x=545+Offset, y=345)
        
        global LblResTarget4
        LblResTarget4 = Label(text="Скопировать в столбец:", background=bckgcolor, font=("Arial", 11))
        LblResTarget4.place(x=445+Offset, y=375)
        
        global TbxResTarget4
        TbxResTarget4 = Entry(fg="black", bg=bckgcolor, width=3, font=("Arial", 11))
        TbxResTarget4.place(x=620+Offset, y=375)
        
        
        
        #global LblSearchRange
        #LblSearchRange = Label(text="Диапазон прогонки искомых:", background=bckgcolor, font=("Arial", 11))
        #LblSearchRange.place(x=470, y=420)
        
        #global TbxSearchRange
        #TbxSearchRange = Entry(fg="black", bg=bckgcolor, width=6, font=("Arial", 11))
        #TbxSearchRange.place(x=675, y=420)
        
        #global LblSearchDataRange
        #LblSearchDataRange = Label(text="Диапазон прогонки поиска:", background=bckgcolor, font=("Arial", 11))
        #LblSearchDataRange.place(x=470, y=450)
        
        #global TbxSearchDataRange
        #TbxSearchDataRange = Entry(fg="black", bg=bckgcolor, width=6, font=("Arial", 11))
        #TbxSearchDataRange.place(x=675, y=450)
        
        
        
        
        global InputFileLbl
        InputFileLbl = Label(text="Выберите файл таблицы:", background=bckgcolor, font=("Arial", 14))
        InputFileLbl.place(x=16, y=420)
        
        global InputFileEntry
        InputFileEntry = Entry(fg="black", bg=bckgcolor, width=27, font=("Arial", 13))
        InputFileEntry.place(x=20, y=450)
        
        global InputFileBtn
        InputFileBtn = Button(text='⮌', command=SelectFile, font=("Arial", 13))
        InputFileBtn.place(x=275, y=450, height=23)
        
        
        global OpenPrevTableBtn
        OpenPrevTableBtn = Button(text='Просмотр', command=OpenPreviewWindow, font=("Arial", 13))
        OpenPrevTableBtn.place(x=320, y=430, width=105, height=43)
        OpenPrevTableBtn.configure(state = DISABLED)
        
        
        global RunBtn
        RunBtn = Button(text='Запуск', command=Run, font=("Arial", 13, "bold"))
        #RunBtn.place(x=640, y=345, width=90, height=50)
        RunBtn.place(x=490, y=430, width=200, height=42)
        
        
        InterfaceTrigger()
        
        #ChBxAdittCond.configure(state = DISABLED)
        #ChBxAdittCond2.configure(state = DISABLED)
        #ChBxAdittCond3.configure(state = DISABLED)
        #ChBxRes2.configure(state = DISABLED)
        #ChBxRes3.configure(state = DISABLED)
        #ChBxRes4.configure(state = DISABLED)

def OpenPreviewWindow():
    PreviewWindow(Path(InputFile).name)

def PreviewWindow( WindowName ):
    #PreviewWindow = Toplevel(root)
    #PreviewWindow.title("New Window 1")
    #PreviewWindow.geometry("300x300")
    
    #table = TableCanvas(t_frame, cellwidth=60, thefont=('Arial',12),rowheight=18, rowheaderwidth=30,
    #                    rowselectedcolor='yellow', editable=True)
    global MaxRowsSearch
    global MaxRowsData
    global PreviewRows
    global PreviewColumns

    PreviewWindow=Tk()
    PreviewWindow.title(WindowName)
    #PreviewWindow.title(Path(InputFile).name)
    PreviewWindow.geometry("1200x550")
    
    t_frame=Frame(PreviewWindow)
    t_frame.pack(fill='both', expand=True)
    
    
    model = TableModel()
    table = TableCanvas(t_frame, model=model, editable=False)
    
    data = table.model.data
    cols = table.model.columnNames #get the current columns
    
    #CheckMaxRows()
    
    book = openpyxl.load_workbook(InputFile)
    sheet = book.active
    
    for i in range(PreviewColumns):
        table.model.addColumn()
    
    for k in range(PreviewRows):
        table.model.addRow()
    
    for r in range(PreviewRows):
        for c in range(PreviewColumns):
            if sheet.cell(row=r+1, column=c+1).value != None:
                table.model.setValueAt(str(sheet.cell(row=r+1, column=c+1).value),r,c)
    
    MaxRowsSearch = PreviewRows
    MaxRowsData = PreviewRows
    
    book.close()
    table.show()


def CheckMaxRows():
    global PreviewRows
    global PreviewColumns

    book = openpyxl.load_workbook(InputFile)
    sheet = book.active
    
    EmptyCellCounter = 0
    EmptyRowCounter = 0
    
    for t in range(100000):
        EmptyCellCounter = 0
        for y in range(PreviewColumns):
            if sheet.cell(row=t+1, column=y+1).value is None:
                EmptyCellCounter = EmptyCellCounter + 1
                #print('t={0}, EmptyCellCounter: {1}'.format(t, EmptyCellCounter))
        if EmptyCellCounter == PreviewColumns:
            EmptyRowCounter = EmptyRowCounter + 1
            #print('EmptyCellCounter == PreviewColumns')
            #print('EmptyRowCounter={0}'.format(EmptyRowCounter))
        if EmptyRowCounter >= 10:
            PreviewRows = t - 9
            #print(t)
            break
    book.close()
    


def BtnCmd():
    print('btn pressed')


def SelectFile():
    global InputFile

    InputFile = ""
    InputFile = filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx")])
    if InputFile:
        InputFileEntry.configure(state = NORMAL)
        InputFileEntry.delete(0,END)
        InputFileEntry.insert(0,str(Path(InputFile).name))
        InputFileEntry.configure(state = DISABLED)
        
        OpenPrevTableBtn.configure(state = NORMAL)
        
        CheckMaxRows()
        PreviewWindow(Path(InputFile).name)
        print('IDC: InputFile : {0}'.format(InputFile))
    else:
        print('IDC: InputFile not selected')





def Run():
    #thread = Thread(target=MainThread)
    #thread.start()
    #thread = ""

#def MainThread():
    global InputFile
    RunBtn.configure(state = DISABLED)
    
    book = openpyxl.load_workbook(InputFile)
    sheet = book.active

    #MaxRowsSearch = int(TbxSearchRange.get())
    #MaxRowsData = int(TbxSearchDataRange.get())
    
    
    # Main data
    ColmSearch = int(TBxSearch.get())
    ColmData = int(TBxSearchData.get())
    
    ColmResult1 = int(TbxRes1.get())
    ColmResultTarget1 = int(TbxResTarget1.get())
    
    ArrSearch = []
    ArrData = []
    
    
    # Main conditions array appending
    for i in range(MaxRowsSearch):
        ArrSearch.append(sheet.cell(row=i+1, column=ColmSearch).value)   
    for k in range(MaxRowsData):
        ArrData.append(sheet.cell(row=k+1, column=ColmData).value)
    
    
    # Additional conditions arrays appending
    if AdittCond1.get() == 1:
        ColmAddSearch1 = int(TBxSearchAd1.get())
        ColmAddData1 = int(TBxSearchDataAd1.get())
        ArrAddSearch1 = []
        ArrAddData1 = []
        for i in range(MaxRowsSearch):
            ArrAddSearch1.append(sheet.cell(row=i+1, column=ColmAddSearch1).value)
            ArrAddData1.append(sheet.cell(row=i+1, column=ColmAddData1).value)
            
    if AdittCond2.get() == 1:
        ColmAddSearch2 = int(TBxSearchAd2.get())
        ColmAddData2 = int(TBxSearchDataAd2.get())
        ArrAddSearch2 = []
        ArrAddData2 = []
        for i in range(MaxRowsSearch):
            ArrAddSearch2.append(sheet.cell(row=i+1, column=ColmAddSearch2).value)
            ArrAddData2.append(sheet.cell(row=i+1, column=ColmAddData2).value)
            
    if AdittCond3.get() == 1:
        ColmAddSearch3 = int(TBxSearchAd3.get())
        ColmAddData3 = int(TBxSearchDataAd3.get())
        ArrAddSearch3 = []
        ArrAddData3 = []
        for i in range(MaxRowsSearch):
            ArrAddSearch3.append(sheet.cell(row=i+1, column=ColmAddSearch3).value)
            ArrAddData3.append(sheet.cell(row=i+1, column=ColmAddData3).value)
    
    
    # Additional results
    if Res2.get() == 1:
        ColmResult2 = int(TbxRes2.get())
        ColmResultTarget2 = int(TbxResTarget2.get())
    if Res3.get() == 1:
        ColmResult3 = int(TbxRes3.get())
        ColmResultTarget3 = int(TbxResTarget3.get())
    if Res4.get() == 1:
        ColmResult4 = int(TbxRes4.get())
        ColmResultTarget4 = int(TbxResTarget4.get())



    # Starting search
    SearchedRowsCount = 0
    FindedRowsCount = 0
    for m in range(MaxRowsSearch):
        SearchedRowsCount = SearchedRowsCount + 1
        if ArrSearch[m]!=None:
            for n in range(MaxRowsData):
                Finded = False
                if ArrSearch[m] == ArrData[n]:
                    Finded = True
                if AdittCond1.get() == 1 and ArrAddSearch1[m] != ArrAddData1[n]:
                    Finded = False
                if AdittCond2.get() == 1 and ArrAddSearch2[m] != ArrAddData2[n]:
                    Finded = False
                if AdittCond3.get() == 1 and ArrAddSearch3[m] != ArrAddData3[n]:
                    Finded = False
                    
                if Finded == True:
                    sheet.cell(row=m+1, column=ColmResultTarget1).value = sheet.cell(row=n+1, column=ColmResult1).value
                    if Res2.get() == 1:
                        sheet.cell(row=m+1, column=ColmResultTarget2).value = sheet.cell(row=n+1, column=ColmResult2).value
                    if Res3.get() == 1:
                        sheet.cell(row=m+1, column=ColmResultTarget3).value = sheet.cell(row=n+1, column=ColmResult3).value
                    if Res4.get() == 1:
                        sheet.cell(row=m+1, column=ColmResultTarget4).value = sheet.cell(row=n+1, column=ColmResult4).value
                    print('m={0}, finded: {1} = {2}'.format(m, ArrSearch[m], ArrData[n]))
                    FindedRowsCount = FindedRowsCount + 1
                    break
                
    book.save(InputFile)
    
    
    RunBtn.configure(state = NORMAL)
    
    msgbxlbl = ['Обработка завершена !', 'Обработано: {0}'.format(SearchedRowsCount), 'Найдено: {0}'.format(FindedRowsCount)]
    messagebox.showinfo("", "\n".join(msgbxlbl))
    PreviewWindow('РЕЗУЛЬТ ОБРАБОТКИ :  {0}'.format(Path(InputFile).name))


def InterfaceTrigger():

    if AdittCond1.get() == 1:
        LblSearchAd1.configure(state = NORMAL)
        LblSearchDataAd1.configure(state = NORMAL)
        TBxSearchAd1.configure(state = NORMAL)
        TBxSearchDataAd1.configure(state = NORMAL)
    else:
        LblSearchAd1.configure(state = DISABLED)
        LblSearchDataAd1.configure(state = DISABLED)
        TBxSearchAd1.configure(state = DISABLED)
        TBxSearchDataAd1.configure(state = DISABLED)


    if AdittCond2.get() == 1:
        LblSearchAd2.configure(state = NORMAL)
        LblSearchDataAd2.configure(state = NORMAL)
        TBxSearchAd2.configure(state = NORMAL)
        TBxSearchDataAd2.configure(state = NORMAL)
    else:
        LblSearchAd2.configure(state = DISABLED)
        LblSearchDataAd2.configure(state = DISABLED)
        TBxSearchAd2.configure(state = DISABLED)
        TBxSearchDataAd2.configure(state = DISABLED)
    
    
    if AdittCond3.get() == 1:
        LblSearchAd3.configure(state = NORMAL)
        LblSearchDataAd3.configure(state = NORMAL)
        TBxSearchAd3.configure(state = NORMAL)
        TBxSearchDataAd3.configure(state = NORMAL)
    else:
        LblSearchAd3.configure(state = DISABLED)
        LblSearchDataAd3.configure(state = DISABLED)
        TBxSearchAd3.configure(state = DISABLED)
        TBxSearchDataAd3.configure(state = DISABLED)


    if Res2.get() == 1:
        LblRes2.configure(state = NORMAL)
        LblResTarget2.configure(state = NORMAL)
        TbxRes2.configure(state = NORMAL)
        TbxResTarget2.configure(state = NORMAL)
    else:
        LblRes2.configure(state = DISABLED)
        LblResTarget2.configure(state = DISABLED)
        TbxRes2.configure(state = DISABLED)
        TbxResTarget2.configure(state = DISABLED)


    if Res3.get() == 1:
        LblRes3.configure(state = NORMAL)
        LblResTarget3.configure(state = NORMAL)
        TbxRes3.configure(state = NORMAL)
        TbxResTarget3.configure(state = NORMAL)
    else:
        LblRes3.configure(state = DISABLED)
        LblResTarget3.configure(state = DISABLED)
        TbxRes3.configure(state = DISABLED)
        TbxResTarget3.configure(state = DISABLED)
    
    
    if Res4.get() == 1:
        LblRes4.configure(state = NORMAL)
        LblResTarget4.configure(state = NORMAL)
        TbxRes4.configure(state = NORMAL)
        TbxResTarget4.configure(state = NORMAL)
    else:
        LblRes4.configure(state = DISABLED)
        LblResTarget4.configure(state = DISABLED)
        TbxRes4.configure(state = DISABLED)
        TbxResTarget4.configure(state = DISABLED)


def resource_path(relative_path):    
    try:       
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)










def RunTest():
    global InputFile
    dataq = Queue()
    
    subprocess = Process(target=SearchProcess, args=(dataq, InputFile))
    subprocess.start()


def SearchProcess(dataq, DivideInputFile):
    
    ProcessWindow=Tk()
    ProcessWindow.title("Process")
    ProcessWindow.geometry("300x200")
    
    app = ProcessGUI(ProcessWindow)
    ProcessWindow.mainloop()
    

    #dataq.put([DivideInputFile])
    print('Process started')
    
    
    
class ProcessGUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background='white')   
        self.parent = parent
        self.parent.title("Add Data v5.1")
        self.pack(fill=BOTH, expand=1)
        self.ProcessInitUI()
        

    
    def ProcessInitUI(self):
        global LblColmName1
        LblColmName1 = Label(text="Настройка столбцов поиска", background='white', font=("Arial", 14))
        LblColmName1.place(x=16, y=5)







if __name__ == '__main__':
    main()
