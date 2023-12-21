from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

import requests, os, sys, csv, re
from pathlib import Path

import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import xml.etree.ElementTree as ET
import barcode
from barcode.writer import ImageWriter
from PIL import Image
from PIL import ImageDraw 

global root
global scrnwparam
global scrnhparam
scrnwparam = 240
scrnhparam = 195

global QuantityIsInThousands


def main():
    global root

    root = Tk()
    root.resizable(False, False)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('370x100+{}+{}'.format(scrnw, scrnh))
        
    app = GUI(root)
    root.mainloop()
    
    

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("Генерация штрихкодов")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):
        IsInputSel = False

        global FindInternalCodes
        FindInternalCodes = False
        

        global UniversalBtn
        UniversalBtn = Button(text='Выбрать файл с кодами', command=SelectFile, font=("Arial", 14))
        UniversalBtn.place(x=30, y=20, width=310, height=50)



def SelectFile():
    extension = '.xlsx .xls'
    extname = extension.upper()
    InputFile = filedialog.askopenfilename(title='Выберите Счет', filetypes=((extname, extension),))
    
    if InputFile:
        print('Файл выбран :' + InputFile)
        Generate(InputFile)
    else:
        print('Файл не выбран')



def Generate(InputFile):
    book = openpyxl.load_workbook(InputFile)
    sheet = book.active
    CodeColumn = AdressInRow("Код",1,sheet)
    image_writer = ImageWriter()
    for myrow in range(2,sheet.max_row+1):
        ean = barcode.get('Code128', sheet.cell(myrow,CodeColumn).value, writer=image_writer)
        filename = ean.save(sheet.cell(myrow,CodeColumn).value,{"module_width":0.35, "module_height":14, "font_size": 28, "text_distance": 15, "quiet_zone": 3})
        print(filename)
        if(sheet.cell(myrow,CodeColumn).value[-1]=="0"):
            #Read the two images
            image1 = Image.open(sheet.cell(myrow,CodeColumn).value+'.png')
            #image1.show()
            image2 = Image.open('DownArrow.png')
            #image2.show()
            #resize, first image
            
            image1_size = image1.size
            image2_size = image2.size
            image2 = image2.resize((200, image1_size[1]))
            new_image = Image.new('RGB',(image1_size[0]+200, image1_size[1]), (250,250,250))
            new_image.paste(image1,(0,0))
            new_image.paste(image2,(image1_size[0],0))
            w, h = new_image.size
            new_image=new_image.crop((0, 0, w, h-30))
            #new_image.show()
            draw = ImageDraw.Draw(new_image, "RGBA")
            draw.rounded_rectangle(((2, 2), (int(new_image.size[0])-2, int(new_image.size[1])-2)), fill=None,outline = (0,0,0,225),width=8, radius = 20)
            new_image.save(sheet.cell(myrow,CodeColumn).value+".png","PNG")
        else:
            image1 = Image.open(sheet.cell(myrow,CodeColumn).value+'.png')
            w, h = image1.size
            image1=image1.crop((0, 0, w, h-30))
            draw = ImageDraw.Draw(image1, "RGBA")
            draw.rounded_rectangle(((2, 2), (int(image1.size[0])-2, int(image1.size[1])-2)), fill=None,outline = (0,0,0,225),width=8,radius = 20)
            image1.save(sheet.cell(myrow,CodeColumn).value+".png","PNG")
    images = [
    Image.open(sheet.cell(myrow,CodeColumn).value+".png")
    for myrow in range(2,sheet.max_row+1)
    ]

    pdf_path = "ФайлПДФ.pdf"
        
    images[0].save(
        pdf_path, "PDF" ,resolution=100.0, save_all=True, append_images=images[1:]
    )
    



def AdressInRow(things,row,sheet):
    for y in range(1,sheet.max_column):
        if str(sheet.cell(row, column=y).value) in things:
            return y
    return -1




if __name__ == '__main__':
    main()
