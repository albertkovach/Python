from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

from pathlib import Path
from openpyxl import Workbook, load_workbook
import os, sys
from datetime import datetime, date, timedelta
from dateutil import parser
from dateutil.relativedelta import relativedelta
from difflib import SequenceMatcher

global root
global scrnwparam
global scrnhparam
scrnwparam = 138
scrnhparam = 110

def main():
    global root

    root = Tk()
    root.resizable(False, False)

    datafile = "icon.ico"
    if not hasattr(sys, "frozen"):
        datafile = os.path.join(os.path.dirname(__file__), datafile)
    else:
        datafile = os.path.join(sys.prefix, datafile)
    root.iconbitmap(default=resource_path(datafile))

    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('275x110+{}+{}'.format(scrnw, scrnh))
        
    app = GUI(root)
    root.mainloop()

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("Resolver")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):
        global InputFileLbl
        InputFileLbl = Label(text="Выберите табель:", background="white", font=("Arial", 10))
        InputFileLbl.place(x=16, y=10)
        
        global InputFileEntry
        InputFileEntry = Entry(fg="black", bg="white", width=30)
        InputFileEntry.place(x=20, y=32)
        InputFileEntry.configure(state = DISABLED)
        
        global InputFileBtn
        InputFileBtn = Button(text='Выбор', command=InputFileChoose)
        InputFileBtn.place(x=210, y=31, height=20)
        
        global StartBtn
        StartBtn = Button(text='Магия !', command=Start)
        StartBtn.place(x=20, y=65, width=75, height=25)
        StartBtn.configure(state = DISABLED)



def InputFileChoose():
    global InputFile
    global IsInputSel

    InputFile = filedialog.askopenfilename(title='Выберите табель:', filetypes=(('Таблица Excel', 'xlsx'),))
    if InputFile:
        InputFileEntry.configure(state = NORMAL)
        InputFileEntry.delete(0,END)
        InputFileEntry.insert(0,str(Path(InputFile).name))
        InputFileEntry.configure(state = DISABLED)
        print('IFC: InputFile :', InputFile)
        
        IsInputSel = True
        StartBtn.configure(state = NORMAL)
    else:
        print('IFC: InputFile not selected')



def Start():
    global InputFile
    global IsInputSel
    
    if IsInputSel:
        inputtable = load_workbook(InputFile, data_only=True)
        inputsheet = inputtable.active
        
        department = str(inputsheet['B8'].value)
        print('Department: {0}'.format(department))
        if department == 'None' or DepartStrConvert(department) == 'error':
            msgbxlbl = ['Файл не подходит !', '(Или поправь исходный код)']
            messagebox.showerror("ОШИБКА !!!!", "\n".join(msgbxlbl))
            StartBtn.configure(state = DISABLED)
        else:
            department = DepartStrConvert(department)
            
            periodstart = inputsheet['AY12'].value
            periodstop = inputsheet['BC12'].value
            print('Period: {0} - {1}'.format(periodstart, periodstop))

            outputtable = Workbook()
            outputsheet = outputtable.active
            outputsheet.title = "Табель"
            outputsheet['A1'] = 'Табель учета рабочего времени {0}: c {1} по {2}'.format(department, periodstart, periodstop)
            outputsheet['A2'] = 'Сотрудник'


            datestart = datetime.strptime(periodstart, "%d.%m.%Y")
            datestop = datetime.strptime(periodstop, "%d.%m.%Y")
            datedelta = datestop - datestart
            
            if int(datestart.strftime("%d"))<=15:
                if int(datestop.strftime("%d"))>15:
                    msgbxlbl = ['Диапазон цепляет дни после 15-го числа !', '{0} - {1}'.format(periodstart, periodstop)]
                    messagebox.showerror("ОШИБКА !!!!", "\n".join(msgbxlbl))

            # Шапка с датами
            for d in range(datedelta.days+1):
                newdate = datestart + relativedelta(days=d)
                outputsheet.cell(row=2, column=2+d).value = DateStrConvert(newdate)
            outputsheet.cell(row=2, column=2+d+1).value = "Сумма"
            
            # Основные данные
            staffoutrow = 2
            for staffinputrow in range(300):
                cellvalue = str(inputsheet.cell(row=5+staffinputrow, column=3).value)
                if cellvalue != 'None':
                    similarity = Similar(cellvalue, 'Фамилия, инициалы, должность (специальность, профессия)')
                    if similarity < 0.9 and len(cellvalue)>2:
                        # Значит, это ячейка с сотрудником
                        staffoutrow = staffoutrow+1
                        fullname = '{0} {1} {2}'.format(cellvalue.split()[0], cellvalue.split()[1], cellvalue.split()[2])
                        outputsheet.cell(row=staffoutrow, column=1).value = fullname
                        print('{0}, {1}'.format(staffinputrow, fullname))
                        
                        # Добавление информации о отработанном времени
                        ci = 1
                        if int(datestart.strftime("%d"))<=15:
                            timerowadd = 0
                        else:
                            timerowadd = 2
                        for c in range(31):
                            workstatuscellvalue = str(inputsheet.cell(row=5+staffinputrow+timerowadd, column=9+c).value)
                            if workstatuscellvalue != 'None' or c == 30:
                                worktimecellvalue = str(inputsheet.cell(row=6+staffinputrow+timerowadd, column=9+c).value)
                                if worktimecellvalue != 'Х':
                                    ci = ci+1
                                    if worktimecellvalue == 'None':
                                        worktimecellvalue = '0'
                                    if '/' in worktimecellvalue:
                                        print("-- Неполный день!")
                                        worktimecellvalue = worktimecellvalue.split("/")[0]
                                    outputsheet.cell(row=staffoutrow, column=ci).value = '{0}:00:00'.format(worktimecellvalue)
                                    #outputsheet.cell(row=staffoutrow, column=ci).value = worktimecellvalue
                

            outputfilepath = Path(Path(InputFile).parent, 'Табель {0} {1}-{2}.xlsx'.format(department, datestart.strftime("%d"), periodstop))
            print('outputfilepath:', outputfilepath)
            outputtable.save(outputfilepath)
            os.startfile(outputfilepath)



def DateStrConvert(date):
    match int(date.strftime("%w")):
        case 0:
            datestr = 'вс. {0}'.format(date.strftime("%d"))
        case 1:
            datestr = 'пн. {0}'.format(date.strftime("%d"))
        case 2:
            datestr = 'вт. {0}'.format(date.strftime("%d"))
        case 3:
            datestr = 'ср. {0}'.format(date.strftime("%d"))
        case 4:
            datestr = 'чт. {0}'.format(date.strftime("%d"))
        case 5:
            datestr = 'пт. {0}'.format(date.strftime("%d"))
        case 6:
            datestr = 'сб. {0}'.format(date.strftime("%d"))
        case _:
            datestr = 'ошибка!'
    return datestr



def DepartStrConvert(department):
    dep1 = Similar(department, 'Лыковская осн.')
    dep2 = Similar(department, 'Отдел цифровой обработки документов')
    dep3 = Similar(department, 'Архив Падиково')
    dep4 = Similar(department, 'ЦОД Падиково')
    
    #print('{0} {1} {2} {3}'.format(dep1, dep2, dep3, dep4))
    if dep1 > 0.9:
        department = 'Лыковская'
    elif dep2 > 0.9:
        department = 'ЦОД'
    elif dep3 > 0.9:
        department = 'Падиково'
    elif dep4 > 0.9:
        department = 'ЦОД Падиково'
    else:
        department = 'error'
    return department



def Similar(a, b):
    return SequenceMatcher(None, a, b).ratio()



def resource_path(relative_path):    
    try:       
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)



if __name__ == '__main__':
    main()
