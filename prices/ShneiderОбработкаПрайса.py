from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

import requests, os, sys, csv, re
from pathlib import Path

import xlrd
import openpyxl
import pyxlsb
from openpyxl import Workbook
from openpyxl.styles import Font
import xml.etree.ElementTree as ET

global root
global scrnwparam
global scrnhparam
scrnwparam = 240
scrnhparam = 195

import pandas as pd

global QuantityIsInThousands


def main():
    global root

    root = Tk()
    root.resizable(False, False)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('370x160+{}+{}'.format(scrnw, scrnh))
        
    app = GUI(root)
    root.mainloop()
    

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("Переделка прайса Schneider")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):
        #global IsInputSel
        IsInputSel = False

        global FindInternalCodes
        FindInternalCodes = False
        
        #global InputFile
        #global extension

        
        global UniversalBtn
        UniversalBtn = Button(text='Нажмите', command=RunUniversal, font=("Arial", 14))
        UniversalBtn.place(x=30, y=30, width=310, height=85)


        try:
            InputFile = sys.argv[1]
            #extension = Path(InputFile).suffix
            print(InputFile)
            print("Файл выбран "+InputFile) 
            Universal(InputFile)
        except:
            print("Файл не выбран") 




def SelectFile(extension):
    #global InputFile
    #global IsInputSel
    extname = extension.upper()

    InputFile = filedialog.askopenfilename(title='Выберите УПД', filetypes=((extname, extension),))
    
    if InputFile:
        print('Файл выбран :' + InputFile)
        
        #print(str(IsInputSel))
        return InputFile
    else:
        print('Файл не выбран')
        return False




def IsInCell(thing,row,column,sheet):
    if str(sheet.cell(row, column).value) == thing:
        return True
    else:
        return False


def AdressInRow(things,row,sheet):
    for y in range(1,sheet.max_column):
        if str(sheet.cell(row, column=y).value) in things:
            return y
        if things[0] in str(sheet.cell(row, column=y).value):
            return y
    return -1


def FindAllInstancesOnSheet(thing,sheet):
    locations = [-1]
    for x in range(1,sheet.max_row):
        for y in range(1,sheet.max_column):
            if thing in str(sheet.cell(row=x, column=y).value): 
                locations.append([x,y])
                return locations
    return locations


def NextNotEmptyInRowAdress(row,startcolumn,sheet):
    for y in range(startcolumn+1,sheet.max_column):
        if (sheet.cell(row, column=y).value)!=None:
            return y
    return -1


def SaveAsXLSX(InputFile):
    encod = False
    book_xls = xlrd.open_workbook(InputFile)
    book_xlsx = Workbook()
    if str(format(book_xls.encoding)) == "iso-8859-1":
        print(str(format(book_xls.encoding))+" Декодирую и сохраняю как xlsx")
        encod = True
    else:
        print(" Сохраняю без декодирования как xlsx")
        encod = False
    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
    if sheet_index == 0:
        sheet_xlsx = book_xlsx.active
        sheet_xlsx.title = sheet_name
    else:
        sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

    for row in range(0, sheet_xls.nrows):
        for col in range(0, sheet_xls.ncols):
            if encod == True:
                sheet_xlsx.cell(row = row+1 , column = col+1).value = Decode(str(sheet_xls.cell_value(row, col)))
            else:
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
    savepath = Path(Path(InputFile).parent, str(Path(InputFile).stem+" Перекодированный.xlsx"))
    #print("Сохранил в " + savepath)
    book_xlsx.save(savepath)
    return savepath



def RunUniversal():
    InputFile = SelectFile('.xlsb .xlsx .xls')
    #print(InputFile)    
    Universal(InputFile)



def Universal(InputFile):
    extension = Path(InputFile).suffix
    QuantityIsInThousands = True
    
    Groupslist = []
    if extension == '.xls':
        savepath = SaveAsXLSX(InputFile)
        book = openpyxl.load_workbook(savepath)
    elif extension == '.xlsb':
        print("Сохраняю xlsb как xlsx")
        df = pd.read_excel(InputFile, engine='pyxlsb')
        savepath = Path(Path(InputFile).parent, str(Path(InputFile).stem+" Перекодированный.xlsx"))
        df.to_excel(savepath)
        book = openpyxl.load_workbook(savepath)
        print("Сохранение завершено")
    else:
        book = openpyxl.load_workbook(InputFile)
    sheet = book.active
    
    expfilename = Path(InputFile).stem + " Результат.xlsx"
    expfilepath = Path(InputFile).parent
    expfilepath = Path(expfilepath, expfilename)
    #print(expfilepath)
    
    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Приход"
    updnumlocs = FindAllInstancesOnSheet("Референс",sheet)
    
    
    #sellertext = ""
    #seller = FindAllInstancesOnSheet("Продавец:",sheet)
    
    #INNtext = ""
    #INN = FindAllInstancesOnSheet("ИНН/КПП продавца:",sheet)
    #if len(seller)!=1:
        
    
    #print(updnumlocs)
    exprowcounter = 0 
    updcounter = len(updnumlocs)-1
    print("Найдено Прайсов в файле:")
    print(len(updnumlocs)-1)
    
    for upd in updnumlocs:
        
        if upd!=-1:

            #exprowcounter=exprowcounter+1
        #                 1         2        3            4              5          6        7       8         9                10
            fieldnames = ['код','Референс', 'Описание Референса', 'Тариф', 'Группа', 'Тариф+Скидка (без НДС)', 'тариф+скидка (С НДС)', '', '', 'Группа', 'Скидка',',C НДС скидка']
           
            exprowcounter=exprowcounter+1
            expsheet.append(fieldnames)
            
            modelcolumn = 1
            quantitycolumn = 1
            pricecolumn = 1
            overallpricecolumn = 1
            gtdnumcolumn = 1
            countrycolumn = 1   
            
            counter = 0
            trigger = False
            startofthenextupd = sheet.max_row
            if updcounter>1:
                startofthenextupd = updnumlocs[updnumlocs.index(upd)+1][0]
                updcounter=updcounter-1
            
            for i in range(upd[0], startofthenextupd+1):

                if trigger:
                        #print(pricecolumn)
                        #print(refcolumn)
                        if has_numbers(str(sheet.cell(row=i, column=pricecolumn).value)):
                            exprowcounter = exprowcounter + 1
                            
                            
                            
                            ref =         str(sheet.cell(row=i, column=refcolumn).value)
                            refdesc =      str(sheet.cell(row=i, column=refdesccolumn).value)
                            price =         str(sheet.cell(row=i, column=pricecolumn).value)
                            group =         str(sheet.cell(row=i, column=groupcolumn).value)
                            if Groupslist.count(group) == 0:
                                Groupslist.append(group)
                                expsheet.cell(row=Groupslist.index(group)+1, column=10).value = group
                                expsheet.cell(row=Groupslist.index(group)+1, column=11).value = '0'
                                expsheet.cell(row=Groupslist.index(group)+1, column=12).value = '0'
                            correctedprice = "=D"+str(exprowcounter)+"-("+"D"+str(exprowcounter)+" /100*"+"K"+str(Groupslist.index(group)+1)+")"
                            correctedpricenal = "=1.2*D"+str(exprowcounter)+"-("+"1.2*D"+str(exprowcounter)+" /100*"+"L"+str(Groupslist.index(group)+1)+")"
                            #overallprice =  str(sheet.cell(row=i, column=overallpricecolumn).value)

                                
                            price = price.replace("-", ",")
                            price = price.replace(" ", "")

                            
                            expsheet.cell(row=exprowcounter, column=2).value = ref        #'Артикул'
                            expsheet.cell(row=exprowcounter, column=3).value = refdesc     #'Количество'
                            expsheet.cell(row=exprowcounter, column=4).value = float(price)        #'Цена'
                            expsheet.cell(row=exprowcounter, column=5).value = group #'Сумма'
                            expsheet.cell(row=exprowcounter, column=6).value = correctedprice
                            expsheet.cell(row=exprowcounter, column=7).value = correctedpricenal                            #'Сумма'

                        else:
                            counter = counter + 1
                #print(str(sheet.cell(i, modelcolumn+1).value))
                
                elif AdressInRow(["Референс","Референс"],i,sheet)!=-1:

                    refcolumn = AdressInRow(["Референс"],i,sheet)

                        
                    refdesccolumn = AdressInRow(["Описание референса"],i,sheet)
                    pricecolumn = AdressInRow(["Тариф"],i,sheet)
                    groupcolumn = AdressInRow(["Коллекция"],i,sheet)

                    #print(countrycolumn)

                    trigger = True
        





            #for i in range(1, expsheet.max_row+1):
            #    origcode = str(expsheet.cell(row=i, column=2).value)

            #    for k in range(0,datasheet.nrows):
            #        datacode = str(datasheet.cell(k, 1).value)
                    
            #        if origcode == datacode:
            #            expsheet.cell(row=i, column=1).value = datasheet.cell(k,0).value
            #            break
  #  msg_box = messagebox.askquestion('Подтверждение', 'Удалить товар, для которого не найдены коды?')
 #   if msg_box == 'yes':
        #    print("Начинаю удалять")
       #     counter = 2
      #      for i in range(2, expsheet.max_row+1):
     #           origcode = str(expsheet.cell(row=i, column=2).value)
                #counter = counter+1
               # if (i%100==0):
              #      print("Пройдено строк "+ str(i))
             #   if not expsheet.cell(row=i, column=1).value:
            #        expsheet.cell(row=i, column=2).value == None
           #         expsheet.cell(row=i, column=3).value == None
          #          expsheet.cell(row=i, column=4).value == None
         #           expsheet.cell(row=i, column=5).value == None
        #            expsheet.cell(row=i, column=6).value == None
       #             expsheet.cell(row=i, column=7).value == None
      #              expsheet.cell(row=i, column=8).value == None
     #               #counter = counter-1
    #            else:
                    #expsheet.cell(row=counter, column=2).value = expsheet.cell(row=i, column=2).value
                   # expsheet.cell(row=counter, column=3).value = expsheet.cell(row=i, column=3).value
                  #  expsheet.cell(row=counter, column=4).value = expsheet.cell(row=i, column=4).value
                 #   expsheet.cell(row=counter, column=5).value = expsheet.cell(row=i, column=5).value
                #    expsheet.cell(row=counter, column=6).value = expsheet.cell(row=i, column=6).value
               #     expsheet.cell(row=counter, column=7).value = expsheet.cell(row=i, column=7).value
              #      expsheet.cell(row=counter, column=8).value = expsheet.cell(row=i, column=8).value
             #       expsheet.cell(row=counter, column=1).value = expsheet.cell(row=i, column=1).value
            #        expsheet.cell(row=i, column=1).value == None
           #         expsheet.cell(row=i, column=2).value == None
          #          expsheet.cell(row=i, column=3).value == None
         #           expsheet.cell(row=i, column=4).value == None
        #            expsheet.cell(row=i, column=5).value == None
       #             expsheet.cell(row=i, column=6).value == None
      #              expsheet.cell(row=i, column=7).value == None
     #               expsheet.cell(row=i, column=8).value == None
    #                counter = counter+1
    SearchCodesFromExcel(expsheet,2,1)
    print("Начинаю сохранять результат")
    try:
        expfile.save(expfilepath)
    except:
        msg_box = messagebox.askquestion('Подтверждение', 'Экспортный файл уже открыт, закройте и попробуйте еще раз')
        if msg_box == 'yes':
            expfile.save(expfilepath)
    print("Открываю файл с результатами...")
    os.system('"{0}"'.format(expfilepath))


def SearchCodesFromExcel(mainsheet, inputcolumn,outputcolumn):
    msg_box = messagebox.askquestion('Подтверждение', 'Выбрать файл расшифровки артикулов ?')
    if msg_box == 'yes':
        databookpath = filedialog.askopenfilename(title='Выберите таблицу с кодами', filetypes=(('.XLS', '.xls'),))
        
        if databookpath:

            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            print("Начинаю поиск кодов")
            Expmassive = []
            Datamassive =[]
            for i in range(1, mainsheet.max_row+1):
                Expmassive.append(str(mainsheet.cell(i, inputcolumn).value))
                #print (Expmassive)
            for k in range(0,datasheet.nrows):
                Datamassive.append(str(datasheet.cell(k, 1).value))
            for i in range(1, mainsheet.max_row+1):
                if (i%100==0):
                    print("Пройдено строк "+ str(i))
                for k in range(0,datasheet.nrows):
                    if Expmassive[i-1]== Datamassive[k]:
                        mainsheet.cell(i, outputcolumn).value = datasheet.cell(k,0).value
                        break

def Decode(data):
    decoded = data.encode('CP1252', errors='ignore').decode('CP1251', errors='ignore')
    return decoded


def enc_test():
    SelectFile('xls')
    
    if IsInputSel:
        book = xlrd.open_workbook(InputFile)
        sheet = book.sheet_by_index(0)
        encoding = book.encoding
        print('***** {0}'.format(book.encoding))
        
        for i in range(0, 5):
            for k in range(0, 5):
                print('r={0},col={1} // res={2} //'.format(i,k,sheet.cell(i,k).value))
        
        text = sheet.cell(3,8).value
        print("#### after: ".format(text))
        decoded = text.encode('CP1252').decode('CP1251', errors='ignore')
        print("#### before: ".format(decoded))



def has_numbers(inputString):
    #number_pattern = re.compile(r'-?\d+(\.\d+)?')
    #numbers = number_pattern.findall(inputString)
    #return numbers
    return bool(re.search(r'\d', inputString))
    
def has_letters(inputString):
    cyrillic_pattern = re.compile('[а-яА-ЯёЁ]')
    contains_cyrillic = bool(cyrillic_pattern.search(inputString))
    return contains_cyrillic

if __name__ == '__main__':
    main()
