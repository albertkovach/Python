from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

global root
global scrnwparam
global scrnhparam
global SuppliersListbox
scrnwparam = 185
scrnhparam = 150

import os, requests,re
from pathlib import Path
import xlrd
import openpyxl
from openpyxl import Workbook
from configparser import ConfigParser

def main():
    global root

    root = Tk()
    root.resizable(False, False)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('600x210+{}+{}'.format(scrnw, scrnh))
 
    app = GUI(root)
    root.mainloop()

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):
        
        global SuppliersListbox
        global languages
        config = ConfigParser()
        config.read('config.ini')
        #for sections in config
        #    languages.append()
        languages = config.sections()
        languages_var = Variable(value=languages)
     
        SuppliersListbox = Listbox(listvariable=languages_var)
        SuppliersListbox.bind('<<ListboxSelect>>', onselect)
        SuppliersListbox.place(x=10, y=10, width=150, height=150)
        
        
        FirstRowY= 10
        DefaultWidth=50
        DefaultHeight=150
        DefaultDistanceX = DefaultWidth+10
        
        global ConfigBox
        ConfigBox = Text(height = 13, width = 40)
        ConfigBox.place(x=180, y=FirstRowY, width=400, height=DefaultHeight)   
        
        
        
        global ExecutePricelist
        ExecutePricelist = Button(text='Обработать прайс', command=Avtek, font=("Arial", 11))
        ExecutePricelist.place(x=180, y=170, width=200, height=30)
        
        global NewConfigBtn
        NewConfigBtn = Button(text='♳', command=GenerateDefaultIni, font=("Arial", 12))
        NewConfigBtn.place(x=550, y=170, height=30)
        
        global UpdateConfigBtn
        UpdateConfigBtn = Button(text='⟳', command=UpdateGUI, font=("Arial", 12))
        UpdateConfigBtn.place(x=510, y=170, height=30)
        



def onselect(evt):
    # Note here that Tkinter passes an event object to onselect()
    w = evt.widget
    index = int(w.curselection()[0])
    value = w.get(index)
    config = ConfigParser()
    
    config.read('config.ini')


    zakupdiscount = float(config.get(value,'СкидкаНаЗакуп'))
    zakupdiscount2 = float(config.get(value,'СкидкаНаЗакуп2'))
    cashdiscount = float(config.get(value,'НаценкаНаНал'))
    noncashdiscount = float(config.get(value,'НаценкаНаБезнал'))
    articlecolumn = int(config.get(value,'НомерСтрокиСАртикулом'))
    pricecolumn = int(config.get(value,'НомерСтрокиСЦеной'))
    namecolumn = int(config.get(value,'НомерСтрокиСНазванием'))
    
    
    print('You selected item %d: "%s"' % (index, value))
    ConfigBox.delete('1.0', END)
    ConfigBox.insert('end', value)
    ConfigBox.insert('end', '\nСкидкаНаЗакуп '+str(zakupdiscount))
    ConfigBox.insert('end', '\nСкидкаНаЗакуп2 '+str(zakupdiscount2))
    ConfigBox.insert('end', '\nНаценкаНаНал '+str(cashdiscount))
    ConfigBox.insert('end', '\nНаценкаНаБезнал '+str(noncashdiscount))
    ConfigBox.insert('end', '\nНомерСтрокиСАртикулом '+str(articlecolumn))
    ConfigBox.insert('end', '\nНомерСтрокиСЦеной ' +str(pricecolumn))
    ConfigBox.insert('end', '\nНомерСтрокиСНазванием '+str(namecolumn))


def TDM():
    dir_path = os.path.dirname(os.path.realpath(__file__))
    filename = 'TDM.xls'
    save_path = os.path.join(dir_path, filename)
    url = 'https://tdme.ru/download/priceTDM.xls'
    response = requests.get(url)

    if response.status_code == 200:
        with open(save_path, 'wb') as f:
            f.write(response.content)
            
        print('File saved successfully.')

    else:
        print('Failed to download file.')
        
        
    book = xlrd.open_workbook(save_path)
    sheet = book.sheet_by_index(0)
    trigger = False
    expfilename = "ТДМ на залив.xlsx"
    expfilepath = Path(Path(save_path).parent, expfilename)
    
    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Приход"
    #               1        2             3           4        5        6    
    fieldnames = ['Код', 'Артикул', 'Наименование', 'Закуп -30 -6', 'Безнал +22', 'Нал +15']
    expsheet.append(fieldnames)
    exprowc = 2
    
    Zakupdiscount = (30/100)*(1-0.06)
    cashinterest = 15/100
    noncashinterest = 22/100
    
    for i in range(0,sheet.nrows):
        celltype = sheet.cell_type(i, 4)
        if celltype != xlrd.XL_CELL_EMPTY:
            if trigger:
                article = sheet.cell(i,2).value
                name = sheet.cell(i,1).value
                
                expsheet.cell(row=exprowc, column=2).value = article
                expsheet.cell(row=exprowc, column=3).value = name
                
                try:
                    price = float(sheet.cell(i,4).value)*(1-Zakupdiscount)
                    cash = price*(1+cashinterest)
                    noncash = price*(1+noncashinterest)
                    
                except:
                    price = float(0)
                    noncash = price
                    cash = price
                    
                expsheet.cell(row=exprowc, column=4).value = price
                expsheet.cell(row=exprowc, column=5).value = noncash
                expsheet.cell(row=exprowc, column=6).value = cash
                
                exprowc = exprowc + 1
            else:
                if "дату" in sheet.cell(i,0).value:
                    ver = sheet.cell(i,0).value
                    print(ver)
                if "Цена" in sheet.cell(i,4).value:
                    trigger = True
 
    SearchCodesFromExcel(expsheet,2,1)
 
    expfile.save(expfilepath)
    print("Opening results file...")
    os.system('"{0}"'.format(expfilepath))
    
def TDMcable():
    dir_path = os.path.dirname(os.path.realpath(__file__))
    filename = 'TDMкабель.xls'
    save_path = os.path.join(dir_path, filename)
    url = 'https://tdme.ru/download/zayavka77kpp.xls'
    response = requests.get(url)

    if response.status_code == 200:
        with open(save_path, 'wb') as f:
            f.write(response.content)
            
        print('File saved successfully.')

    else:
        print('Failed to download file.')
        
        
    book = xlrd.open_workbook(save_path)
    sheet = book.sheet_by_index(0)
    trigger = False
    expfilename = "ТДМ кабель на залив.xlsx"
    expfilepath = Path(Path(save_path).parent, expfilename)
    
    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Приход"
    #               1        2             3           4        5        6    
    fieldnames = ['Код', 'Артикул', 'Наименование', 'Закуп', 'Безнал +7', 'Нал +5']
    expsheet.append(fieldnames)
    exprowc = 2
    
    Zakupdiscount = 0
    cashinterest = 5
    noncashinterest = 7
    
    for i in range(0,sheet.nrows):
        celltype = sheet.cell_type(i, 8)
        if celltype != xlrd.XL_CELL_EMPTY:
            if trigger:
                article = sheet.cell(i,0).value
                name = sheet.cell(i,1).value
                
                expsheet.cell(row=exprowc, column=2).value = article
                expsheet.cell(row=exprowc, column=3).value = name
                
                try:
                    price = float(sheet.cell(i,8).value)*(1-Zakupdiscount)
                    cash = price*(1+cashinterest)
                    noncash = price*(1+noncashinterest)
                    
                except:
                    price = float(0)
                    noncash = price
                    cash = price
                    
                expsheet.cell(row=exprowc, column=4).value = price
                expsheet.cell(row=exprowc, column=5).value = noncash
                expsheet.cell(row=exprowc, column=6).value = cash
                
                exprowc = exprowc + 1
            else:
                if "дату" in sheet.cell(i,0).value:
                    ver = sheet.cell(i,0).value
                    print(ver)
                if "Артикул" in sheet.cell(i,0).value:
                    trigger = True
    
    SearchCodesFromExcel(expsheet,2,1)
    
    expfile.save(expfilepath)
    print("Opening results file...")
    os.system('"{0}"'.format(expfilepath))

def Avtek():
    selection = SuppliersListbox.curselection()
    config = ConfigParser()
    
    config.read('config.ini')
    #IsUsingURL = config.get(languages[selection[0]],'СкачиватьПоURL')
    zakupdiscount = float(config.get(languages[selection[0]],'СкидкаНаЗакуп'))
    zakupdiscount2 = float(config.get(languages[selection[0]],'СкидкаНаЗакуп2'))
    cashdiscount = float(config.get(languages[selection[0]],'НаценкаНаНал'))
    noncashdiscount = float(config.get(languages[selection[0]],'НаценкаНаБезнал'))
    articlecolumn = int(config.get(languages[selection[0]],'НомерСтрокиСАртикулом'))
    pricecolumn = int(config.get(languages[selection[0]],'НомерСтрокиСЦеной'))
    namecolumn = int(config.get(languages[selection[0]],'НомерСтрокиСНазванием'))
    
    
    
    try:
        url = config.get(languages[selection[0]],'url')
        dir_path = os.path.dirname(os.path.realpath(__file__))
        filename = 'ЗагруженныйПрайс.xls'
        InputFile = os.path.join(dir_path, filename)
    
    
    #url = 'https://drive.google.com/uc?export=download&confirm=no_antivirus&id=1NtdgO2iatpD17EAGzJEawvz0myYO7mXN'
        response = requests.get(url)


        if response.status_code == 200:
            with open(InputFile, 'wb') as f:
                f.write(response.content)
            
            print('File saved successfully.')

        else:
            print('Failed to download file.')
    except:
        InputFile = SelectFile('.xlsb .xlsx .xls')
    
    
    
    
    
    
    extension = Path(InputFile).suffix
    if extension == '.xls':
        savepath = SaveAsXLSX(InputFile)
        book = openpyxl.load_workbook(savepath)
    elif extension == '.xlsb':
        print("Сохраняю xlsb как xlsx")
        df = pd.read_excel(InputFile, engine='pyxlsb')
        savepath = Path(Path(InputFile).parent, str(Path(InputFile).stem+" Перекодированный.xlsx"))
        df.to_excel(savepath)
        book = openpyxl.load_workbook(savepath)
        print("Сохранение завершено")
    else:
        book = openpyxl.load_workbook(InputFile)    
    #book = xlrd.open_workbook(save_path)
    sheet = book.active
    #trigger = False
    
    expfilename = "Прайс на залив.xlsx"
    expfilepath = Path(Path(InputFile).parent, expfilename)
    
    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Приход"
    #               1        2             3           4        5        6    
    fieldnames = ['Код', 'Артикул', 'Наименование','Закуп изначальный' ,'Закуп  '+str(zakupdiscount)+'%  '+str(zakupdiscount2)+'%', 'Безнал '+str(noncashdiscount)+'%', 'Нал '+str(cashdiscount)+'%','8','9','10','Закуп  '+str(zakupdiscount)+'%  ','Закуп 2  '+str(zakupdiscount2)+'%','Безнал '+str(noncashdiscount)+'%','Нал '+str(cashdiscount)+'%']
    expsheet.append(fieldnames)
    
    expsheet.cell(row=2, column=11).value = zakupdiscount
    expsheet.cell(row=2, column=12).value = zakupdiscount2
    expsheet.cell(row=2, column=13).value = noncashdiscount
    expsheet.cell(row=2, column=14).value = cashdiscount
    
    
    
    
    exprowc = 2
    

    
    #noncashdiscount = 5
    trigger = False
    print(pricecolumn)
    for i in range(2,sheet.max_row+1):
        #celltype = sheet.cell_type(i, 4)
        if has_numbers(str(sheet.cell(i,pricecolumn).value)):
            #print(i)

            #print(trigger)
            article = sheet.cell(i,articlecolumn).value
            name = sheet.cell(i,namecolumn).value
            
            expsheet.cell(exprowc, 2).value = article
            expsheet.cell(exprowc, 3).value = name
            price1 = 0
            try:
                price1 = float(sheet.cell(i,pricecolumn).value)
                percentage1 = (100+zakupdiscount)/100
                percentage2 = (100+zakupdiscount2)/100
                price = price1*percentage1*percentage2
                cash = price*(100+cashdiscount)/100
                noncash = price*(100+noncashdiscount)/100
            except:
                price1 = 0
                price = 0
                noncash = 0
                cash = 0
             #   percentage1 = (100+zakupdiscount)/100

            #noncash = price1
            #cash = price1
                
            expsheet.cell(row=exprowc, column=4).value = price1
            expsheet.cell(row=exprowc, column=5).value = price
            expsheet.cell(row=exprowc, column=6).value = noncash
            expsheet.cell(row=exprowc, column=7).value = cash
            
            exprowc = exprowc + 1

                    
    SearchCodesFromExcel(expsheet,2,1)
    
    print("Начинаю сохранять результат")
    try:
        expfile.save(expfilepath)
    except:
        msg_box = messagebox.askquestion('Подтверждение', 'Экспортный файл уже открыт, закройте и попробуйте еще раз')
        if msg_box == 'yes':
            expfile.save(expfilepath)
    print("Открываю файл с результатами...")
    os.system('"{0}"'.format(expfilepath))
  





def GenerateDefaultIni():
    config = ConfigParser()

    config.read('config.ini')
    config.add_section('TDM')
    #config.set('TDM', 'СкачиватьПоURL', 'Да')
    config.set('TDM', 'url', 'https://tdme.ru/download/priceTDM.xls')
    config.set('TDM', 'СкидкаНаЗакуп', '-30')
    config.set('TDM', 'СкидкаНаЗакуп2', '-6')
    config.set('TDM', 'НаценкаНаБезнал', '25')
    config.set('TDM', 'НаценкаНаНал', '15')
    config.set('TDM', 'НомерСтрокиСАртикулом', '2')
    config.set('TDM', 'НомерСтрокиСЦеной', '5')
    config.set('TDM', 'НомерСтрокиСНазванием', '3')

    config.add_section('TDMcable')
    config.set('TDM', 'СкачиватьПоURL', 'Да')
    config.set('TDMcable', 'url', 'https://tdme.ru/download/zayavka77kpp.xls')
    config.set('TDMcable', 'СкидкаНаЗакуп', '0')
    config.set('TDMcable', 'СкидкаНаЗакуп2', '0')
    config.set('TDMcable', 'НаценкаНаБезнал', '5')
    config.set('TDMcable', 'НаценкаНаНал', '7')
    config.set('TDMcable', 'НомерСтрокиСАртикулом', '1')
    config.set('TDMcable', 'НомерСтрокиСЦеной', '7')
    config.set('TDMcable', 'НомерСтрокиСНазванием', '2')
    
    config.add_section('Avtek')
    config.set('TDM', 'СкачиватьПоURL', 'Да')
    config.set('Avtek', 'url', 'https://drive.google.com/uc?export=download&confirm=no_antivirus&id=1NtdgO2iatpD17EAGzJEawvz0myYO7mXN')
    config.set('Avtek', 'СкидкаНаЗакуп', '0')
    config.set('Avtek', 'СкидкаНаЗакуп2', '0')
    config.set('Avtek', 'НаценкаНаБезнал', '5')
    config.set('Avtek', 'НаценкаНаНал', '5')
    config.set('Avtek', 'НомерСтрокиСАртикулом', '3')
    config.set('Avtek', 'НомерСтрокиСЦеной', '4')
    config.set('Avtek', 'НомерСтрокиСНазванием', '1')
    with open('config.ini', "w") as f:
        config.write(f)



def SaveAsXLSX(InputFile):
    encod = False
    book_xls = xlrd.open_workbook(InputFile)
    book_xlsx = Workbook()
    if str(format(book_xls.encoding)) == "iso-8859-1":
        print(str(format(book_xls.encoding))+" Декодирую и сохраняю как xlsx")
        encod = True
    else:
        print(" Сохраняю без декодирования как xlsx")
        encod = False
    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
    if sheet_index == 0:
        sheet_xlsx = book_xlsx.active
        sheet_xlsx.title = sheet_name
    else:
        sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

    for row in range(0, sheet_xls.nrows):
        for col in range(0, sheet_xls.ncols):
            if encod == True:
                sheet_xlsx.cell(row = row+1 , column = col+1).value = Decode(str(sheet_xls.cell_value(row, col)))
            else:
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
    savepath = Path(Path(InputFile).parent, str(Path(InputFile).stem+" Перекодированный.xlsx"))
    #print("Сохранил в " + savepath)
    book_xlsx.save(savepath)
    return savepath



def UpdateGUI():
    config = ConfigParser()
    config.read('config.ini')
    #for sections in config
    #    languages.append()
    languages = config.sections()
    languages_var = Variable(value=languages)
 
    SuppliersListbox.delete(0, END)
    for n in languages:
        SuppliersListbox.insert(languages.index(n), n)



def SelectFile(extension):
    #global InputFile
    #global IsInputSel
    extname = extension.upper()

    InputFile = filedialog.askopenfilename(title='Выберите УПД', filetypes=((extname, extension),))
    
    if InputFile:
        print('Файл выбран :' + InputFile)
        
        #print(str(IsInputSel))
        return InputFile
    else:
        print('Файл не выбран')
        return False



def SearchCodesFromExcel(mainsheet, inputcolumn,outputcolumn):
    msg_box = messagebox.askquestion('Подтверждение', 'Выбрать файл расшифровки артикулов ?')
    if msg_box == 'yes':
        databookpath = filedialog.askopenfilename(title='Выберите таблицу с кодами', filetypes=(('.XLS', '.xls'),))
        
        if databookpath:

            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            print("Начинаю поиск кодов")
            Expmassive = []
            Datamassive =[]
            for i in range(1, mainsheet.max_row+1):
                Expmassive.append(str(mainsheet.cell(i, inputcolumn).value))
                #print (Expmassive)
            for k in range(0,datasheet.nrows):
                Datamassive.append(str(datasheet.cell(k, 1).value))
            for i in range(1, mainsheet.max_row+1):
                if (i%100==0):
                    print("Пройдено строк "+ str(i))
                for k in range(0,datasheet.nrows):
                    if Expmassive[i-1]== Datamassive[k]:
                        mainsheet.cell(i, outputcolumn).value = datasheet.cell(k,0).value
                        break



def has_numbers(inputString):
    #number_pattern = re.compile(r'-?\d+(\.\d+)?')
    #numbers = number_pattern.findall(inputString)
    #return numbers
    return bool(re.search(r'\d', inputString))



if __name__ == '__main__':
    main()
