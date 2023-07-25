from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

global root
global scrnwparam
global scrnhparam
scrnwparam = 185
scrnhparam = 150

import os, requests
from pathlib import Path
import xlrd
import openpyxl
from openpyxl import Workbook


def main():
    global root

    root = Tk()
    root.resizable(False, False)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('240x130+{}+{}'.format(scrnw, scrnh))
        
    app = GUI(root)
    root.mainloop()

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):
        global AvtekBtn
        AvtekBtn = Button(text='Автэк', command=Avtek, font=("Arial", 11))
        AvtekBtn.place(x=15, y=15, width=100, height=30)
        
        global TDMBtn
        TDMBtn = Button(text='ТДМ', command=TDM, font=("Arial", 11))
        TDMBtn.place(x=15, y=50, width=100, height=30)
       
        global TDMcableBtn
        TDMcableBtn = Button(text='ТДМ кабель', command=TDMcable, font=("Arial", 11))
        TDMcableBtn.place(x=15, y=85, width=100, height=30)


def TDM():
    dir_path = os.path.dirname(os.path.realpath(__file__))
    filename = 'TDM.xls'
    save_path = os.path.join(dir_path, filename)
    url = 'https://tdme.ru/download/priceTDM.xls'
    response = requests.get(url)

    if response.status_code == 200:
        with open(save_path, 'wb') as f:
            f.write(response.content)
            
        print('File saved successfully.')

    else:
        print('Failed to download file.')
        
        
    book = xlrd.open_workbook(save_path)
    sheet = book.sheet_by_index(0)
    trigger = False
    expfilename = "ТДМ на залив.xlsx"
    expfilepath = Path(Path(save_path).parent, expfilename)
    
    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Приход"
    #               1        2             3           4        5        6    
    fieldnames = ['Код', 'Артикул', 'Наименование', 'Закуп -30 -6', 'Безнал +22', 'Нал +15']
    expsheet.append(fieldnames)
    exprowc = 2
    
    Zakupdiscount = (30/100)*(1-0.06)
    cashinterest = 15/100
    noncashinterest = 22/100
    
    for i in range(0,sheet.nrows):
        celltype = sheet.cell_type(i, 4)
        if celltype != xlrd.XL_CELL_EMPTY:
            if trigger:
                article = sheet.cell(i,2).value
                name = sheet.cell(i,1).value
                
                expsheet.cell(row=exprowc, column=2).value = article
                expsheet.cell(row=exprowc, column=3).value = name
                
                try:
                    price = float(sheet.cell(i,4).value)*(1-Zakupdiscount)
                    cash = price*(1+cashinterest)
                    noncash = price*(1+noncashinterest)
                    
                except:
                    price = float(0)
                    noncash = price
                    cash = price
                    
                expsheet.cell(row=exprowc, column=4).value = price
                expsheet.cell(row=exprowc, column=5).value = noncash
                expsheet.cell(row=exprowc, column=6).value = cash
                
                exprowc = exprowc + 1
            else:
                if "дату" in sheet.cell(i,0).value:
                    ver = sheet.cell(i,0).value
                    print(ver)
                if "Цена" in sheet.cell(i,4).value:
                    trigger = True
    
    expfile.save(expfilepath)
    print("Opening results file...")
    os.system('"{0}"'.format(expfilepath))
    
def TDMcable():
    dir_path = os.path.dirname(os.path.realpath(__file__))
    filename = 'TDMкабель.xls'
    save_path = os.path.join(dir_path, filename)
    url = 'https://tdme.ru/download/zayavka77kpp.xls'
    response = requests.get(url)

    if response.status_code == 200:
        with open(save_path, 'wb') as f:
            f.write(response.content)
            
        print('File saved successfully.')

    else:
        print('Failed to download file.')
        
        
    book = xlrd.open_workbook(save_path)
    sheet = book.sheet_by_index(0)
    trigger = False
    expfilename = "ТДМ кабель на залив.xlsx"
    expfilepath = Path(Path(save_path).parent, expfilename)
    
    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Приход"
    #               1        2             3           4        5        6    
    fieldnames = ['Код', 'Артикул', 'Наименование', 'Закуп', 'Безнал +7', 'Нал +5']
    expsheet.append(fieldnames)
    exprowc = 2
    
    Zakupdiscount = 0
    cashinterest = 5
    noncashinterest = 7
    
    for i in range(0,sheet.nrows):
        celltype = sheet.cell_type(i, 8)
        if celltype != xlrd.XL_CELL_EMPTY:
            if trigger:
                article = sheet.cell(i,0).value
                name = sheet.cell(i,1).value
                
                expsheet.cell(row=exprowc, column=2).value = article
                expsheet.cell(row=exprowc, column=3).value = name
                
                try:
                    price = float(sheet.cell(i,8).value)*(1-Zakupdiscount)
                    cash = price*(1+cashinterest)
                    noncash = price*(1+noncashinterest)
                    
                except:
                    price = float(0)
                    noncash = price
                    cash = price
                    
                expsheet.cell(row=exprowc, column=4).value = price
                expsheet.cell(row=exprowc, column=5).value = noncash
                expsheet.cell(row=exprowc, column=6).value = cash
                
                exprowc = exprowc + 1
            else:
                if "дату" in sheet.cell(i,0).value:
                    ver = sheet.cell(i,0).value
                    print(ver)
                if "Артикул" in sheet.cell(i,0).value:
                    trigger = True
    
    expfile.save(expfilepath)
    print("Opening results file...")
    os.system('"{0}"'.format(expfilepath))


def Avtek():

    dir_path = os.path.dirname(os.path.realpath(__file__))
    filename = 'Автэк.xls'
    save_path = os.path.join(dir_path, filename)

    url = 'https://drive.google.com/uc?export=download&confirm=no_antivirus&id=1NtdgO2iatpD17EAGzJEawvz0myYO7mXN'
    response = requests.get(url)

    if response.status_code == 200:
        with open(save_path, 'wb') as f:
            f.write(response.content)
            
        print('File saved successfully.')

    else:
        print('Failed to download file.')
        
        
    book = xlrd.open_workbook(save_path)
    sheet = book.sheet_by_index(0)
    trigger = False
    
    expfilename = "Автэк на залив.xlsx"
    expfilepath = Path(Path(save_path).parent, expfilename)
    
    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Приход"
    #               1        2             3           4        5        6    
    fieldnames = ['Код', 'Артикул', 'Наименование', 'Закуп', 'Безнал -5', 'Нал -5']
    expsheet.append(fieldnames)
    exprowc = 2
    
    noncashdiscount = 5
    
    for i in range(0,sheet.nrows):
        celltype = sheet.cell_type(i, 0)
        if celltype != xlrd.XL_CELL_EMPTY:
            if trigger:
                article = sheet.cell(i,2).value
                name = sheet.cell(i,0).value
                
                expsheet.cell(row=exprowc, column=2).value = article
                expsheet.cell(row=exprowc, column=3).value = name
                
                try:
                    price = float(sheet.cell(i,3).value)/1000
                    noncash = price-(noncashdiscount/100)*price
                    cash = noncash
                except:
                    price = float(0)
                    noncash = price
                    cash = price
                    
                expsheet.cell(row=exprowc, column=4).value = price
                expsheet.cell(row=exprowc, column=5).value = noncash
                expsheet.cell(row=exprowc, column=6).value = cash
                
                exprowc = exprowc + 1
            else:
                if "дату" in sheet.cell(i,0).value:
                    ver = sheet.cell(i,0).value
                    print(ver)
                if "ТМЦ" in sheet.cell(i,0).value:
                    trigger = True
    
    expfile.save(expfilepath)
    print("Opening results file...")
    os.system('"{0}"'.format(expfilepath))
    



if __name__ == '__main__':
    main()
