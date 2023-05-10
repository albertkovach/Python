from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

import requests, os, sys, csv
from pathlib import Path

import xml.etree.ElementTree as ET

from openpyxl import Workbook

global root
global scrnwparam
global scrnhparam


scrnwparam = 185
scrnhparam = 150


def main():
    global root

    root = Tk()
    root.resizable(True, True)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('240x200+{}+{}'.format(scrnw, scrnh))
        
    app = GUI(root)
    root.mainloop()

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("XML DECODE")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):
        global dir_path
        dir_path = Path(os.path.dirname(os.path.realpath(__file__)))
        print(dir_path)
    
        global Decode
        Decode = True
        
        global DecodeFile
        global IsDecodeFileSelected
        IsDecodeFileSelected = False
    
        
        global FeronBtn
        FeronBtn = Button(text='Feron', command=DecodeFeron, font=("Arial", 11))
        FeronBtn.place(x=15, y=15, width=100, height=30)

        global GetFeronBtn
        GetFeronBtn = Button(text='get', command=GetFeronXml, font=("Arial", 11))
        GetFeronBtn.place(x=125, y=15, width=40, height=30)
        
        
        global Makel25Btn
        Makel25Btn = Button(text='Makel25', command=DecodeMakel, font=("Arial", 11))
        Makel25Btn.place(x=15, y=55, width=100, height=30)

        global GetMakel25Btn
        GetMakel25Btn = Button(text='get', command=GetMakel25Xml, font=("Arial", 11))
        GetMakel25Btn.place(x=125, y=55, width=40, height=30)
        
        
        global DecodeXlsxBtn
        DecodeXlsxBtn = Button(text='Decode in Excel', command=DecodeXLSXsf, font=("Arial", 11))
        DecodeXlsxBtn.place(x=30, y=110, width=160, height=30)
        

        global DecodeTxtBtn
        DecodeTxtBtn = Button(text='Decode in TXT', command=DecodeTXTsf, font=("Arial", 11))
        DecodeTxtBtn.place(x=30, y=150, width=160, height=30)







def GetMakel25Xml():
    msg_box = messagebox.askquestion('Подтверждение', 'Уверены, что хотите скачать файл ?')
    
    if msg_box == 'yes':
        price_url = "https://www.makel25.ru/bitrix/catalog_export/export_dVz.xml"
        
        reqfilepath = Path(dir_path, 'makel25.xml')
        revisedreqfile = reqfilepath.as_posix()
        
        response = requests.get(price_url)
        
        if response.status_code == 200:
            with open(revisedreqfile,'wb') as reqfile:
                reqfile.write(response.content)
                
            messagebox.showinfo('Успех', 'Файл загружен !')
            print(revisedreqfile)
        else:
            messagebox.showinfo('Ошибка', 'Запрос прервался с кодом ошибки: {0}'.format(response.status_code)) 


def GetFeronXml():
    msg_box = messagebox.askquestion('Подтверждение', 'Уверены, что хотите скачать файл ?')
    
    if msg_box == 'yes':
        price_url = "https://shop.feron.ru/bitrix/catalog_export/im.xml"
        
        reqfilepath = Path(dir_path, 'feron.xml')
        revisedreqfile = reqfilepath.as_posix()
        
        response = requests.get(price_url)
        
        if response.status_code == 200:
            with open(revisedreqfile,'wb') as reqfile:
                reqfile.write(response.content)
                
            messagebox.showinfo('Succes', 'File downloaded !')
            print(revisedreqfile)
        else:
            messagebox.showinfo('Ошибка', 'Запрос прервался с кодом ошибки: {0}'.format(response.status_code)) 
   

def DecodeMakel():
    global DecodeFile
    DecodeFile = Path(dir_path, 'makel25.xml')
    
    if os.path.isfile(DecodeFile):
        DecodeXLSX()
    else:
        messagebox.showinfo('Ошибка', 'Отсутствует файл, сначала загрузите его !')


def DecodeFeron():
    global DecodeFile
    DecodeFile = Path(dir_path, 'feron.xml')
    
    if os.path.isfile(DecodeFile):
        DecodeXLSX()
    else:
        messagebox.showinfo('Ошибка', 'Отсутствует файл, сначала загрузите его !')


def DecodeXLSXsf():
    SelectFile()
    if IsDecodeFileSelected:
        DecodeXLSX()   
    
    
def DecodeXLSX():
    global DecodeFile
    global IsDecodeFileSelected
    
    reqfilepath = Path(DecodeFile)
    revisedreqfile = reqfilepath.as_posix()
    
    tree = ET.parse(revisedreqfile)
    root = tree.getroot()
    
    wbpath = Path(Path(DecodeFile).parent, 'DECODED {0}.xlsx'.format(Path(DecodeFile).name))
    wb = Workbook()
    sheetitems = wb.active
    sheetitems.title = "Items"
    sheetcategory = wb.create_sheet("Category")
    

    itemsfieldnames = ['id', 'path', 'picture', 'more_ph', 'param']
    sheetitems.append(itemsfieldnames)
    
    categoryfieldnames = ['category', 'id', 'parentId']
    sheetcategory.append(categoryfieldnames)
    
    rowitem = 1
    rowcategory = 1
    categorypath = []
    for L1 in root:
        for L2 in L1:
            for L3 in L2:
                
                if L3.tag == 'category':
                
                    rowcategory = rowcategory + 1
                    
                    catname = L3.text
                    sheetcategory.cell(row=rowcategory, column=1).value = catname
                    
                    catid = L3.get('id')
                    sheetcategory.cell(row=rowcategory, column=2).value = catid
                    
                    catparentid = L3.get('parentId')
                    sheetcategory.cell(row=rowcategory, column=3).value = catparentid
                    
                    categorypath.append(catname)
                    if sheetcategory.cell(row=rowcategory, column=3).value:
                        for a in range(1, 700):
                            if sheetcategory.cell(row=rowcategory, column=3).value == sheetcategory.cell(row=a, column=2).value:
                                categorypath.append(sheetcategory.cell(row=a, column=1).value)
                                
                                if sheetcategory.cell(row=a, column=3).value:
                                    for b in range(1, 700):
                                        if sheetcategory.cell(row=a, column=3).value == sheetcategory.cell(row=b, column=2).value:
                                            categorypath.append(sheetcategory.cell(row=b, column=1).value)
                                            
                                            if sheetcategory.cell(row=b, column=3).value:
                                                for c in range(1, 700):
                                                    if sheetcategory.cell(row=b, column=3).value == sheetcategory.cell(row=c, column=2).value:
                                                        categorypath.append(sheetcategory.cell(row=c, column=1).value)
                                                        
                                                        if sheetcategory.cell(row=c, column=3).value:
                                                            for d in range(1, 700):
                                                                if sheetcategory.cell(row=c, column=3).value == sheetcategory.cell(row=d, column=2).value:
                                                                    categorypath.append(sheetcategory.cell(row=d, column=1).value)
                                                                    
                                                                    if sheetcategory.cell(row=d, column=3).value:
                                                                        for e in range(1, 700):
                                                                            if sheetcategory.cell(row=d, column=3).value == sheetcategory.cell(row=e, column=2).value:
                                                                                categorypath.append(sheetcategory.cell(row=e, column=1).value)
                                                                                
                                                                                if sheetcategory.cell(row=e, column=3).value:
                                                                                    for f in range(1, 700):
                                                                                        if sheetcategory.cell(row=e, column=3).value == sheetcategory.cell(row=f, column=2).value:
                                                                                            categorypath.append(sheetcategory.cell(row=f, column=1).value)
                    
                    categorypath.reverse()
                    sheetcategory.cell(row=rowcategory, column=4).value = "|".join(categorypath)
                    categorypath = []
            
            
            
            
                if L3.tag == 'offer':
                    if Decode: print("**** Item : ", L3.attrib)
                    if Decode: print("")
                    
                    rowitem = rowitem + 1
                    
                    name = L3.attrib.get('id')
                    sheetitems.cell(row=rowitem, column=1).value = name

                    for L4 in L3:
                        if itemsfieldnames.count(L4.tag) == 0:
                            itemsfieldnames.append(L4.tag)
                            sheetitems.cell(row=1, column=itemsfieldnames.index(L4.tag)+1).value = L4.tag
                            
                    
                    for L4 in L3:
                        if L4.tag == 'categoryId':
                            sheetitems.cell(row=rowitem, column=itemsfieldnames.index(L4.tag)+1).value = L4.text
                            
                            for i in range(1, 700):
                                if sheetitems.cell(row=rowitem, column=itemsfieldnames.index(L4.tag)+1).value == sheetcategory.cell(row=i, column=2).value:
                                    sheetitems.cell(row=rowitem, column=2).value = sheetcategory.cell(row=i, column=4).value
                            
                        elif L4.tag == 'picture':
                            
                            if sheetitems.cell(row=rowitem, column=3).value:
                                if sheetitems.cell(row=rowitem, column=4).value:
                                    sheetitems.cell(row=rowitem, column=4).value = sheetitems.cell(row=rowitem, column=4).value + ';' + L4.text
                                else:
                                    sheetitems.cell(row=rowitem, column=4).value = L4.text
                            else:
                                sheetitems.cell(row=rowitem, column=3).value = L4.text
                            
                        elif L4.tag == 'param':
                        
                            name = L4.get('name')
                            value = L4.text
                            paramstr = str(name + ': ' + value + '\n')
                            
                            if sheetitems.cell(row=rowitem, column=5).value:
                                sheetitems.cell(row=rowitem, column=5).value = str(sheetitems.cell(row=rowitem, column=5).value + paramstr)
                            else:
                                sheetitems.cell(row=rowitem, column=5).value = paramstr
                                

                        
                        
                        else:
                            sheetitems.cell(row=rowitem, column=itemsfieldnames.index(L4.tag)+1).value = L4.text
                            
                            #columnitem=columnitem+1
    wb.save(wbpath)
    messagebox.showinfo("Внимание !", "Выполнено !")
    os.system('"{0}"'.format(wbpath))






def DecodeTXTsf():
    SelectFile()
    if IsDecodeFileSelected:
        DecodeTXT()


def DecodeTXT():
    global DecodeFile
    global IsDecodeFileSelected
    
    DecodeFile = Path(DecodeFile)
    revisedreqfile = DecodeFile.as_posix()
    
    tree = ET.parse(revisedreqfile)
    root = tree.getroot()
    
    
    ResultTxt = Path(Path(DecodeFile).parent, 'DECODED {0}.txt'.format(Path(DecodeFile).name))

    with open(ResultTxt, 'w') as f:
        ## LAYER 1
        for L1 in root:
            try:
                textrow = "|--> L1: {0} || {1} || {2}".format(L1.tag, L1.attrib, L1.text)
                f.write(textrow + '\n')
                print(textrow)
            except:
                textrow = "|--> L1: DECODE ERROR"
                f.write(textrow + '\n')
                print(textrow)
            ## LAYER 2
            for L2 in L1:
                try:
                    textrow = "|-----> L2: {0} || {1} || {2}".format(L2.tag, L2.attrib, L2.text)
                    f.write(textrow + '\n')
                    print(textrow)
                except:
                    textrow = "|-----> L2: DECODE ERROR"
                    f.write(textrow + '\n')
                    print(textrow)
                ## LAYER 3
                for L3 in L2:
                    try:
                        textrow = "|--------> L3: {0} || {1} || {2}".format(L3.tag, L3.attrib, L3.text)
                        f.write(textrow + '\n')
                        print(textrow)
                    except:
                        textrow = "|--------> L3: DECODE ERROR"
                        f.write(textrow + '\n')
                        print(textrow)
                    ## LAYER 4
                    for L4 in L3:
                        try:
                            textrow = "|-----------> L4: {0} || {1} || {2}".format(L4.tag, L4.attrib, L4.text)
                            f.write(textrow + '\n')
                            print(textrow)
                        except:
                            textrow = "|-----------> L4: DECODE ERROR"
                            f.write(textrow + '\n')
                            print(textrow)
    os.system('"{0}"'.format(ResultTxt))






def SelectFile():
    global DecodeFile
    global IsDecodeFileSelected

    DecodeFile = filedialog.askopenfilename(title='Выберите файл обмена', filetypes=(('XML', 'xml'),))
    if DecodeFile:
        print('SF: DecodeFile :', DecodeFile)
        
        IsDecodeFileSelected = True
    else:
        print('SF: DecodeFile not selected')






def Makel25Xml():
    reqfilepath = Path(dir_path, 'makel25.xml')
    revisedreqfile = reqfilepath.as_posix()
    
    tree = ET.parse(revisedreqfile)
    root = tree.getroot()
    
    wbpath = Path(dir_path, 'makel25.xlsx')
    wb = Workbook()
    sheetitems = wb.active
    sheetitems.title = "Items"
    sheetcategory = wb.create_sheet("Category")
    
    #                       1               2           3         4          5           6               7           8          9           10
    itemsfieldnames = ['description', 'model', 'vendor', 'picture', 'more_ph', 'categoryId', 'category_path', 'url', 'price mitc', 'available']
    sheetitems.append(itemsfieldnames)
    
    categoryfieldnames = ['category', 'id', 'parentId']
    sheetcategory.append(categoryfieldnames)
    
    rowitem = 1
    rowcategory = 1
    categorypath = []
    for L1 in root:
        for L2 in L1:
            for L3 in L2:
                
                if L3.tag == 'category':
                
                    rowcategory = rowcategory + 1
                    
                    catname = L3.text
                    sheetcategory.cell(row=rowcategory, column=1).value = catname
                    
                    catid = L3.get('id')
                    sheetcategory.cell(row=rowcategory, column=2).value = catid
                    
                    catparentid = L3.get('parentId')
                    sheetcategory.cell(row=rowcategory, column=3).value = catparentid
                    
                    categorypath.append(catname)
                    if sheetcategory.cell(row=rowcategory, column=3).value:
                        for a in range(1, 700):
                            if sheetcategory.cell(row=rowcategory, column=3).value == sheetcategory.cell(row=a, column=2).value:
                                categorypath.append(sheetcategory.cell(row=a, column=1).value)
                                
                                if sheetcategory.cell(row=a, column=3).value:
                                    for b in range(1, 700):
                                        if sheetcategory.cell(row=a, column=3).value == sheetcategory.cell(row=b, column=2).value:
                                            categorypath.append(sheetcategory.cell(row=b, column=1).value)
                                            
                                            if sheetcategory.cell(row=b, column=3).value:
                                                for c in range(1, 700):
                                                    if sheetcategory.cell(row=b, column=3).value == sheetcategory.cell(row=c, column=2).value:
                                                        categorypath.append(sheetcategory.cell(row=c, column=1).value)
                                                        
                                                        if sheetcategory.cell(row=c, column=3).value:
                                                            for d in range(1, 700):
                                                                if sheetcategory.cell(row=c, column=3).value == sheetcategory.cell(row=d, column=2).value:
                                                                    categorypath.append(sheetcategory.cell(row=d, column=1).value)
                                                                    
                                                                    if sheetcategory.cell(row=d, column=3).value:
                                                                        for e in range(1, 700):
                                                                            if sheetcategory.cell(row=d, column=3).value == sheetcategory.cell(row=e, column=2).value:
                                                                                categorypath.append(sheetcategory.cell(row=e, column=1).value)
                                                                                
                                                                                if sheetcategory.cell(row=e, column=3).value:
                                                                                    for f in range(1, 700):
                                                                                        if sheetcategory.cell(row=e, column=3).value == sheetcategory.cell(row=f, column=2).value:
                                                                                            categorypath.append(sheetcategory.cell(row=f, column=1).value)
                    
                    categorypath.reverse()
                    sheetcategory.cell(row=rowcategory, column=4).value = "|".join(categorypath)
                    categorypath = []
                    
                if L3.tag == 'offer':
                    if Decode: print("**** Item : ", L3.attrib)
                    if Decode: print("")
                    
                    rowitem = rowitem + 1
                    
                    #                        1               2           3         4          5           6               7           8          9           10
                    #itemsfieldnames = ['description', 'model', 'vendor', 'picture', 'more_ph', 'categoryId', 'category_path', 'url', 'price mitc', 'available']
    
                    name = L3.attrib.get('available')
                    sheetitems.cell(row=rowitem, column=10).value = name
                    
                    for L4 in L3:
                        if L4.tag == 'name':
                            sheetitems.cell(row=rowitem, column=1).value=L4.text
                            
                        if L4.tag == 'url':
                            sheetitems.cell(row=rowitem, column=8).value=L4.text
                            
                        if L4.tag == 'categoryId':
                            sheetitems.cell(row=rowitem, column=6).value=L4.text
                            for i in range(1, 700):
                                if sheetitems.cell(row=rowitem, column=6).value == sheetcategory.cell(row=i, column=2).value:
                                    sheetitems.cell(row=rowitem, column=7).value = sheetcategory.cell(row=i, column=4).value
                            
                        if L4.tag == 'picture':
                            if sheetitems.cell(row=rowitem, column=4).value:
                                if sheetitems.cell(row=rowitem, column=5).value:
                                    sheetitems.cell(row=rowitem, column=5).value = sheetitems.cell(row=rowitem, column=5).value + ';' + L4.text
                                else:
                                    sheetitems.cell(row=rowitem, column=5).value = L4.text
                            else:
                                sheetitems.cell(row=rowitem, column=4).value = L4.text
                            
                        if L4.tag == 'model':
                            sheetitems.cell(row=rowitem, column=2).value = L4.text
    
    wb.save(wbpath)
    
def FeronXml():
    reqfilepath = Path(dir_path, 'feron.xml')
    revisedreqfile = reqfilepath.as_posix()
    
    tree = ET.parse(revisedreqfile)
    root = tree.getroot()
    
    wbpath = Path(dir_path, 'feron.xlsx')
    wb = Workbook()
    sheetitems = wb.active
    sheetitems.title = "Items"
    sheetcategory = wb.create_sheet("Category")
    
    ResultTxt = Path(dir_path, 'DECODED FERON.txt', encoding="utf-8")
    txtfile = open(ResultTxt, "w")
    
    #                       1            2         3         4          5           6               7           8          9           10           11
    itemsfieldnames = ['description', 'model', 'vendor', 'picture', 'more_ph', 'categoryId', 'category_path', 'url', 'price mitc', 'available', 'article']
    sheetitems.append(itemsfieldnames)
    
    categoryfieldnames = ['category', 'id', 'parentId']
    sheetcategory.append(categoryfieldnames)
    
    rowitem = 1
    rowcategory = 1
    categorypath = []
    for L1 in root:
        for L2 in L1:
            for L3 in L2:
                
                if L3.tag == 'category':
                
                    rowcategory = rowcategory + 1
                    
                    catname = L3.text
                    sheetcategory.cell(row=rowcategory, column=1).value = catname
                    
                    catid = L3.get('id')
                    sheetcategory.cell(row=rowcategory, column=2).value = catid
                    
                    catparentid = L3.get('parentId')
                    sheetcategory.cell(row=rowcategory, column=3).value = catparentid
                    
                    categorypath.append(catname)
                    if sheetcategory.cell(row=rowcategory, column=3).value:
                        for a in range(1, 700):
                            if sheetcategory.cell(row=rowcategory, column=3).value == sheetcategory.cell(row=a, column=2).value:
                                categorypath.append(sheetcategory.cell(row=a, column=1).value)
                                
                                if sheetcategory.cell(row=a, column=3).value:
                                    for b in range(1, 700):
                                        if sheetcategory.cell(row=a, column=3).value == sheetcategory.cell(row=b, column=2).value:
                                            categorypath.append(sheetcategory.cell(row=b, column=1).value)
                                            
                                            if sheetcategory.cell(row=b, column=3).value:
                                                for c in range(1, 700):
                                                    if sheetcategory.cell(row=b, column=3).value == sheetcategory.cell(row=c, column=2).value:
                                                        categorypath.append(sheetcategory.cell(row=c, column=1).value)
                                                        
                                                        if sheetcategory.cell(row=c, column=3).value:
                                                            for d in range(1, 700):
                                                                if sheetcategory.cell(row=c, column=3).value == sheetcategory.cell(row=d, column=2).value:
                                                                    categorypath.append(sheetcategory.cell(row=d, column=1).value)
                                                                    
                                                                    if sheetcategory.cell(row=d, column=3).value:
                                                                        for e in range(1, 700):
                                                                            if sheetcategory.cell(row=d, column=3).value == sheetcategory.cell(row=e, column=2).value:
                                                                                categorypath.append(sheetcategory.cell(row=e, column=1).value)
                                                                                
                                                                                if sheetcategory.cell(row=e, column=3).value:
                                                                                    for f in range(1, 700):
                                                                                        if sheetcategory.cell(row=e, column=3).value == sheetcategory.cell(row=f, column=2).value:
                                                                                            categorypath.append(sheetcategory.cell(row=f, column=1).value)
                    
                    categorypath.reverse()
                    sheetcategory.cell(row=rowcategory, column=4).value = "|".join(categorypath)
                    categorypath = []
                    
                if L3.tag == 'offer':
                    print("** Item : " + str(L3.attrib))
                    textrow = str("**** Item : " + str(L3.attrib))
                    try:
                        txtfile.write(textrow + '\n')
                    except:
                        txtfile.write("ERROR!!!!!!! Item" + '\n')
                    
                    rowitem = rowitem + 1
                    
    #                       1            2         3         4          5           6               7           8          9           10           11
    #itemsfieldnames = ['description', 'model', 'vendor', 'picture', 'more_ph', 'categoryId', 'category_path', 'url', 'price mitc', 'available', 'article']
    
                    name = L3.attrib.get('available')
                    sheetitems.cell(row=rowitem, column=10).value = name
                    
                    for L4 in L3:
                        if L4.tag == 'url':
                            sheetitems.cell(row=rowitem, column=8).value=L4.text
                            textrow = str("========= url : " + L4.text)
                            try:
                                txtfile.write(textrow + '\n')
                            except:
                                txtfile.write("ERROR!!!!!!!" + '\n')
                            
                        if L4.tag == 'categoryId':
                            sheetitems.cell(row=rowitem, column=6).value=L4.text
                            for i in range(1, 700):
                                if sheetitems.cell(row=rowitem, column=6).value == sheetcategory.cell(row=i, column=2).value:
                                    sheetitems.cell(row=rowitem, column=7).value = sheetcategory.cell(row=i, column=4).value
                            
                        if L4.tag == 'picture':
                            if sheetitems.cell(row=rowitem, column=4).value:
                                if sheetitems.cell(row=rowitem, column=5).value:
                                    sheetitems.cell(row=rowitem, column=5).value = sheetitems.cell(row=rowitem, column=5).value + ';' + L4.text
                                else:
                                    sheetitems.cell(row=rowitem, column=5).value = L4.text
                            else:
                                sheetitems.cell(row=rowitem, column=4).value = L4.text
                            
                        if L4.tag == 'vendor':
                            sheetitems.cell(row=rowitem, column=3).value = L4.text
                            textrow = str("========= vendor : " + L4.text)
                            try:
                                txtfile.write(textrow + '\n')
                            except:
                                txtfile.write("ERROR!!!!!!! vendor" + '\n')
                            
                        if L4.tag == 'model':
                            sheetitems.cell(row=rowitem, column=11).value = L4.text
                            textrow = str("========= model : " + L4.text)
                            try:
                                txtfile.write(textrow + '\n')
                            except:
                                txtfile.write("ERROR!!!!!!! model" + '\n')
                            
                        if L4.tag == 'vendorCode':
                            sheetitems.cell(row=rowitem, column=2).value = L4.text
                            textrow = str("========= vendorCode : " + L4.text)
                            try:
                                txtfile.write(textrow + '\n')
                            except:
                                txtfile.write("ERROR!!!!!!! vendorCode" + '\n')
                            
                        if L4.tag == 'description':
                            sheetitems.cell(row=rowitem, column=1).value = L4.text
                            textrow = str("========= description : " + L4.text)
                            try:
                                txtfile.write(textrow + '\n')
                            except:
                                txtfile.write("ERROR!!!!!!! description" + '\n')
                            
                        if L4.tag == 'param':
                            name = L4.get('name')
                            if name == 'МИЦ (мин. интернет-цена)':
                                sheetitems.cell(row=rowitem, column=9).value = L4.text
                                textrow = str("========= МИЦ : " + L4.text)
                            try:
                                txtfile.write(textrow + '\n')
                            except:
                                txtfile.write("ERROR!!!!!!! МИЦ" + '\n')
    
    wb.save(wbpath)
    txtfile.close()
    

if __name__ == '__main__':
    main()
