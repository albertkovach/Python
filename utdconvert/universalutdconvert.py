from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

import requests, os, sys, csv, re
from pathlib import Path

import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import xml.etree.ElementTree as ET

global root
global scrnwparam
global scrnhparam
scrnwparam = 240
scrnhparam = 195

global QuantityIsInThousands


def main():
    global root

    root = Tk()
    root.resizable(False, False)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('370x160+{}+{}'.format(scrnw, scrnh))
        
    app = GUI(root)
    root.mainloop()
    

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("Переделка УПД под формат УТ, v2.0")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):
        #global IsInputSel
        IsInputSel = False

        global FindInternalCodes
        FindInternalCodes = False
        
        #global InputFile
        #global extension

        
        global UniversalBtn
        UniversalBtn = Button(text='Сделать все одной кнопкой', command=RunUniversal, font=("Arial", 14))
        UniversalBtn.place(x=30, y=30, width=310, height=85)


        try:
            InputFile = sys.argv[1]
            #extension = Path(InputFile).suffix
            print(InputFile)
            print("Файл выбран "+InputFile) 
            Universal(InputFile)
        except:
            print("Файл не выбран") 




def SelectFile(extension):
    #global InputFile
    #global IsInputSel
    extname = extension.upper()

    InputFile = filedialog.askopenfilename(title='Выберите УПД', filetypes=((extname, extension),))
    
    if InputFile:
        print('Файл выбран :' + InputFile)
        
        #print(str(IsInputSel))
        return InputFile
    else:
        print('Файл не выбран')
        return False




def IsInCell(thing,row,column,sheet):
    if str(sheet.cell(row, column).value) == thing:
        return True
    else:
        return False


def AdressInRow(things,row,sheet):
    for y in range(1,sheet.max_column):
        if str(sheet.cell(row, column=y).value) in things:
            return y
    return -1


def FindAllInstancesOnSheet(thing,sheet):
    locations = [-1]
    for x in range(1,sheet.max_row):
        for y in range(1,sheet.max_column):
            if thing in str(sheet.cell(row=x, column=y).value): 
                locations.append([x,y])
    return locations


def NextNotEmptyInRowAdress(row,startcolumn,sheet):
    for y in range(startcolumn+1,sheet.max_column):
        if (sheet.cell(row, column=y).value)!=None:
            return y
    return -1


def SaveAsXLSX(InputFile):
    encod = False
    book_xls = xlrd.open_workbook(InputFile)
    book_xlsx = Workbook()
    if str(format(book_xls.encoding)) == "iso-8859-1":
        print(str(format(book_xls.encoding))+" Декодирую и сохраняю как xlsx")
        encod = True
    else:
        print(" Сохраняю без декодирования как xlsx")
        encod = False
    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
    if sheet_index == 0:
        sheet_xlsx = book_xlsx.active
        sheet_xlsx.title = sheet_name
    else:
        sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

    for row in range(0, sheet_xls.nrows):
        for col in range(0, sheet_xls.ncols):
            if encod == True:
                sheet_xlsx.cell(row = row+1 , column = col+1).value = Decode(str(sheet_xls.cell_value(row, col)))
            else:
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
    savepath = Path(Path(InputFile).parent, str(Path(InputFile).stem+" Перекодированный.xlsx"))
    #print("Сохранил в " + savepath)
    book_xlsx.save(savepath)
    return savepath



def RunUniversal():
    InputFile = SelectFile('.xlsx .xls')
    #print(InputFile)    
    Universal(InputFile)



def Universal(InputFile):
    extension = Path(InputFile).suffix
    QuantityIsInThousands = True
    

    if extension == '.xls':
        savepath = SaveAsXLSX(InputFile)
        book = openpyxl.load_workbook(savepath)
        
    else:
        book = openpyxl.load_workbook(InputFile)
    sheet = book.active
    
    expfilename = Path(InputFile).stem + " Результат.xlsx"
    expfilepath = Path(InputFile).parent
    expfilepath = Path(expfilepath, expfilename)
    #print(expfilepath)
    
    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Приход"
    updnumlocs = FindAllInstancesOnSheet("Счет-фактура №",sheet)
    
    
    sellertext = ""
    seller = FindAllInstancesOnSheet("Продавец:",sheet)
    
    INNtext = ""
    INN = FindAllInstancesOnSheet("ИНН/КПП продавца:",sheet)
    #if len(seller)!=1:
        
    
    #print(updnumlocs)
    exprowcounter = 0 
    updcounter = len(updnumlocs)-1
    print("Найдено УПД в файле:")
    print(len(updnumlocs)-1)
    
    for upd in updnumlocs:
        
        if upd!=-1:
            #print(upd[1])
            #print(SearchNextNotEmptyToRight(upd[0],upd[1],sheet))
            UTDnumber = ""
            if NextNotEmptyInRowAdress(upd[0],upd[1],sheet)!=-1:
                UTDnumber = sheet.cell(upd[0],NextNotEmptyInRowAdress(upd[0],upd[1],sheet)).value
            DAte = ""
            if AdressInRow("от",upd[0],sheet)!=-1:
                if NextNotEmptyInRowAdress(upd[0],AdressInRow("от",upd[0],sheet),sheet)!=-1:
                    DAte = sheet.cell(upd[0],NextNotEmptyInRowAdress(upd[0],AdressInRow("от",upd[0],sheet),sheet)).value
            
            sellercolumn = NextNotEmptyInRowAdress(seller[len(updnumlocs)-updcounter][0],seller[len(updnumlocs)-updcounter][1],sheet)
            if sellercolumn!=-1:    
                sellertext = sheet.cell(seller[len(updnumlocs)-updcounter][0],sellercolumn).value
            
            INNcolumn = NextNotEmptyInRowAdress(INN[len(updnumlocs)-updcounter][0],INN[len(updnumlocs)-updcounter][1],sheet)
            if INNcolumn!=-1:    
                INNtext = sheet.cell(INN[len(updnumlocs)-updcounter][0],INNcolumn).value

            sellernametext = sheet.cell(seller[len(updnumlocs)-updcounter][0],seller[len(updnumlocs)-updcounter][1]).value
            INNnametext = sheet.cell(INN[len(updnumlocs)-updcounter][0],INN[len(updnumlocs)-updcounter][1]).value
            

            expsheet.append([sheet.cell(upd[0],upd[1]).value,UTDnumber,"от",DAte,sellernametext,sellertext,INNnametext,INNtext])
            exprowcounter=exprowcounter+1
        #                 1         2        3            4              5          6        7       8         9                10
            fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
            exprowcounter=exprowcounter+1
            expsheet.append(fieldnames)
            
            modelcolumn = 1
            quantitycolumn = 1
            pricecolumn = 1
            overallpricecolumn = 1
            gtdnumcolumn = 1
            countrycolumn = 1   
            
            counter = 0
            trigger = False
            startofthenextupd = sheet.max_row
            if updcounter>1:
                startofthenextupd = updnumlocs[updnumlocs.index(upd)+1][0]
                updcounter=updcounter-1
            
            for i in range(upd[0], startofthenextupd+1):

                if trigger:
                        if has_numbers(str(sheet.cell(row=i, column=pricecolumn).value)) and not IsInCell(11,i,gtdnumcolumn,sheet) and not IsInCell("11",i,gtdnumcolumn,sheet) and not IsInCell("А",i,modelcolumn,sheet):
                            exprowcounter = exprowcounter + 1
                            
                            model =         str(sheet.cell(row=i, column=modelcolumn).value)
                            quantity =      str(sheet.cell(row=i, column=quantitycolumn).value)
                            quantity = quantity.replace(" ", "")
                            overallprice =  str(sheet.cell(row=i, column=overallpricecolumn).value)
                            gtdnum =        str(sheet.cell(row=i, column=gtdnumcolumn).value)
                            country =       str(sheet.cell(row=i, column=countrycolumn).value)
                            overallprice = overallprice.replace("-", ".")
                            overallprice = overallprice.replace(" ", "")
                            price =         str(float(overallprice)/float(quantity))
                            #price =         price.replace(" ", "")
                            #overallprice = overallprice.replace(".", ",")
                            
                            
                            if not has_numbers(str(gtdnum)):
                                gtdnum = ""
                            if not has_letters(str(country)):
                                country = ""
                            if has_letters(str(country)) and not has_numbers(str(gtdnum)):
                                gtdnum = "-"
                            price = price.replace("-", ".")
                            price = price.replace(" ", "")

                            quantity = quantity.replace(" ", "")
                            gtdnum = gtdnum.replace(" ","")
                            
                            if float(quantity)%1000!=0:
                                QuantityIsInThousands = False
                            
                            expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                            expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                            expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                            expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                            expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                            expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                        else:
                            counter = counter + 1
                #print(str(sheet.cell(i, modelcolumn+1).value))
                
                elif AdressInRow(["A","А"],i,sheet)!=-1:

                    modelcolumn = AdressInRow(["A","А"],i,sheet)
                    if AdressInRow(["Б","б"],i,sheet)==NextNotEmptyInRowAdress(i,modelcolumn,sheet):
                        modelcolumn = AdressInRow(["Б","б"],i,sheet)
                        
                    if IsInCell("1",i+1,modelcolumn,sheet) or IsInCell(1,i+1,modelcolumn,sheet):
                        modelcolumn = NextNotEmptyInRowAdress(i,modelcolumn,sheet) 
                        
                    quantitycolumn = AdressInRow(["3"],i,sheet)
                    pricecolumn = AdressInRow(["4"],i,sheet)
                    overallpricecolumn = AdressInRow(["9"],i,sheet)
                    gtdnumcolumn = AdressInRow(["11"],i,sheet)
                    countrycolumn = AdressInRow(["10a","10а"],i,sheet)
                    temp = AdressInRow(["10","10"],i,sheet)
                    
                    
                    if sheet.cell(i,NextNotEmptyInRowAdress(i,temp,sheet)).value==10:
                        countrycolumn = NextNotEmptyInRowAdress(i,temp,sheet)
                    #print(countrycolumn)

                    trigger = True
        
    #QuantityIsInThousands = True
    if len(FindAllInstancesOnSheet("Лезард",sheet)) == 1 and len(FindAllInstancesOnSheet("ЭНЕРГОИМПУЛЬС",sheet)) ==1:
        QuantityIsInThousands = False


    if QuantityIsInThousands:
        for xy in range(3,expsheet.max_row+1):
            expsheet.cell(xy,5).value = int(expsheet.cell(xy,5).value)/1000


    SearchCodesFromExcel(expsheet, 3,2)
    

    print("Начинаю сохранять результат")
    try:
        expfile.save(expfilepath)
    except:
        msg_box = messagebox.askquestion('Подтверждение', 'Экспортный файл уже открыт, закройте и попробуйте еще раз')
        if msg_box == 'yes':
            expfile.save(expfilepath)
    print("Открываю файл с результатами...")
    os.system('"{0}"'.format(expfilepath))




def Decode(data):
    decoded = data.encode('CP1252', errors='ignore').decode('CP1251', errors='ignore')
    return decoded


def enc_test():
    SelectFile('xls')
    
    if IsInputSel:
        book = xlrd.open_workbook(InputFile)
        sheet = book.sheet_by_index(0)
        encoding = book.encoding
        print('***** {0}'.format(book.encoding))
        
        for i in range(0, 5):
            for k in range(0, 5):
                print('r={0},col={1} // res={2} //'.format(i,k,sheet.cell(i,k).value))
        
        text = sheet.cell(3,8).value
        print("#### after: ".format(text))
        decoded = text.encode('CP1252').decode('CP1251', errors='ignore')
        print("#### before: ".format(decoded))



def has_numbers(inputString):
    #number_pattern = re.compile(r'-?\d+(\.\d+)?')
    #numbers = number_pattern.findall(inputString)
    #return numbers
    return bool(re.search(r'\d', inputString))
    
def has_letters(inputString):
    cyrillic_pattern = re.compile('[а-яА-ЯёЁ]')
    contains_cyrillic = bool(cyrillic_pattern.search(inputString))
    return contains_cyrillic
    
def SearchCodesFromExcel(mainsheet, inputcolumn,outputcolumn):
    msg_box = messagebox.askquestion('Подтверждение', 'Выбрать файл расшифровки артикулов ?')
    if msg_box == 'yes':
        databookpath = filedialog.askopenfilename(title='Выберите таблицу с кодами', filetypes=(('.XLS', '.xls'),))
        
        if databookpath:

            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            print("Начинаю поиск кодов")
            Expmassive = []
            Datamassive =[]
            for i in range(1, mainsheet.max_row+1):
                Expmassive.append(str(mainsheet.cell(i, inputcolumn).value))
                #print (Expmassive)
            for k in range(0,datasheet.nrows):
                Datamassive.append(str(datasheet.cell(k, 1).value))
            for i in range(1, mainsheet.max_row+1):
                if (i%100==0):
                    print("Пройдено строк "+ str(i))
                First_found = False
                First_k = 0
                Second_found = False
                for k in range(0,datasheet.nrows):
                    if Expmassive[i-1]== Datamassive[k]:
                        if First_found == False:
                            First_found = True
                            First_k=k
                        else: 
                            Second_found = True
                            break
                if First_found and not Second_found:
                    mainsheet.cell(i, outputcolumn).value = datasheet.cell(First_k,0).value
                elif Second_found:
                    mainsheet.cell(i, outputcolumn).value = "ВНИМАНИЕ!!! КОПИЯ" + datasheet.cell(First_k,0).value

if __name__ == '__main__':
    main()
