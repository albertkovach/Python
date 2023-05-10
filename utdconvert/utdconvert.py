from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

import requests, os, sys, csv, re
from pathlib import Path

import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import xml.etree.ElementTree as ET

global root
global scrnwparam
global scrnhparam
scrnwparam = 240
scrnhparam = 195

def main():
    global root

    root = Tk()
    root.resizable(False, False)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('480x390+{}+{}'.format(scrnw, scrnh))
        
    app = GUI(root)
    root.mainloop()
    

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("Переделка УПД под формат УТ, v1.3")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):
        global IsInputSel
        IsInputSel = False
        
        global FindInternalCodes
        FindInternalCodes = False
        

        global AllurBtn
        AllurBtn = Button(text='Аллюр', command=Allur, font=("Arial", 14))
        AllurBtn.place(x=30, y=15, width=180, height=35)
        
        global TDMBtn
        TDMBtn = Button(text='TDM', command=TDM, font=("Arial", 14))
        TDMBtn.place(x=30, y=60, width=180, height=35) # y + 45
        
        global TrelvanBtn
        TrelvanBtn = Button(text='Трелван', command=Trelvan, font=("Arial", 14))
        TrelvanBtn.place(x=30, y=105, width=180, height=35)
        
        global TransEnBtn
        TransEnBtn = Button(text='ТрансЭнергетик', command=TransEn, font=("Arial", 14))
        TransEnBtn.place(x=30, y=150, width=180, height=35)
        
        global TransEnMultBtn
        TransEnMultBtn = Button(text='ТрЭн Склеенный', command=TransEnMult, font=("Arial", 14))
        TransEnMultBtn.place(x=30, y=195, width=180, height=35)
        
        global KartaSvetaBtn
        KartaSvetaBtn = Button(text='КартаСвета', command=KartaSveta, font=("Arial", 14))
        KartaSvetaBtn.place(x=30, y=240, width=180, height=35)
        
        global GreenelBtn
        GreenelBtn = Button(text='Greenel', command=Greenel, font=("Arial", 14))
        GreenelBtn.place(x=30, y=285, width=180, height=35)
        
        global SDSBtn
        SDSBtn = Button(text='СДС', command=SDSv2, font=("Arial", 14))
        SDSBtn.place(x=30, y=330, width=180, height=35)
        
        
        
        
        global SonelBtn
        SonelBtn = Button(text='Сонэл', command=Sonel, font=("Arial", 14))
        SonelBtn.place(x=270, y=15, width=180, height=35)
        
        global LezardBtn
        LezardBtn = Button(text='Lezard/ЭнИмп', command=Lezard, font=("Arial", 14))
        LezardBtn.place(x=270, y=60, width=180, height=35)

        global LezardMultiBtn
        LezardMultiBtn = Button(text='Lezard Склеенный', command=LezardMult, font=("Arial", 14))
        LezardMultiBtn.place(x=270, y=105, width=180, height=35)
        
        global EraBtn
        EraBtn = Button(text='ВипМаркет', command=Era, font=("Arial", 14))
        EraBtn.place(x=270, y=150, width=180, height=35)
        
        global BironiBtn
        BironiBtn = Button(text='ПК Ретро', command=Bironi, font=("Arial", 14))
        BironiBtn.place(x=270, y=195, width=180, height=35)
        
        global AbBatteriesBtn
        AbBatteriesBtn = Button(text='АБ Бэттэрис', command=AbBatteries, font=("Arial", 14))
        AbBatteriesBtn.place(x=270, y=240, width=180, height=35)

        global ArsiLightBtn
        ArsiLightBtn = Button(text='Арсилайт/Белсвет', command=ArsiLight, font=("Arial", 14))
        ArsiLightBtn.place(x=270, y=285, width=180, height=35)
        
        global GaussBtn
        GaussBtn = Button(text='Gauss', command=Gauss, font=("Arial", 14))
        GaussBtn.place(x=270, y=330, width=180, height=35)
        
        


def SelectFile(extension):
    global InputFile
    global IsInputSel
    extname = extension.upper()

    InputFile = filedialog.askopenfilename(title='Выберите УПД', filetypes=((extname, extension),))
    
    if InputFile:
        print('SF: InputFile :', InputFile)
        
        IsInputSel = True
    else:
        print('SF: InputFile not selected')






def Allur():
    SelectFile('xls')
    
    if IsInputSel:
    
        book = xlrd.open_workbook(InputFile)
        sheet = book.sheet_by_index(0)
        
        expfilename = "Аллюр № {0} от {1}.xlsx".format(sheet.cell(0,15).value, sheet.cell(0,24).value).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10,sheet.nrows):
            
            modelrow = 1
            celltype = sheet.cell_type(i, modelrow)

            if trigger:
                if celltype == xlrd.XL_CELL_EMPTY:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(i,modelrow).value
                        quantity =      sheet.cell(i,26).value
                        price =         sheet.cell(i,29).value
                        overallprice =  sheet.cell(i,57).value
                        gtdnum =        sheet.cell(i,68).value
                        country =       sheet.cell(i,66).value
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(i,modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, sheet.cell(i,modelrow).value, trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def TDM():
    SelectFile('xls')
    
    if IsInputSel:
    
        book = xlrd.open_workbook(InputFile)
        sheet = book.sheet_by_index(0)
        
        expfilename = "ТДМ № {0} от {1}.xlsx".format(sheet.cell(2,11).value, sheet.cell(2,18).value).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10,sheet.nrows):
        
            modelrow = 1
            celltype = sheet.cell_type(i, modelrow)

            if trigger:
                if celltype == xlrd.XL_CELL_EMPTY or "№" in sheet.cell(i,modelrow).value:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(i,modelrow).value
                        quantity =      sheet.cell(i,27).value
                        price =         sheet.cell(i,32).value
                        overallprice =  sheet.cell(i,49).value
                        gtdnum =        sheet.cell(i,65).value
                        country =       sheet.cell(i,53).value
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(i,modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, sheet.cell(i,modelrow).value, trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def Trelvan():
    SelectFile('xls')
    
    if IsInputSel:
    
        book = xlrd.open_workbook(InputFile)
        sheet = book.sheet_by_index(0)
        
        expfilename = "Трелван № {0} от {1}.xlsx".format(sheet.cell(0,21).value, sheet.cell(0,30).value).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10,sheet.nrows):
        
            modelrow = 1
            celltype = sheet.cell_type(i, modelrow)

            if trigger:
                if celltype == xlrd.XL_CELL_EMPTY or "№" in sheet.cell(i,modelrow).value:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(i,modelrow).value
                        quantity =      sheet.cell(i,35).value
                        price =         sheet.cell(i,37).value
                        overallprice =  sheet.cell(i,65).value
                        gtdnum =        sheet.cell(i,74).value
                        country =       sheet.cell(i,72).value
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(i,modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, sheet.cell(i,modelrow).value, trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def TransEn():
    SelectFile('xls')
    
    if IsInputSel:
    
        book = xlrd.open_workbook(InputFile, encoding_override="iso-8859-1")
        sheet = book.sheet_by_index(0)
        
        expfilename = "ТрансЭнергетик № {0} от {1}.xlsx".format(Decode(str(sheet.cell(1,30).value)), Decode(str(sheet.cell(1,45).value))).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10,sheet.nrows):
        
            modelrow = 1
            celltype = sheet.cell_type(i, modelrow)

            if trigger:
                if celltype == xlrd.XL_CELL_EMPTY or "," in sheet.cell(i,modelrow).value:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(i,modelrow).value
                        quantity =      sheet.cell(i,58).value
                        price =         sheet.cell(i,65).value
                        overallprice =  sheet.cell(i,113).value
                        gtdnum =        sheet.cell(i,141).value
                        country =       sheet.cell(i,130).value
                        
                        model = Decode(str(model))
                        gtdnum = Decode(str(gtdnum))
                        country = Decode(str(country))
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if Decode(str(sheet.cell(i,modelrow).value)) == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, Decode(str(sheet.cell(i,modelrow).value)), trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def TransEnMult():
    SelectFile('xls')
    
    if IsInputSel:
    
        book = xlrd.open_workbook(InputFile, encoding_override="iso-8859-1")
        sheet = book.sheet_by_index(0)
        
        expfilename = "ТрансЭнергетик склеенный от {0}.xlsx".format(Decode(str(sheet.cell(1,45).value))).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        
        partname = str("№ {0} от {1}".format(Decode(str(sheet.cell(1,30).value)), Decode(str(sheet.cell(1,45).value))))
        expsheet.cell(row=1, column=2).value = partname
        expsheet.cell(row=1, column=2).font = Font(size=12, bold=True)
        
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 2
        for i in range(10,sheet.nrows):
            
            partnamecelltype = sheet.cell_type(i, 30)
            if partnamecelltype != xlrd.XL_CELL_EMPTY:
                if has_letters(Decode(str(sheet.cell(i, 30).value))): 
                    exprowcounter = exprowcounter + 2
                    
                    partname = str("№ {0} от {1}".format(Decode(str(sheet.cell(i,30).value)), Decode(str(sheet.cell(i,45).value))))
                    expsheet.cell(row=exprowcounter, column=2).value = partname
                    expsheet.cell(row=exprowcounter, column=2).font = Font(size=12, bold=True)
                    expsheet.append(fieldnames)
                    
                    exprowcounter = exprowcounter + 1
                
        
            modelrow = 6
            celltype = sheet.cell_type(i, modelrow)

            if trigger:
                if celltype == xlrd.XL_CELL_EMPTY or "," in sheet.cell(i,modelrow).value:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(i,modelrow).value
                        quantity =      sheet.cell(i,58).value
                        price =         sheet.cell(i,65).value
                        overallprice =  sheet.cell(i,113).value
                        gtdnum =        sheet.cell(i,141).value
                        country =       sheet.cell(i,130).value
                        
                        model = Decode(str(model))
                        gtdnum = Decode(str(gtdnum))
                        country = Decode(str(country))
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if Decode(str(sheet.cell(i,modelrow).value)) == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, Decode(str(sheet.cell(i,modelrow).value)), trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def KartaSveta():
    SelectFile('xls')
    
    if IsInputSel:
    
        book = xlrd.open_workbook(InputFile)
        sheet = book.sheet_by_index(0)
        
        expfilename = "КартаСвета № {0} от {1}.xlsx".format(sheet.cell(0,21).value, sheet.cell(0,30).value).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10,sheet.nrows):
        
            modelrow = 1
            celltype = sheet.cell_type(i, modelrow)

            if trigger:
                if celltype == xlrd.XL_CELL_EMPTY or "№" in sheet.cell(i,modelrow).value:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(i, modelrow).value
                        quantity =      sheet.cell(i, 35).value
                        price =         sheet.cell(i, 37).value
                        overallprice =  sheet.cell(i, 65).value
                        gtdnum =        sheet.cell(i, 74).value
                        country =       sheet.cell(i, 72).value
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(i,modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, sheet.cell(i,modelrow).value, trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def Greenel():
    SelectFile('xls')
    
    if IsInputSel:
    
        book = xlrd.open_workbook(InputFile, encoding_override="iso-8859-1")
        sheet = book.sheet_by_index(0)
        
        expfilename = "Гринел № {0} от {1}.xlsx".format(Decode(str(sheet.cell(0,30).value)), Decode(str(sheet.cell(0,45).value))).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10,sheet.nrows):
        
            modelrow = 5
            celltype = sheet.cell_type(i, modelrow)

            if trigger:
                if celltype == xlrd.XL_CELL_EMPTY or "," in sheet.cell(i,modelrow).value or "Б" in sheet.cell(i,modelrow).value:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 4:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(i,modelrow).value
                        quantity =      sheet.cell(i,55).value
                        price =         sheet.cell(i,62).value
                        overallprice =  sheet.cell(i,103).value
                        gtdnum =        sheet.cell(i,126).value
                        country =       sheet.cell(i,118).value
                        
                        model = Decode(str(model))
                        gtdnum = Decode(str(gtdnum))
                        country = Decode(str(country))
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if Decode(str(sheet.cell(i,modelrow).value)) == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, Decode(str(sheet.cell(i,modelrow).value)), trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def SDSxml():
    SelectFile('xml')
    
    if IsInputSel:
        tree = ET.parse(InputFile)
        root = tree.getroot()

        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        
        rowitem = 1
        rowcategory = 1
        categorypath = []
        for L1 in root:
            for L2 in L1:
                name = L2.get('НомерСчФ')
                if name:
                    utdnum = name
                name = L2.get('ДатаСчФ')
                if name:
                    utddate = name
                
                for L3 in L2:
                    if L3.tag == 'СведТов':
                        
                        rowitem = rowitem + 1

                        name = str(L3.attrib.get('КолТов'))                                  #'Количество'
                        name = name.replace('.', ',')
                        expsheet.cell(row=rowitem, column=5).value = name
                        
                        name = str(L3.attrib.get('ЦенаТов'))                                 #'Цена'
                        name = name.replace('.', ',')
                        expsheet.cell(row=rowitem, column=6).value = name
                        
                        name = str(L3.attrib.get('СтТовУчНал'))                              #'Сумма'
                        name = name.replace('.', ',')
                        expsheet.cell(row=rowitem, column=7).value = name
                        
                        for L4 in L3:
                            if L4.tag == 'СвТД':
                                name = str(L4.attrib.get('НомерТД'))                         #'Номер ГТД'
                                expsheet.cell(row=rowitem, column=9).value = name
                                
                            if L4.tag == 'ДопСведТов':
                                name = str(L4.attrib.get('КрНаимСтрПр'))                     #'Страна происхождения'
                                if not has_letters(name):
                                    name = "-"
                                expsheet.cell(row=rowitem, column=10).value = name
                                
                            if L4.tag == 'ДопСведТов':                                       #'Артикул'
                                name = str(L4.attrib.get('КодТов'))
                                expsheet.cell(row=rowitem, column=3).value = name
                                
        for i in range(2,rowitem-1):
            if not has_numbers(str(expsheet.cell(row=i+1, column=9).value)):
                expsheet.cell(row=i+1, column=9).value = "-"
        
        expfilename = str("СДС № {0} от {1}.xlsx".format(utdnum, utddate))
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        
        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def SDS():
    SelectFile('xlsx')
    
    if IsInputSel:
    
        book = openpyxl.load_workbook(InputFile)
        sheet = book.active
        
        expfilename = "СДС № {0} от {1}.xlsx".format(sheet.cell(row=4, column=18).value, sheet.cell(row=4, column=31).value).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10, sheet.max_row):
        
            modelrow = 2

            if trigger:
                if sheet.cell(row=i, column=modelrow).value is None or "А" in sheet.cell(row=i, column=modelrow).value:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(row=i, column=modelrow).value
                        quantity =      sheet.cell(row=i, column=33).value
                        price =         sheet.cell(row=i, column=37).value
                        overallprice =  sheet.cell(row=i, column=70).value
                        gtdnum =        sheet.cell(row=i, column=85).value
                        country =       sheet.cell(row=i, column=81).value
                        
                        if not has_numbers(str(gtdnum)):
                            gtdnum = "-"
                        if not has_letters(str(country)):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(row=i, column=modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, sheet.cell(i,modelrow).value, trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))

def SDSv2():
    SelectFile('xlsx')
    
    if IsInputSel:
    
        book = openpyxl.load_workbook(InputFile)
        sheet = book.active
        
        expfilename = "СДС № {0} от {1}.xlsx".format(sheet.cell(row=4, column=23).value, sheet.cell(row=4, column=38).value).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10, sheet.max_row):
        
            modelrow = 2

            if trigger:
                if sheet.cell(row=i, column=modelrow).value is None or "А" in sheet.cell(row=i, column=modelrow).value:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(row=i, column=modelrow).value
                        quantity =      sheet.cell(row=i, column=41).value
                        price =         sheet.cell(row=i, column=46).value
                        overallprice =  sheet.cell(row=i, column=86).value
                        gtdnum =        sheet.cell(row=i, column=104).value
                        country =       sheet.cell(row=i, column=99).value
                        
                        if not has_numbers(str(gtdnum)):
                            gtdnum = "-"
                        if not has_letters(str(country)):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(row=i, column=modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, sheet.cell(i,modelrow).value, trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def Sonel():
    SelectFile('.xlsx .xls')
    extension = Path(InputFile).suffix
    
    if IsInputSel:
        if extension == '.xlsx':
            SonelXLSX()
        if extension == '.xls':
            SonelXLS()

def SonelXLSX():
    book = openpyxl.load_workbook(InputFile)
    sheet = book.active
    
    expfilename = "Сонэл № {0} от {1}.xlsx".format(sheet.cell(row=1, column=22).value, sheet.cell(row=1, column=31).value).replace("/", "-")
    expfilepath = Path(InputFile).parent
    expfilepath = Path(expfilepath, expfilename)
    print(expfilepath)
    
    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Приход"
    #                 1         2        3            4              5          6        7       8         9                10
    fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
    expsheet.append(fieldnames)
    
    counter = 0
    trigger = False
    exprowcounter = 1
    for i in range(10, sheet.max_row):
    
        modelrow = 2

        if trigger:
            if sheet.cell(row=i, column=modelrow).value is None or "№" in sheet.cell(row=i, column=modelrow).value:
                if counter >= 1:
                    counter = counter + 1
                else:
                    counter = 0
                    trigger = False
            else:
                if counter == 3:
                    exprowcounter = exprowcounter + 1
                    
                    model =         sheet.cell(row=i, column=modelrow).value
                    quantity =      sheet.cell(row=i, column=36).value
                    price =         sheet.cell(row=i, column=38).value
                    overallprice =  sheet.cell(row=i, column=66).value
                    gtdnum =        sheet.cell(row=i, column=75).value
                    country =       sheet.cell(row=i, column=73).value
                    
                    if not has_numbers(str(gtdnum)):
                        gtdnum = "-"
                    if not has_letters(str(country)):
                        country = "-"
                    
                    expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                    expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                    expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                    expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                    expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                    expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                else:
                    counter = counter + 1
        
        if sheet.cell(row=i, column=modelrow).value == "Код товара/ работ, услуг":
            counter = 1
            trigger = True
        
        #print("|| {0} || {1} || {2} || {3}".format(celltype, sheet.cell(i,modelrow).value, trigger, counter))

    if FindInternalCodes:
        databookpath = Path(Path(__file__).parent, 'все коды.xls')
        databook = xlrd.open_workbook(databookpath)
        datasheet = databook.sheet_by_index(0)
        
        for i in range(1, expsheet.max_row+1):
            origcode = str(expsheet.cell(row=i, column=3).value)
            
            for k in range(0,datasheet.nrows):
                datacode = str(datasheet.cell(k, 0).value)
                
                if origcode == datacode:
                    expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

    expfile.save(expfilepath)
    print("Opening results file...")
    os.system('"{0}"'.format(expfilepath))

def SonelXLS():
    book = xlrd.open_workbook(InputFile)
    sheet = book.sheet_by_index(0)
    
    expfilename = "Сонэл № {0} от {1}.xlsx".format(sheet.cell(0, 21).value, sheet.cell(0, 30).value).replace("/", "-")
    expfilepath = Path(InputFile).parent
    expfilepath = Path(expfilepath, expfilename)
    print(expfilepath)
    
    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Приход"
    #                 1         2        3            4              5          6        7       8         9                10
    fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
    expsheet.append(fieldnames)
    
    counter = 0
    trigger = False
    exprowcounter = 1
    for i in range(10, sheet.nrows):
    
        modelrow = 1
        celltype = sheet.cell_type(i, modelrow)
        
        if trigger:
            if celltype == xlrd.XL_CELL_EMPTY or "№" in sheet.cell(i, modelrow).value:
                if counter >= 1:
                    counter = counter + 1
                else:
                    counter = 0
                    trigger = False
            else:
                if counter == 3:
                    exprowcounter = exprowcounter + 1
                    
                    model =         sheet.cell(i, modelrow).value
                    quantity =      sheet.cell(i, 35).value
                    price =         sheet.cell(i, 37).value
                    overallprice =  sheet.cell(i, 65).value
                    gtdnum =        sheet.cell(i, 74).value
                    country =       sheet.cell(i, 72).value
                    
                    if not has_numbers(str(gtdnum)):
                        gtdnum = "-"
                    if not has_letters(str(country)):
                        country = "-"
                    
                    expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                    expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                    expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                    expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                    expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                    expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                else:
                    counter = counter + 1
        
        if sheet.cell(i, modelrow).value == "Код товара/ работ, услуг":
            counter = 1
            trigger = True
        
        #print("|| {0} || {1} || {2} || {3}".format(celltype, sheet.cell(i,modelrow).value, trigger, counter))

    if FindInternalCodes:
        databookpath = Path(Path(__file__).parent, 'все коды.xls')
        databook = xlrd.open_workbook(databookpath)
        datasheet = databook.sheet_by_index(0)
        
        for i in range(1, expsheet.max_row+1):
            origcode = str(expsheet.cell(row=i, column=3).value)
            
            for k in range(0,datasheet.nrows):
                datacode = str(datasheet.cell(k, 0).value)
                
                if origcode == datacode:
                    expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

    expfile.save(expfilepath)
    print("Opening results file...")
    os.system('"{0}"'.format(expfilepath))




def Bironi():
    SelectFile('xlsx')
    
    if IsInputSel:
    
        book = openpyxl.load_workbook(InputFile)
        sheet = book.active
        
        expfilename = "ПК Ретро № {0} от {1}.xlsx".format(sheet.cell(row=1, column=16).value, sheet.cell(row=1, column=25).value).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10, sheet.max_row):
        
            modelrow = 2

            if trigger:
                if sheet.cell(row=i, column=modelrow).value is None or "PR" in sheet.cell(row=i, column=modelrow).value:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(row=i, column=modelrow).value
                        quantity =      sheet.cell(row=i, column=27).value
                        price =         sheet.cell(row=i, column=30).value
                        overallprice =  sheet.cell(row=i, column=59).value
                        gtdnum =        sheet.cell(row=i, column=70).value
                        country =       sheet.cell(row=i, column=63).value
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(row=i, column=modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, Decode(str(sheet.cell(i,modelrow).value)), trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def Lezard():
    SelectFile('xlsx')
    
    if IsInputSel:
    
        book = openpyxl.load_workbook(InputFile)
        sheet = book.active
        
        expfilename = "Лезадр-Энергоимпульс № {0} от {1}.xlsx".format(sheet.cell(row=1, column=22).value, sheet.cell(row=1, column=31).value).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10, sheet.max_row):
        
            modelrow = 2

            if trigger:
                if sheet.cell(row=i, column=modelrow).value is None or "№" in sheet.cell(row=i, column=modelrow).value or "А" in sheet.cell(row=i, column=modelrow).value:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(row=i, column=modelrow).value
                        quantity =      sheet.cell(row=i, column=36).value
                        price =         sheet.cell(row=i, column=38).value
                        overallprice =  sheet.cell(row=i, column=66).value
                        gtdnum =        sheet.cell(row=i, column=75).value
                        country =       sheet.cell(row=i, column=73).value
                        
                        quantity = int(int(quantity)/1000)
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(row=i, column=modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            print("|| {0} || {1} || {2}".format(sheet.cell(row=i, column=modelrow).value, trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def LezardMult():
    SelectFile('xls')
    
    if IsInputSel:
    
        book = xlrd.open_workbook(InputFile)
        sheet = book.sheet_by_index(0)
        
        expfilename = "Лезадр-Энергоимпульс склеенный от {0}.xlsx".format(sheet.cell(0, 30).value)
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        
        partname = str("№ {0} от {1}.xlsx".format(sheet.cell(0, 21).value, sheet.cell(0, 30).value))
        expsheet.cell(row=1, column=2).value = partname
        expsheet.cell(row=1, column=2).font = Font(size=12, bold=True)
        
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 2
        for i in range(10, sheet.nrows):
        
            partnamecelltype = sheet.cell_type(i, 21)
            if partnamecelltype != xlrd.XL_CELL_EMPTY:
                if has_letters(str(sheet.cell(i, 21).value)):
                    exprowcounter = exprowcounter + 2
                    
                    partname = str("№ {0} от {1}".format(Decode(str(sheet.cell(i,21).value)), Decode(str(sheet.cell(i,30).value))))
                    expsheet.cell(row=exprowcounter, column=2).value = partname
                    expsheet.cell(row=exprowcounter, column=2).font = Font(size=12, bold=True)
                    expsheet.append(fieldnames)
                    
                    exprowcounter = exprowcounter + 1
        
        
        
            modelrow = 1
            celltype = sheet.cell_type(i, modelrow)
            
            if trigger:
                if celltype == xlrd.XL_CELL_EMPTY or "№" in sheet.cell(i, modelrow).value or "А" in sheet.cell(i, modelrow).value:
                    if counter != 0:
                        counter = counter + 1
                    if counter > 3 or counter == 0:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(i, modelrow).value
                        quantity =      sheet.cell(i, 35).value
                        price =         sheet.cell(i, 37).value
                        overallprice =  sheet.cell(i, 65).value
                        gtdnum =        sheet.cell(i, 74).value
                        country =       sheet.cell(i, 72).value
                        
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(i, modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2}".format(sheet.cell(i, modelrow).value, trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))

def LezardMultX():
    SelectFile('xlsx')
    
    if IsInputSel:
    
        book = openpyxl.load_workbook(InputFile)
        sheet = book.active
        
        expfilename = "Лезадр-Энергоимпульс склеенный от {0}.xlsx".format(sheet.cell(row=1, column=31).value)
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        
        partname = str("№ {0} от {1}.xlsx".format(sheet.cell(row=1, column=22).value, sheet.cell(row=1, column=31).value))
        expsheet.cell(row=1, column=2).value = partname
        expsheet.cell(row=1, column=2).font = Font(size=12, bold=True)
        
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 2
        for i in range(10, sheet.max_row):
        
            if not sheet.cell(row=i, column=22).value is None:
                if has_letters(str(sheet.cell(row=i, column=22).value)):
                    exprowcounter = exprowcounter + 2
                    
                    partname = str("№ {0} от {1}".format(Decode(str(sheet.cell(row=i, column=22).value)), Decode(str(sheet.cell(row=i, column=31).value))))
                    expsheet.cell(row=exprowcounter, column=2).value = partname
                    expsheet.cell(row=exprowcounter, column=2).font = Font(size=12, bold=True)
                    expsheet.append(fieldnames)
                    
                    exprowcounter = exprowcounter + 1
        
            modelrow = 1
            
            if trigger:
                if sheet.cell(row=i, column=modelrow).value is None or "№" in sheet.cell(row=i, column=modelrow).value or "А" in sheet.cell(row=i, column=modelrow).value:
                    if counter != 0:
                        counter = counter + 1
                    if counter > 3 or counter == 0:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(row=i, column=modelrow).value
                        quantity =      sheet.cell(row=i, column=36).value
                        price =         sheet.cell(row=i, column=38).value
                        overallprice =  sheet.cell(row=i, column=66).value
                        gtdnum =        sheet.cell(row=i, column=75).value
                        country =       sheet.cell(row=i, column=73).value
                        
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(row=i, column=modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2}".format(sheet.cell(i, modelrow).value, trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def ArsiLight():
    messagebox.showinfo('Внимание !', 'Не распознает несколько склеенных УПД в одном файле !')
    SelectFile('xlsx')
    
    if IsInputSel:
    
        book = openpyxl.load_workbook(InputFile)
        sheet = book.active
        
        expfilename = "Арсилайт-Белсвет № {0} от {1}.xlsx".format(sheet.cell(row=2, column=11).value, sheet.cell(row=2, column=19).value).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10, sheet.max_row):
        
            modelrow = 3

            if trigger:
                if sheet.cell(row=i, column=modelrow).value is None or sheet.cell(row=i, column=modelrow).value == 1:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(row=i, column=modelrow).value
                        quantity =      sheet.cell(row=i, column=20).value
                        price =         sheet.cell(row=i, column=23).value
                        overallprice =  sheet.cell(row=i, column=40).value
                        gtdnum =        sheet.cell(row=i, column=48).value
                        country =       sheet.cell(row=i, column=46).value
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(row=i, column=modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2}".format(sheet.cell(row=i, column=modelrow).value, trigger, counter))
            
            if FindInternalCodes:
                databookpath = Path(Path(__file__).parent, 'все коды.xls')
                databook = xlrd.open_workbook(databookpath)
                datasheet = databook.sheet_by_index(0)
                
                for i in range(1, expsheet.max_row+1):
                    origcode = str(expsheet.cell(row=i, column=3).value)
                    
                    for k in range(0,datasheet.nrows):
                        datacode = str(datasheet.cell(k, 0).value)
                        
                        if origcode == datacode:
                            expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def AbBatteries():
    SelectFile('xls')
    
    if IsInputSel:
    
        book = xlrd.open_workbook(InputFile, encoding_override="iso-8859-1")
        sheet = book.sheet_by_index(0)
        
        expfilename = "АбБэтэрис № {0} от {1}.xlsx".format(Decode(str(sheet.cell(0,30).value)), Decode(str(sheet.cell(0,45).value))).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10,sheet.nrows):
        
            modelrow = 1
            celltype = sheet.cell_type(i, modelrow)

            if trigger:
                if celltype == xlrd.XL_CELL_EMPTY:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(i, modelrow).value
                        quantity =      sheet.cell(i, 68).value
                        price =         sheet.cell(i, 75).value
                        overallprice =  sheet.cell(i, 123).value
                        gtdnum =        sheet.cell(i, 151).value
                        country =       sheet.cell(i, 140).value
                        
                        model = Decode(str(model))
                        gtdnum = Decode(str(gtdnum))
                        country = Decode(str(country))
                        
                        if not has_numbers(gtdnum):
                            gtdnum = "-"
                        if not has_letters(country):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if Decode(str(sheet.cell(i,modelrow).value)) == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, Decode(str(sheet.cell(i,modelrow).value)), trigger, counter))
            
        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 1).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,0).value
        
        
        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))


def Era():
    msgbxlbl = ['Выберите файл с кодами ЭРА !', '', 'Первый столбец - код Б', 'Второй столбец - внешний код']
    messagebox.showinfo("Внимание !", "\n".join(msgbxlbl))

    EraDataFile = filedialog.askopenfilename(title='Выберите файл', filetypes=(('XLS', 'xls'),))
    
    if EraDataFile:
        messagebox.showinfo('Внимание !', 'Выберите файл УПД !')

        SelectFile('xls')
        
        if IsInputSel:
        
            book = xlrd.open_workbook(InputFile)
            sheet = book.sheet_by_index(0)
            
            expfilecode = sheet.cell(3,8).value
            expfilecode = expfilecode[13:]
            expfilename = "ВИП Маркет {0}.xlsx".format(expfilecode).replace("/", "-")
            expfilepath = Path(InputFile).parent
            expfilepath = Path(expfilepath, expfilename)
            print(expfilepath)
            
            expfile = Workbook()
            expsheet = expfile.active
            expsheet.title = "Приход"
            #                 1         2        3            4              5          6        7       8         9                10
            fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
            expsheet.append(fieldnames)
            
            counter = 0
            trigger = False
            exprowcounter = 1
            for i in range(10,sheet.nrows):
                
                modelrow = 1
                celltype = sheet.cell_type(i, modelrow)

                if trigger:
                    if celltype == xlrd.XL_CELL_EMPTY:
                        if counter >= 1:
                            counter = counter + 1
                        else:
                            counter = 0
                            trigger = False
                    else:
                        if counter == 3:
                            exprowcounter = exprowcounter + 1
                            
                            model =         sheet.cell(i,modelrow).value
                            quantity =      sheet.cell(i,22).value
                            price =         sheet.cell(i,25).value
                            overallprice =  sheet.cell(i,39).value
                            gtdnum =        sheet.cell(i,45).value
                            country =       sheet.cell(i,44).value
                            
                            price = price.replace("-", ",")
                            price = price.replace(" ", "")
                            overallprice = overallprice.replace("-", ",")
                            overallprice = overallprice.replace(" ", "")
                            
                            if not has_numbers(gtdnum):
                                gtdnum = "-"
                            if not has_letters(country):
                                country = "-"
                            
                            expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                            expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                            expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                            expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                            expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                            expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                        else:
                            counter = counter + 1
                
                if "Код" in sheet.cell(i,modelrow).value:
                    counter = 1
                    trigger = True
                
                #print("|| {0} || {1} || {2} || {3}".format(celltype, sheet.cell(i,modelrow).value, trigger, counter))
            
            
            
            databook = xlrd.open_workbook(EraDataFile)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value
                        expsheet.cell(row=i, column=3).value = ""
            
            
            expfile.save(expfilepath)
            print("Opening results file...")
            os.system('"{0}"'.format(expfilepath))


def Gauss():
    SelectFile('xlsx')
    
    if IsInputSel:
    
        book = openpyxl.load_workbook(InputFile)
        sheet = book.active
        
        expfilename = "Gauss № {0} от {1}.xlsx".format(sheet.cell(row=1, column=22).value, sheet.cell(row=1, column=31).value).replace("/", "-")
        expfilepath = Path(InputFile).parent
        expfilepath = Path(expfilepath, expfilename)
        print(expfilepath)
        
        expfile = Workbook()
        expsheet = expfile.active
        expsheet.title = "Приход"
        #                 1         2        3            4              5          6        7       8         9                10
        fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена', 'Сумма', 'НДС', 'Номер ГТД', 'Страна происхождения']
        expsheet.append(fieldnames)
        
        counter = 0
        trigger = False
        exprowcounter = 1
        for i in range(10, sheet.max_row):
        
            modelrow = 2

            if trigger:
                triggercell = sheet.cell(row=i, column=modelrow).value
                if triggercell is None or "№" in triggercell or "A" in triggercell:
                    if counter >= 1:
                        counter = counter + 1
                    else:
                        counter = 0
                        trigger = False
                else:
                    if counter == 3:
                        exprowcounter = exprowcounter + 1
                        
                        model =         sheet.cell(row=i, column=modelrow).value
                        quantity =      sheet.cell(row=i, column=36).value
                        price =         sheet.cell(row=i, column=38).value
                        overallprice =  sheet.cell(row=i, column=66).value
                        gtdnum =        sheet.cell(row=i, column=75).value
                        country =       sheet.cell(row=i, column=73).value
                        
                        if not has_numbers(str(gtdnum)):
                            gtdnum = "-"
                        if not has_letters(str(country)):
                            country = "-"
                        
                        expsheet.cell(row=exprowcounter, column=3).value = model        #'Артикул'
                        expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                        expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'
                        expsheet.cell(row=exprowcounter, column=7).value = overallprice #'Сумма'
                        expsheet.cell(row=exprowcounter, column=9).value = gtdnum       #'Номер ГТД'
                        expsheet.cell(row=exprowcounter, column=10).value = country     #'Страна происхождения'
                    else:
                        counter = counter + 1
            
            if sheet.cell(row=i, column=modelrow).value == "Код товара/ работ, услуг":
                counter = 1
                trigger = True
            
            #print("|| {0} || {1} || {2} || {3}".format(celltype, sheet.cell(i,modelrow).value, trigger, counter))

        if FindInternalCodes:
            databookpath = Path(Path(__file__).parent, 'все коды.xls')
            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            
            for i in range(1, expsheet.max_row+1):
                origcode = str(expsheet.cell(row=i, column=3).value)
                
                for k in range(0,datasheet.nrows):
                    datacode = str(datasheet.cell(k, 0).value)
                    
                    if origcode == datacode:
                        expsheet.cell(row=i, column=2).value = datasheet.cell(k,1).value

        expfile.save(expfilepath)
        print("Opening results file...")
        os.system('"{0}"'.format(expfilepath))








def Placeholder():
    messagebox.showinfo('В работе', 'Еще не готово !')


def Decode(data):
    decoded = data.encode('CP1252', errors='ignore').decode('CP1251', errors='ignore')
    return decoded


def enc_test():
    SelectFile('xls')
    
    if IsInputSel:
        book = xlrd.open_workbook(InputFile)
        sheet = book.sheet_by_index(0)
        encoding = book.encoding
        print('***** {0}'.format(encoding))
        
        for i in range(0, 5):
            for k in range(0, 5):
                print('r={0},col={1} // res={2} //'.format(i,k,sheet.cell(i,k).value))
        
        text = sheet.cell(3,8).value
        print("#### after: ".format(text))
        decoded = text.encode('CP1252').decode('CP1251', errors='ignore')
        print("#### before: ".format(decoded))



def has_numbers(inputString):
    #number_pattern = re.compile(r'-?\d+(\.\d+)?')
    #numbers = number_pattern.findall(inputString)
    #return numbers
    return bool(re.search(r'\d', inputString))
    
def has_letters(inputString):
    cyrillic_pattern = re.compile('[а-яА-ЯёЁ]')
    contains_cyrillic = bool(cyrillic_pattern.search(inputString))
    return contains_cyrillic

if __name__ == '__main__':
    main()
