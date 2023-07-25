from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

import requests, os, sys, csv, re
from pathlib import Path

import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import xml.etree.ElementTree as ET

global root
global scrnwparam
global scrnhparam
scrnwparam = 240
scrnhparam = 195

global QuantityIsInThousands


def main():
    global root

    root = Tk()
    root.resizable(False, False)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('370x100+{}+{}'.format(scrnw, scrnh))
        
    app = GUI(root)
    root.mainloop()
    

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("Перенос заказа ИНФО -> 1С")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):
        #global IsInputSel
        IsInputSel = False

        global FindInternalCodes
        FindInternalCodes = False
        
        #global InputFile
        #global extension

        
        global UniversalBtn
        UniversalBtn = Button(text='Выбрать файл заказа', command=RunUniversal, font=("Arial", 14))
        UniversalBtn.place(x=30, y=20, width=310, height=50)


        try:
            InputFile = sys.argv[1]
            #extension = Path(InputFile).suffix
            print(InputFile)
            print("Файл выбран "+InputFile) 
            Universal(InputFile)
        except:
            print("Файл не выбран") 




def SelectFile(extension):
    #global InputFile
    #global IsInputSel
    extname = extension.upper()

    InputFile = filedialog.askopenfilename(title='Выберите Счет', filetypes=((extname, extension),))
    
    if InputFile:
        print('Файл выбран :' + InputFile)
        
        #print(str(IsInputSel))
        return InputFile
    else:
        print('Файл не выбран')
        return False




def IsInCell(thing,row,column,sheet):
    if str(sheet.cell(row, column).value) == thing:
        return True
    else:
        return False


def AdressInRow(things,row,sheet):
    for y in range(1,sheet.max_column):
        if str(sheet.cell(row, column=y).value) in things:
            return y
    return -1


def FindAllInstancesOnSheet(thing,sheet):
    locations = [-1]
    for x in range(1,sheet.max_row):
        for y in range(1,sheet.max_column):
            if thing in str(sheet.cell(row=x, column=y).value): 
                locations.append([x,y])
    return locations


def NextNotEmptyInRowAdress(row,startcolumn,sheet):
    for y in range(startcolumn+1,sheet.max_column):
        if (sheet.cell(row, column=y).value)!=None:
            return y
    return -1


def SaveAsXLSX(InputFile):
    encod = False
    book_xls = xlrd.open_workbook(InputFile)
    book_xlsx = Workbook()
    if str(format(book_xls.encoding)) == "iso-8859-1":
        print(str(format(book_xls.encoding))+" Декодирую и сохраняю как xlsx")
        encod = True
    else:
        print(" Сохраняю без декодирования как xlsx")
        encod = False
    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
    if sheet_index == 0:
        sheet_xlsx = book_xlsx.active
        sheet_xlsx.title = sheet_name
    else:
        sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

    for row in range(0, sheet_xls.nrows):
        for col in range(0, sheet_xls.ncols):
            if encod == True:
                sheet_xlsx.cell(row = row+1 , column = col+1).value = Decode(str(sheet_xls.cell_value(row, col)))
            else:
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
    savepath = Path(Path(InputFile).parent, str(Path(InputFile).stem+" Перекодированный.xlsx"))
    #print("Сохранил в " + savepath)
    book_xlsx.save(savepath)
    return savepath



def RunUniversal():
    InputFile = SelectFile('.xlsx .xls')
    #print(InputFile)    
    Universal(InputFile)



def Universal(InputFile):
    extension = Path(InputFile).suffix
    QuantityIsInThousands = True
    

    if extension == '.xls':
        savepath = SaveAsXLSX(InputFile)
        book = openpyxl.load_workbook(savepath)
        
    else:
        book = openpyxl.load_workbook(InputFile)
    sheet = book.active
    
    expfilename = Path(InputFile).stem + " Результат.xlsx"
    expfilepath = Path(InputFile).parent
    expfilepath = Path(expfilepath, expfilename)
    #print(expfilepath)
    
    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Счет"
    updnumlocs = FindAllInstancesOnSheet("Код",sheet)
    
    
    #sellertext = ""
    #seller = FindAllInstancesOnSheet("Продавец:",sheet)
    
    #INNtext = ""
    #INN = FindAllInstancesOnSheet("ИНН/КПП продавца:",sheet)
    #if len(seller)!=1:
        
    
    #print(updnumlocs)
    exprowcounter = 0 
    updcounter = len(updnumlocs)-1
    print("Найдено Счетов в файле:")
    print(len(updnumlocs)-1)
    
    for upd in updnumlocs:
        
        if upd!=-1:

            #exprowcounter=exprowcounter+1
        #                 1         2        3            4              5          6        7       8         9                10
            fieldnames = ['Штрихкод', 'Код', 'Артикул', 'Номенклатура', 'Количество', 'Цена']
            exprowcounter=exprowcounter+1
            expsheet.append(fieldnames)

            codecolumn = 1
            modelcolumn = 1
            quantitycolumn = 1
            pricecolumn = 1
            namecolumn = 1
            
            counter = 0
            trigger = False
            startofthenextupd = sheet.max_row
            if updcounter>1:
                startofthenextupd = updnumlocs[updnumlocs.index(upd)+1][0]
                updcounter=updcounter-1
            
            for i in range(upd[0], startofthenextupd+1):

                if trigger:
                        if True: #has_numbers(str(sheet.cell(row=i, column=pricecolumn).value)) and not IsInCell(11,i,gtdnumcolumn,sheet) and not IsInCell("11",i,gtdnumcolumn,sheet) and not IsInCell("А",i,modelcolumn,sheet):
                            exprowcounter = exprowcounter + 1
                            
                            code = str(sheet.cell(row=i, column=codecolumn).value)
                            model =         str(sheet.cell(row=i, column=modelcolumn).value)
                            name =          str(sheet.cell(row=i, column=namecolumn).value)
                            quantity =      str(sheet.cell(row=i, column=quantitycolumn).value)
                            price =         str(sheet.cell(row=i, column=pricecolumn).value)
                            price =         price.replace(" ", "")

                            
                                
                            price = price.replace("-", ",")
                            price = price.replace(" ", "")
                            quantity = quantity.replace("ролл", "")
                            quantity = quantity.replace("гра", "")
                            quantity = quantity.replace("метр", "")                            
                            quantity = quantity.replace(" ", "")
                            quantity = quantity.replace("шт", "")
                            quantity = quantity.replace("м", "")
                            quantity = quantity.replace("упак", "")
                            quantity = quantity.replace("бухта", "")
                            quantity = quantity.replace("бухт", "")
                            quantity = quantity.replace("коплект", "")
                            quantity = quantity.replace("гра", "")
                            quantity = quantity.replace(".", "")
                            #if float(quantity)%1000!=0:
                            #    QuantityIsInThousands = False
                                
                                
                            expsheet.cell(row=exprowcounter, column=2).value = code 
                            expsheet.cell(row=exprowcounter, column=3).value = model
                            expsheet.cell(row=exprowcounter, column=4).value = name        #'Артикул'
                            expsheet.cell(row=exprowcounter, column=5).value = quantity     #'Количество'
                            expsheet.cell(row=exprowcounter, column=6).value = price        #'Цена'

                        else:
                            counter = counter + 1
                #print(str(sheet.cell(i, modelcolumn+1).value))
                
                elif AdressInRow(["Код","Код"],i,sheet)!=-1:
                    codecolumn =AdressInRow(["Код","Код"],i,sheet)
                    modelcolumn = AdressInRow(["Артикул","Артикул"],i,sheet)
                    namecolumn = AdressInRow(["ТМЦ, услуга","ТМЦ"],i,sheet)

                        
                    quantitycolumn = AdressInRow(["Кол"],i,sheet)
                    pricecolumn = AdressInRow(["Цена"],i,sheet)


                    trigger = True
        


    print("Начинаю сохранять результат")
    try:
        expfile.save(expfilepath)
    except:
        msg_box = messagebox.askquestion('Подтверждение', 'Экспортный файл уже открыт, закройте и попробуйте еще раз')
        if msg_box == 'yes':
            expfile.save(expfilepath)
    print("Открываю файл с результатами...")
    os.system('"{0}"'.format(expfilepath))




def Decode(data):
    decoded = data.encode('CP1252', errors='ignore').decode('CP1251', errors='ignore')
    return decoded


def enc_test():
    SelectFile('xls')
    
    if IsInputSel:
        book = xlrd.open_workbook(InputFile)
        sheet = book.sheet_by_index(0)
        encoding = book.encoding
        print('***** {0}'.format(book.encoding))
        
        for i in range(0, 5):
            for k in range(0, 5):
                print('r={0},col={1} // res={2} //'.format(i,k,sheet.cell(i,k).value))
        
        text = sheet.cell(3,8).value
        print("#### after: ".format(text))
        decoded = text.encode('CP1252').decode('CP1251', errors='ignore')
        print("#### before: ".format(decoded))



def has_numbers(inputString):
    #number_pattern = re.compile(r'-?\d+(\.\d+)?')
    #numbers = number_pattern.findall(inputString)
    #return numbers
    return bool(re.search(r'\d', inputString))
    
def has_letters(inputString):
    cyrillic_pattern = re.compile('[а-яА-ЯёЁ]')
    contains_cyrillic = bool(cyrillic_pattern.search(inputString))
    return contains_cyrillic

if __name__ == '__main__':
    main()
