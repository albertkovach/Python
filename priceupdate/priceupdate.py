from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import tkinter.font as TkFont

from pathlib import Path
import requests, os, sys, csv
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

global root
global scrnwparam
global scrnhparam
scrnwparam = 185
scrnhparam = 150

def main():
    global root

    root = Tk()
    root.resizable(False, False)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('220x100+{}+{}'.format(scrnw, scrnh))
        
    app = GUI(root)
    root.mainloop()

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):
        global dir_path
        dir_path = Path(os.path.dirname(os.path.realpath(__file__)))
        
        global bn_path
        bn_path = Path(dir_path, 'BN.xml')
        global n_path
        n_path = Path(dir_path, 'N.xml')
        global wbn_path
        wbn_path = Path(dir_path, 'БН.xlsx')
        global wn_path
        wn_path = Path(dir_path, 'Н.xlsx')
        
        customfont= TkFont.Font(family='Helvetica', size=12, weight='bold')
        
        global StartBtn
        StartBtn = Button(text='Запуск', command=BtnCmd, font=customfont, bg="#eff252")
        StartBtn.place(x=35, y=25, height=35, width=150)

        


def BtnCmd():
    msg_box = messagebox.askquestion('Подтверждение', 'Уверены ?')
    
    if msg_box == 'yes':
        GetFiles()


def GetFiles():
    global bn_path
    global n_path
    
    print("Загрузка прайсов...")

    price_bn_url = "https://www.makel25.ru/bitrix/catalog_export/export_bn.xml"
    revisedreqfile = bn_path.as_posix()
    response = requests.get(price_bn_url)
    if response.status_code == 200:
        with open(revisedreqfile,'wb') as reqfile:
            reqfile.write(response.content)
        print("Файл БН прайса загружен !")
    else:
        print('Ошибка', 'Запрос БН прайса прервался с кодом ошибки: {0}'.format(response.status_code))
        return
        
        
    price_n_url = "https://www.makel25.ru/bitrix/catalog_export/export_bn.xml"
    revisedreqfile = n_path.as_posix()
    response = requests.get(price_n_url)
    if response.status_code == 200:
        with open(revisedreqfile,'wb') as reqfile:
            reqfile.write(response.content)   
        print("Оба прайса загружены !")
    else:
        print('Ошибка', 'Запрос Н прайса прервался с кодом ошибки: {0}'.format(response.status_code))
        return

    MakeFile()

    
def MakeFile():
    global bn_path
    global n_path
    global wbn_path
    global wn_path
    
    
    
    print("Сборка файла БН ...")
    
    wb = Workbook()
    sheetitems = wb.active
    sheetitems.title = "Прайс-лист"

    itemsfieldnames = ['Код', 'Артикул', 'Наименование', 'Цена', 'Кратность', 'Ед.изм']
    sheetitems.append(itemsfieldnames) 
    
    for i in range(1,7):
        sheetitems.cell(row=1, column=i).font = Font(bold=True, size=12)
    sheetitems.column_dimensions['A'].width = 9
    sheetitems.column_dimensions['B'].width = 20
    sheetitems.column_dimensions['C'].width = 80
    sheetitems.column_dimensions['D'].width = 10
    sheetitems.column_dimensions['E'].width = 10
    sheetitems.column_dimensions['F'].width = 10

    revisedreqfile = bn_path.as_posix()
    tree = ET.parse(revisedreqfile)
    root = tree.getroot()
    rowitem = 1

    for L1 in root:
        for L2 in L1:
            for L3 in L2:

                if L3.tag == 'offer':

                    rowitem = rowitem + 1
                    if rowitem%100 == 0:
                        print("Запись БН ", rowitem)
                    
                    code = L3.attrib.get('id') ### Внешний код
                    sheetitems.cell(row=rowitem, column=1).value = int(code)
                    
                    for L4 in L3:
                        if L4.tag == 'description': ### Артикул
                            sheetitems.cell(row=rowitem, column=2).value = L4.text
                            
                        elif L4.tag == 'name': ### Наименование
                            sheetitems.cell(row=rowitem, column=3).value = L4.text
                            
                        elif L4.tag == 'price': ### Цена
                            sheetitems.cell(row=rowitem, column=4).value = float(L4.text)
                            
                        elif L4.tag == 'vendorCode': ### Кратность
                            try:
                                sheetitems.cell(row=rowitem, column=5).value = int(L4.text)
                            except:
                                sheetitems.cell(row=rowitem, column=5).value = L4.text
                            
                        elif L4.tag == 'vendor': ### Ед.изм
                            sheetitems.cell(row=rowitem, column=6).value = L4.text

    wb.save(wbn_path)
    print("Готово !")
    
    
    
    
    
    print("Сборка файла Н ...")
    
    wb = Workbook()
    sheetitems = wb.active
    sheetitems.title = "Прайс-лист"

    itemsfieldnames = ['Код', 'Артикул', 'Наименование', 'Цена', 'Кратность', 'Ед.изм']
    sheetitems.append(itemsfieldnames) 
    
    for i in range(1,7):
        sheetitems.cell(row=1, column=i).font = Font(bold=True, size=12)
    sheetitems.column_dimensions['A'].width = 9
    sheetitems.column_dimensions['B'].width = 20
    sheetitems.column_dimensions['C'].width = 80
    sheetitems.column_dimensions['D'].width = 10
    sheetitems.column_dimensions['E'].width = 10
    sheetitems.column_dimensions['F'].width = 10

    revisedreqfile = n_path.as_posix()
    tree = ET.parse(revisedreqfile)
    root = tree.getroot()
    rowitem = 1

    for L1 in root:
        for L2 in L1:
            for L3 in L2:

                if L3.tag == 'offer':

                    rowitem = rowitem + 1
                    if rowitem%100 == 0:
                        print("Запись Н ", rowitem)
                    
                    code = L3.attrib.get('id') ### Внешний код
                    sheetitems.cell(row=rowitem, column=1).value = int(code)
                    
                    for L4 in L3:
                        if L4.tag == 'description': ### Артикул
                            sheetitems.cell(row=rowitem, column=2).value = L4.text
                            
                        elif L4.tag == 'name': ### Наименование
                            sheetitems.cell(row=rowitem, column=3).value = L4.text
                            
                        elif L4.tag == 'price': ### Цена
                            sheetitems.cell(row=rowitem, column=4).value = float(L4.text)
                            
                        elif L4.tag == 'vendorCode': ### Кратность
                            try:
                                sheetitems.cell(row=rowitem, column=5).value = int(L4.text)
                            except:
                                sheetitems.cell(row=rowitem, column=5).value = L4.text
                            
                        elif L4.tag == 'vendor': ### Ед.изм
                            sheetitems.cell(row=rowitem, column=6).value = L4.text

    wb.save(wn_path)
    print("Готово !")

    
    messagebox.showinfo("Внимание !", "Выполнено !")







if __name__ == '__main__':
    main()
