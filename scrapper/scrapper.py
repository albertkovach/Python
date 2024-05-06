from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import os, requests,re
from pathlib import Path
import xlrd
import openpyxl
from openpyxl import Workbook
from configparser import ConfigParser
from urllib.request import Request, urlopen
import time

global root
global scrnwparam
global scrnhparam
scrnwparam = 185
scrnhparam = 150

global myURL
global page
global soup



import requests
import re
from bs4 import BeautifulSoup

def main():
    global root
    global counter
    root = Tk()
    root.resizable(True, True)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('1030x570+{}+{}'.format(scrnw, scrnh))
    #root.resizable(height = None, width = None)
    
    app = GUI(root)
    root.mainloop()

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):

        global RunBtn
        RunBtn = Button(text='RUN', command=RUN, font=("Arial", 11))
        RunBtn.place(x=10, y=15, width=60, height=25)
        
        global URLText
        URLText = Text()
        URLText.place(x=10, y=50, width=1000, height=500)
        URLText.insert(END,"https://sk-electro.com/shop/goods/ramka_2_m_Cosmo_alyum_ABB_612_011000_226-64523")



def RUN():
    print('****** Starting...  {0}'.format(time.ctime()))

    expfile = Workbook()
    expsheet = expfile.active
    expsheet.title = "Приход"
    fieldnames = ['Артикул', 'Штрихкод']
    expsheet.append(fieldnames)

    dir_path = os.path.dirname(os.path.realpath(__file__))
    filename = 'SCRAPPER.xlsx'
    save_path = os.path.join(dir_path, filename)


    urls = URLText.get('1.0', END).splitlines()


    counter = 1
    for url in urls:
        counter = counter + 1
        try:
            print('**** {0} of {1} : {2}'.format(counter-1, len(urls), url))
            #time.sleep(15)
            
            agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
            request = Request(url, headers={'User-Agent': agent})
            time.sleep(1)
            html = urlopen(request).read().decode()
            time.sleep(1)
            soup = BeautifulSoup(html, 'html.parser')
            

            try:
                block_features = soup.find('div',{'class' : 'features'})
                lines = str(block_features).splitlines()

                for i in range(len(lines)):
                    if 'Штрих код' in str(lines[i]):
                        barcode = lines[i+1]
                        start = barcode.rfind('<span>') + 6
                        stop = len(barcode) - 13
                        barcode = barcode[start:stop]
                        expsheet.cell(counter,2).value = barcode
                        
                        print('* ', barcode)
                    if 'Код производителя:' in str(lines[i]):
                        article = lines[i+1]
                        start = article.rfind('<span>') + 6
                        stop = len(article) - 13
                        article = article[start:stop]
                        expsheet.cell(counter,1).value = article
                        
                        print('* ', article)

            except Exception as error:
                print('Exception while exec: {0} /n {1}'.format(type(error).__name__, error))


        except:
            print('CRITICAL FAILURE')
    

    expfile.save(save_path)
    print('****** DONE !   {0}'.format(time.ctime()))



if __name__ == '__main__':
    main()
