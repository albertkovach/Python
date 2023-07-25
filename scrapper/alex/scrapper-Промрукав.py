from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import os, requests,re
from pathlib import Path
import xlrd
import openpyxl
from openpyxl import Workbook
from configparser import ConfigParser
from urllib.request import Request, urlopen
import time

global root
global scrnwparam
global scrnhparam
scrnwparam = 185
scrnhparam = 150

global myURL
global page
global soup



import requests
import re
from bs4 import BeautifulSoup

def main():
    global nextpaging
    global root
    global counter_products
    
    nextpaging = True #Вести ли перебор страниц списка товаров
    nazvanie = '-Elevel-rozetki.xlsx' #
    
    
    root = Tk()
    root.resizable(True, True)
    
    
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('500x500+{}+{}'.format(scrnw, scrnh))
    #root.resizable(height = None, width = None)
    
    app = GUI(root)
    root.mainloop()

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):
        global main_site_adress
        global counter_products # количество заскрапленных товаров (строка в файле эксель)
        counter_products = 1
        main_site_adress = "https://www.promrukav.ru"
        
        
        
        global p3Btn
        p3Btn = Button(text='p3', command=p3)
        p3Btn.place(x=10, y=70, width=40, height=20)
        
        global URLText
        URLText = Text()
        URLText.place(x=10, y=100, width=1000, height=500)
        URLText.insert(END,"https://www.promrukav.ru/catalog/kabelnyy-kanal/") #что введено в форму по умолчанию
        myURL = URLText.get("1.0","1.0 lineend")
        print(myURL)
        global expfile
        expfile = Workbook()
        global ParameterNames
        ParameterNames = []



def iterate():
    
    while(True):
        try:
            p3()
        except:
            print("ОООООШШШШИИИИИББККККААА НА "+str(counter_products))

def p3():
    global nextpage
    global counter_products
    global nextpaging
    global main_site_adress
    dir_path = os.path.dirname(os.path.realpath(__file__))
    nmn = main_site_adress.replace("/","")
    nmn = nmn.replace(".","-")
    nmn = "---"+nmn.replace(":","-")
    print(nmn)


    #for i in range(int(URLText.index('end').split(".",1)[0])):
    myURL = URLText.get("1.0","1.0 lineend")
    #print(MyURL)
        
    agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.36\(KHTML, like Gecko) Chrome/35.0.1916.47 Safari/537.36'

    request = Request(myURL, headers={'User-Agent': agent})
    time.sleep(1)
    html = urlopen(request).read().decode()
    time.sleep(1)
    soup = BeautifulSoup(html, 'html.parser')


    global expfile
    global ParameterNames


    expsheet = expfile.active
    #               1        2             3           4        5        6    
    urls = []       
    next_page_exists = True
    counter_products1 = 0
    while(next_page_exists):
        links = soup.find_all('div', {'class' : 'product block'})
        linkz = []
        for link in links:
            linkz.append(link.find('a', {'class' : 'image'}))
        #print(linkz)
        for link in linkz:
            #print(link['href'])
            urls.append(main_site_adress+link['href'])
        print(urls)
        nextpagez = soup.find('li',{'class':"bx-pag-next"})
        nextpage = nextpagez.find('a')
        next_page_exists = False
        if nextpaging:
            if nextpage and counter_products1<100:
                #print(nextpage['href'])
                page = requests.get(main_site_adress+nextpage['href'])
                counter_products1=counter_products1+1
                soup = BeautifulSoup(page.content, 'html.parser')
                lastnextpage= nextpage
                next_page_exists = True
            else:           
                URLText.delete('1.0', END)
                URLText.insert(END,main_site_adress+lastnextpage['href'])
                URLText.update()
            #next_page_exists = False
    #<a class="pagination__next" href="/catalog/nizkvoltnoe-oborudovanie/avtomaticheskie-vyklyuchateli/&amp;PAGEN_1=2">След.</a>
    #print(urls)

    fieldnames = ['Описание', 'Фото', 'Наименование','ССылка', 'Запасная ячейка 5', 'Запасная ячейка 6','Цены']
    expsheet.append(fieldnames)

    
    for url in urls:
        
        page = requests.get(url)
        
        soup = BeautifulSoup(page.content, 'html.parser')


        counter_products=counter_products+1


        try:
            nomenkl = soup.find('meta',{'property' : 'og:title'})
            expsheet.cell(counter_products,3).value = nomenkl["content"]
        except:
            print("FAILURE at nomenkl")
        
        
        try:
            expsheet.cell(counter_products,4).value = str(url)
        except:
            print("FAILURE at URL")                
        
        
        try:
            description = soup.find('div',{'class' : 'description'})
            characteristics_soup = soup.find('div', {'class' : 'properties-table'})
            characteristics = characteristics_soup.find_all('div',{'class':'row'})
            charstring = ""
            for characteristic in characteristics:
                name = characteristic.find('div', {'class' : 'name'})
                value = characteristic.find('div', {'class' : 'value'})

                charstring = charstring+"\n"+name.text+":"+value.text
            expsheet.cell(counter_products,1).value = description.text+charstring
        except:
            print("FAILURE at descr")
        
        
        

        try:
            price = soup.find('span',{'class' : 'price'})
            
            expsheet.cell(counter_products,7).value = price.text
        except:
            print("FAILURE")
     

        try:
            counter_productsX = 2
            images1 = soup.find_all('img', {'class' : 'lb-image'})
            #print(images1)
            images2 = soup.find_all('meta', {'property' : 'og:image'})
            #print(images2)
            images= images1+images2
            print(images)
            for image in images:
                #img = image.find('img')
                imagestring = str(expsheet.cell(counter_products,counter_productsX).value)+"; "+image['content']
                imagestring = imagestring.replace("None;","")
                expsheet.cell(counter_products,counter_productsX).value = imagestring
                #counter_productsX=counter_productsX+1
                #print(img['src'])
        except:
            print("FAILURE at images")    
            


        try:    
            characteristics_soup = soup.find('div', {'class' : 'properties-table'})
            characteristics = characteristics_soup.find_all('div',{'class':'row'})
            for characteristic in characteristics:
                name = characteristic.find('div', {'class' : 'name'})
                value = characteristic.find('div', {'class' : 'value'})

                #print(name.text+value.text)
                Nametext = name.text.strip()
                if not name.text.strip() in ParameterNames:
                    ParameterNames.append(Nametext)
                    expsheet.cell(1,ParameterNames.index(Nametext)+9).value = Nametext
                
                expsheet.cell(counter_products,ParameterNames.index(Nametext)+9).value = value.text.strip()
        except:
            print("FAILURE at characteristics")
        print(counter_products)
    filename = str(counter_products)+ nmn + ".xlsx"
    save_path = os.path.join(dir_path, filename)
    expfile.save(save_path)
    print("The End")

if __name__ == '__main__':
    main()
