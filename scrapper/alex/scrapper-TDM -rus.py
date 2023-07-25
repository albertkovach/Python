from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import os, requests,re
from pathlib import Path
import xlrd
import openpyxl
from openpyxl import Workbook
from configparser import ConfigParser

global root
global scrnwparam
global scrnhparam
scrnwparam = 185
scrnhparam = 150

global myURL
global page
global soup



import requests
import re
from bs4 import BeautifulSoup

def main():
    global root
    global counter
    root = Tk()
    root.resizable(True, True)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('500x500+{}+{}'.format(scrnw, scrnh))
    #root.resizable(height = None, width = None)
    
    app = GUI(root)
    root.mainloop()

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):

        global counter
        counter = 2
        
        global p3Btn
        p3Btn = Button(text='p3', command=p3)
        p3Btn.place(x=10, y=70, width=40, height=20)
        
        global URLText
        URLText = Text()
        URLText.place(x=10, y=100, width=1000, height=500)
        URLText.insert(END,"https://www.tdm-rus.ru/category/avtomati/20")
        myURL = URLText.get("1.0","1.0 lineend")
        print(myURL)
        
        global expfile
        expfile = Workbook()
        global ParameterNames
        ParameterNames = []



def iterate():
    while(True):
        p3()


def p3():
    global counter
    dir_path = os.path.dirname(os.path.realpath(__file__))
    filename = str(counter)+'TDM-rus.xlsx'
    save_path = os.path.join(dir_path, filename)
    print(save_path)
    #for i in range(int(URLText.index('end').split(".",1)[0])):
    myURL = URLText.get("1.0","1.0 lineend")
    #print(MyURL)
        
    page = requests.get(myURL)

    soup = BeautifulSoup(page.content, 'html.parser')


    global expfile
    global ParameterNames


    expsheet = expfile.active
    expsheet.title = "Приход"
    #               1        2             3           4        5        6    
    urls = []       
    nextpaging = True
    counter1 = 0
    while(nextpaging):
        nextpaging = False
        #grid = soup.find('div',{'class' : 'grid'})
        links = soup.find_all('div', {'class' : 'product'})
        #print(links)


        #print(links)
        for link in links:
            lnk = link.find('a')
            urls.append("https://www.tdm-rus.ru"+lnk['href'])
            print("https://www.tdm-rus.ru"+lnk['href'])
            
        #print(url)
        if counter1<100:
            
            try:
                nextpagex =soup.find('li',{'class':'page-next'})

                nextpage = nextpagex.find('a',{'class':'page-link'})
                print(nextpage['href'])
                page = requests.get("https://www.tdm-rus.ru"+nextpage['href'])
                counter1=counter1+1
                soup = BeautifulSoup(page.content, 'html.parser')
                nextpaging = True
            except:
                print("no next")
                
            #if nextpage and counter1<99:
    #<a class="pagination__next" href="/catalog/nizkvoltnoe-oborudovanie/avtomaticheskie-vyklyuchateli/&amp;PAGEN_1=2">След.</a>
    #print(urls)

    fieldnames = ['Описание', 'Фото', 'Наименование','ССылка', 'Видео', 'Название видео','Цены']
    expsheet.append(fieldnames)

    
    for url in urls:
        
        page = requests.get(url)
        

        soup = BeautifulSoup(page.content, 'html.parser')
        #articlespan = soup.find_all('span', {'class' : 'changeArticle'})
        #for data in articlespan:
        #    article = data.text

       #     n=data.text
            
        #print(n)


        counter=counter+1
        #expsheet.cell(counter,1).value = n
        print(url)
        try:
            nomenkl = soup.find('meta',{'property':'og:title'})
            print(nomenkl)
            #print(nomenkl[1])
            #print(nomenkl[1]['content'])
            expsheet.cell(counter,3).value = nomenkl['content']
        except:
            print("FAILURE1")
        try:
            expsheet.cell(counter,4).value = str(url)
        except:
            print("FAILURE")                
        try:
            description = soup.find('div',{'class' : 'product-description'})
            #for span in description("span"):
             #   span.decompose()
            #for h2 in description("h2"):
            #    h2.decompose()
            #for li in description("li"):
            #    li.decompose()
            #for ul in description("ul"):
            #    ul.decompose()
           # for p in description("p"):
            #    p.decompose()
                #for li in description("li"):
                #li.decompose()
            #ul = description.find('ul')
            #li = ul.find_all('span')
            #txtx=""
            #for span in li:
            #    txtx = txtx+span.text
            expsheet.cell(counter,1).value = description.get_text(separator="\n")
            #print(description.get_text())
            #<meta property="og:title" content="Tdm Плавкая вставка Н520Б 6А 250В SQ0738-0014">
        except:
            print("FAILURE2")
        try:
            video = soup.find('div',{'class' : 'product-articul'})
            videoa = video.find('span')
            expsheet.cell(counter,5).value = videoa.text
            #expsheet.cell(counter,6).value = videoa['title']
        except:
            print("FAILURE")
        #<span data-v-2033c1fe="" class="w-1/2 block font-bold py-1 border-gray-300 border text-center text-sm sm:text-base text-secondary">21.14 ₽</span>
        try:
            price = soup.find('span',{'class' : 'w-1/2'})
            
            expsheet.cell(counter,7).value = price.text
        except:
            print("FAILURE3")
     


        counterX = 2
        images = soup.find_all('meta', {'property' : 'og:image'})
        for image in images:
            #img = image.find('img')
            expsheet.cell(counter,counterX).value = str(expsheet.cell(counter,counterX).value)+";"+str(image['content'])
            #counterX=counterX+1
            #print(img['src'])
            
            
#<img data-v-1b90539c="" alt="product-img" class="object-cover swiper-lazy swiper-lazy-loaded" src="https://tdme.ru/download/WebImage/TDM-SQ0319-0005.jpg">
#<tr data-v-2033c1fe="" class="border-b odd:bg-white even:bg-gray-200"><td data-v-2033c1fe="" class="py-4 px-6 text-xs lg:text-sm text-gray-700">
  #                  Цоколь (патрон) лампы
  #                </td> <td data-v-2033c1fe="" class="lg:w-1/3 lg:w-auto py-4 lg:px-6 text-xs break-all lg:break-normal lg:text-sm text-gray-700"><span data-v-2033c1fe="">
  #                    E14
    #                </span></td> <td data-v-2033c1fe="" class="lg:w-1/3 lg:w-auto py-4 lg:px-6 text-xs break-all lg:break-normal lg:text-sm text-gray-700"><span data-v-2033c1fe="">
   #                   EF000048
     #               </span></td></tr>
            
        characteristics = soup.find_all('li', {'class' : 'product-property'})
        for characteristic in characteristics:
            name = characteristic.find('span', {'itemprop':'name'})
            #values = characteristic.find_all('td', {'class' : 'card-tabs-table__cell'})
            value = characteristic.find('span',{'class','product-property-value'})
            #valuea = value.find('a')
            #print(name.text+value.text)
            Nametext = name.text.strip()
            if not name.text.strip() in ParameterNames:
                ParameterNames.append(Nametext)
                expsheet.cell(1,ParameterNames.index(Nametext)+9).value = Nametext
            
            expsheet.cell(counter,ParameterNames.index(Nametext)+9).value = value.text.strip()
        print(counter)

    expfile.save(save_path)
    print("The End")

if __name__ == '__main__':
    main()
