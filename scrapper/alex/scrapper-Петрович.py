from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import os, requests,re
from pathlib import Path
import xlrd
import openpyxl
from openpyxl import Workbook
from configparser import ConfigParser
import time

global root
global scrnwparam
global scrnhparam
scrnwparam = 185
scrnhparam = 150

global myURL
global page
global soup



import requests
import re
from bs4 import BeautifulSoup

def main():
    global root
    global counter
    root = Tk()
    root.resizable(True, True)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('500x500+{}+{}'.format(scrnw, scrnh))
    #root.resizable(height = None, width = None)
    
    app = GUI(root)
    root.mainloop()

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):

        global counter
        counter = 2
        
        global p3Btn
        p3Btn = Button(text='p3', command=p3)
        p3Btn.place(x=10, y=70, width=40, height=20)
        
        global URLText
        URLText = Text()
        URLText.place(x=10, y=100, width=1000, height=500)
        URLText.insert(END,"https://petrovich.ru/catalog/10168/makel/")
        myURL = URLText.get("1.0","1.0 lineend")
        print(myURL)
        
        global expfile
        expfile = Workbook()
        global ParameterNames
        ParameterNames = []



def iterate():
    urltest = "https://petrovich.ru/catalog/10168/makel/"
    page = requests.get(urltest)
    soup = BeautifulSoup(page.content, 'html.parser')
    print(soup)
    #while(True):
    #    try:
    #        p3()
     #   except:
     #       print("ОООООШШШШИИИИИББККККААА НА "+str(counter))

def p3():
    global nextpage
    global counter
    dir_path = os.path.dirname(os.path.realpath(__file__))
    
    filename = str(counter)+'-Петрович-Макел-rozetki.xlsx'
    save_path = os.path.join(dir_path, filename)

    #for i in range(int(URLText.index('end').split(".",1)[0])):
    myURL = URLText.get("1.0","1.0 lineend")
    #print(MyURL)
        
    page = requests.get(myURL)

    soup = BeautifulSoup(page.content, 'html.parser')


    global expfile
    global ParameterNames


    expsheet = expfile.active
    expsheet.title = "Приход"
    #               1        2             3           4        5        6    
    urls = []       
    nextpaging = True
    counter1 = 0
    while(nextpaging):
        links = soup.find_all('div', {'class' : 'fade-in-list'})



        #print(links)
        for l in links:
            link = l.find('a',{'data-test':'product-image'})
            #print(link['href'])
            urls.append("https://petrovich.ru"+link['href'])
        #print(url)
        nextpage = soup.find('a',{'class':"pagination__link pagination__link--next"})
        
        if nextpage and counter1<100:
            print(nextpage['href'])
            page = requests.get("https://petrovich.ru"+nextpage['href'])
            counter1=counter1+1
            soup = BeautifulSoup(page.content, 'html.parser')
            lastnextpage= nextpage
        else:           
            #URLText.delete('1.0', END)
            #URLText.insert(END,"https://petrovich.ru"+lastnextpage['href'])
            #URLText.update()
            nextpaging = False
    #<a class="pagination__next" href="/catalog/nizkvoltnoe-oborudovanie/avtomaticheskie-vyklyuchateli/&amp;PAGEN_1=2">След.</a>
    
    if counter == 2:
        fieldnames = ['Описание', 'Фото', 'Наименование','ССылка', 'Видео', 'Название видео','Цены']
        expsheet.append(fieldnames)

    
    for url in urls:
        time.sleep(15)
        page = requests.get(url)
        

        soup = BeautifulSoup(page.content, 'html.parser')
        #articlespan = soup.find_all('span', {'class' : 'changeArticle'})
        #for data in articlespan:
        #    article = data.text
        print(soup)
       #     n=data.text
            
        #print(n)

        counter=counter+1
        #expsheet.cell(counter,1).value = n
        print(url)
        try:
            nomenkl = soup.find('meta',{'property' : 'og:title'})
            expsheet.cell(counter,3).value = nomenkl["content"]
        except:
            print("FAILURE")
        try:
            expsheet.cell(counter,4).value = str(url)
        except:
            print("FAILURE")                
        try:
            description = soup.find('p',{'class' : 'product-description-text'})
            expsheet.cell(counter,1).value = description.text
        except:
            print("FAILURE")
        try:
            video = soup.find('div',{'class' : 'card-tabs__videos'})
            videoa = video.find('iframe')
            expsheet.cell(counter,5).value = videoa['src']
            expsheet.cell(counter,6).value = videoa['title']
        except:
            print("FAILURE")
        try:
            price = soup.find('span',{'class' : 'card-price__current'})
            
            expsheet.cell(counter,7).value = price.text
        except:
            print("FAILURE")
     


        counterX = 2
        images = soup.find_all('div', {'class' : 'content-slide'})
        control = []
        for i in images:
            #img = image.find('img')
            image = i.find('img')
            if not image in control:
                control.append(image)
                expsheet.cell(counter,counterX).value = str(expsheet.cell(counter,counterX).value)+";"+image['data-src']
            #counterX=counterX+1
            #print(img['src'])
            
            


        ul = soup.find('ul', {'class' : 'product-properties-list'})    
        characteristics = ul.find_all('li', {'class' : 'data-item'})
        for characteristic in characteristics:
            #print(characteristic)
            name = characteristic.find('div', {'class' : 'title'})
            value = characteristic.find('div', {'class' : 'value'})
            #value = values[1]
            #valuea = value.find('a')
            #print(name.text+value.text)
            Nametext = name.text.strip()
            if not name.text.strip() in ParameterNames:
                ParameterNames.append(Nametext)
                expsheet.cell(1,ParameterNames.index(Nametext)+9).value = Nametext
            
            expsheet.cell(counter,ParameterNames.index(Nametext)+9).value = value.text.strip()
        print(counter)
    
    expfile.save(save_path)
    print("The End")

if __name__ == '__main__':
    main()
