from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import os, requests,re
from pathlib import Path
import xlrd
import openpyxl
from openpyxl import Workbook
from configparser import ConfigParser

global root
global scrnwparam
global scrnhparam
scrnwparam = 185
scrnhparam = 150

global myURL
global page
global soup



import requests
import re
from bs4 import BeautifulSoup

def main():
    global root
    global counter
    root = Tk()
    root.resizable(True, True)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('500x500+{}+{}'.format(scrnw, scrnh))
    #root.resizable(height = None, width = None)
    
    app = GUI(root)
    root.mainloop()

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):

        global counter
        counter = 2
        
        global p1Btn
        p1Btn = Button(text='p1', command=p1)
        p1Btn.place(x=10, y=10, width=40, height=20)
        
        global p2Btn
        p2Btn = Button(text='p2', command=p2)
        p2Btn.place(x=10, y=40, width=40, height=20)
        
        global p3Btn
        p3Btn = Button(text='p3', command=p3)
        p3Btn.place(x=10, y=70, width=40, height=20)
        
        global URLText
        URLText = Text()
        URLText.place(x=10, y=100, width=500, height=500)
        URLText.insert(END,"https://bironi.ru/catalog/retro-vyklyuchateli/")
        myURL = URLText.get("1.0","1.0 lineend")
        print(myURL)
        
        global expfile
        expfile = Workbook()
        global ParameterNames
        ParameterNames = []


def p1():
    articlespan = soup.find_all('span', {'class' : 'changeArticle'})

    for data in articlespan:
        article = data.text
        #print(data.text)
        n=data.text
        
    print(n)
    n = n.replace("-","_")
    print(n)



def p2():

    articlespan = soup.find_all('span', {'class' : 'changeArticle'})

    for data in articlespan:
        article = data.text
        #print(data.text)
        n=data.text
        
    print(n)
    n = n.replace("-","_")
    print(n)


    images = soup.findAll('img')
    
    #for image in images:
        




def p3():
    global counter
    dir_path = os.path.dirname(os.path.realpath(__file__))
    filename = 'САЙТ.xlsx'
    save_path = os.path.join(dir_path, filename)

    for i in range(int(URLText.index('end').split(".",1)[0])):
        myURL = URLText.get(str(i)+".0",str(i)+".0 lineend")
        #print(MyURL)
            
        page = requests.get(myURL)

        soup = BeautifulSoup(page.content, 'html.parser')


        global expfile
        global ParameterNames


        expsheet = expfile.active
        expsheet.title = "Приход"
        #               1        2             3           4        5        6    
        links = soup.find_all('a', {'class' : 'picture'})
        urls = []
        print(links)
        for link in links:
            print(link['href'])
            urls.append("https://bironi.ru"+link['href'])
            #print(url)
        #counter = 2
        


        fieldnames = ['Артикул', 'Фото', 'Наименование','ССылка']
        expsheet.append(fieldnames)

        for url in urls:
            
            page = requests.get(url)
            

            soup = BeautifulSoup(page.content, 'html.parser')
            articlespan = soup.find_all('span', {'class' : 'changeArticle'})
            for data in articlespan:
                article = data.text

                n=data.text
                
            print(n)

            counter=counter+1
            expsheet.cell(counter,1).value = n
            
            try:
                nomenkl = soup.find('meta',{'property' : 'og:title'})
                expsheet.cell(counter,3).value = nomenkl['content']
            except:
                print("FAILURE")
            try:
                expsheet.cell(counter,4).value = str(url)
            except:
                print("FAILURE")                
            
            counterX = 2
            images = soup.find_all('meta', {'property' : 'og:image'})
            for image in images:

                expsheet.cell(counter,counterX).value = image['content']
                counterX=counterX+1
                print(image["content"])
                
            images = soup.find_all('div', {'class' : 'propertyTable'})
            for image in images:
                name = image.find('div', {'class' : 'propertyName'})
                value = image.find('div', {'class' : 'propertyValue'})
                print(name.text+value.text)
                Nametext = name.text.strip()
                if not name.text.strip() in ParameterNames:
                    ParameterNames.append(Nametext)
                    expsheet.cell(1,ParameterNames.index(Nametext)+5).value = Nametext
                
                expsheet.cell(counter,ParameterNames.index(Nametext)+5).value = value.text.strip()


    expfile.save(save_path)
    

if __name__ == '__main__':
    main()
