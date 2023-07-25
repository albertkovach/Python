from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

import requests, os, sys, csv, re
from pathlib import Path

import xlrd
import openpyxl
import pyxlsb
from openpyxl import Workbook
from openpyxl.styles import Font
import xml.etree.ElementTree as ET





def SelectFile(extension):
    #global InputFile
    #global IsInputSel
    extname = extension.upper()

    InputFile = filedialog.askopenfilename(title='Выберите УПД', filetypes=((extname, extension),))
    
    if InputFile:
        print('Файл выбран :' + InputFile)
        
        #print(str(IsInputSel))
        return InputFile
    else:
        print('Файл не выбран')
        return False

def SearchCodesFromExcel(mainsheet, inputcolumn,outputcolumn):
    msg_box = messagebox.askquestion('Подтверждение', 'Выбрать файл расшифровки артикулов ?')
    if msg_box == 'yes':
        databookpath = filedialog.askopenfilename(title='Выберите таблицу с кодами', filetypes=(('.XLS', '.xls'),))
        
        if databookpath:

            databook = xlrd.open_workbook(databookpath)
            datasheet = databook.sheet_by_index(0)
            print("Начинаю поиск кодов")
            Expmassive = []
            Datamassive =[]
            
            for i in range(1, mainsheet.max_row+1):
                Expmassive.append(str(mainsheet.cell(i, inputcolumn).value))
                #print (Expmassive)
            for k in range(0,datasheet.nrows):
                Datamassive.append(str(datasheet.cell(k, 1).value))
            
            print(Expmassive)
            print(Datamassive)
            for i in range(1, mainsheet.max_row+1):
                if (i%100==0):
                    print("Пройдено строк "+ str(i))
                for k in range(0,datasheet.nrows):
                    if Expmassive[i-1]== Datamassive[k] or str(Expmassive[i-1])== str(Datamassive[k]) or str(Expmassive[i-1])+".0"==(str(Datamassive[k])):
                        mainsheet.cell(i, outputcolumn).value = datasheet.cell(k,0).value
                        break
                        
def SaveAsXLSX(InputFile):
    encod = False
    book_xls = xlrd.open_workbook(InputFile)
    book_xlsx = Workbook()
    if str(format(book_xls.encoding)) == "iso-8859-1":
        print(str(format(book_xls.encoding))+" Декодирую и сохраняю как xlsx")
        encod = True
    else:
        print(" Сохраняю без декодирования как xlsx")
        encod = False
    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
    if sheet_index == 0:
        sheet_xlsx = book_xlsx.active
        sheet_xlsx.title = sheet_name
    else:
        sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

    for row in range(0, sheet_xls.nrows):
        for col in range(0, sheet_xls.ncols):
            if encod == True:
                sheet_xlsx.cell(row = row+1 , column = col+1).value = Decode(str(sheet_xls.cell_value(row, col)))
            else:
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
    savepath = Path(Path(InputFile).parent, str(Path(InputFile).stem+" Перекодированный.xlsx"))
    #print("Сохранил в " + savepath)
    book_xlsx.save(savepath)
    return savepath


input("Press Enter to continue...")    
InputFile = SelectFile('.xlsx .xls')
extension = Path(InputFile).suffix
if extension == '.xls':
    savepath = SaveAsXLSX(InputFile)
    book = openpyxl.load_workbook(savepath)
elif extension == '.xlsb':
    print("Сохраняю xlsb как xlsx")
    df = pd.read_excel(InputFile, engine='pyxlsb')
    savepath = Path(Path(InputFile).parent, str(Path(InputFile).stem+" Перекодированный.xlsx"))
    df.to_excel(savepath)
    book = openpyxl.load_workbook(savepath)
    print("Сохранение завершено")
else:
    #input("Press Enter to continue...")
    book = openpyxl.load_workbook(InputFile)

input("Press Enter to continue...")
sheet = book.active
SearchCodesFromExcel(sheet,2,1)
try:
    book.save(InputFile)
except:
    msg_box = messagebox.askquestion('Подтверждение', 'Экспортный файл уже открыт, закройте и попробуйте еще раз')
    if msg_box == 'yes':
        book.save(InputFile)
print("Открываю файл с результатами...")
os.system('"{0}"'.format(InputFile))