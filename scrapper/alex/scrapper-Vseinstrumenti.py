from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import os, requests,re
from pathlib import Path
import xlrd
import openpyxl
from openpyxl import Workbook
from configparser import ConfigParser
import time
from urllib.request import Request, urlopen


global root
global scrnwparam
global scrnhparam
scrnwparam = 185
scrnhparam = 150

global myURL
global page
global soup



import requests
import re
from bs4 import BeautifulSoup

def main():
    global root
    global counter
    root = Tk()
    root.resizable(True, True)
        
    scrnw = (root.winfo_screenwidth()//2) - scrnwparam
    scrnh = (root.winfo_screenheight()//2) - scrnhparam
    root.geometry('500x500+{}+{}'.format(scrnw, scrnh))
    #root.resizable(height = None, width = None)
    
    app = GUI(root)
    root.mainloop()

class GUI(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent, background="white")   
        self.parent = parent
        self.parent.title("")
        self.pack(fill=BOTH, expand=1)
        self.initUI()
    
    def initUI(self):

        global counter
        counter = 2
        
        global p3Btn
        p3Btn = Button(text='p3', command=p3)
        p3Btn.place(x=10, y=70, width=40, height=20)
        
        global URLText
        URLText = Text()
        URLText.place(x=10, y=100, width=1000, height=500)
        URLText.insert(END,"https://agisprofile.ru/kabel-kanal-belyi")
        
        myURL = URLText.get("1.0","1.0 lineend")
        print(myURL)
        
        global expfile
        expfile = Workbook()
        global ParameterNames
        ParameterNames = []



def iterate():
    urltest = "https://www.sds-group.ru/ru/catalogue/21588"
    page = requests.get(urltest)
    soup = BeautifulSoup(page.content, 'html.parser')
    print(soup)
    #while(True):
    #    try:
    #        p3()
     #   except:
     #       print("ОООООШШШШИИИИИББККККААА НА "+str(counter))

def p3():
    global nextpage
    global counter
    dir_path = os.path.dirname(os.path.realpath(__file__))
    
    filename = str(counter)+'-Agisprofile.xlsx'
    save_path = os.path.join(dir_path, filename)

    #for i in range(int(URLText.index('end').split(".",1)[0])):
    myURL = URLText.get("1.0","1.0 lineend")
    #print(MyURL)
    agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.36\(KHTML, like Gecko) Chrome/35.0.1916.47 Safari/537.36'

    request = Request(myURL, headers={'User-Agent': agent})
    time.sleep(5)
    html = urlopen(request).read().decode()
    time.sleep(5)
    soup = BeautifulSoup(html, 'html.parser')
    #soup = BeautifulSoup(page)    
    #page = requests.get(myURL)
    #print(7)
    #soup = BeautifulSoup(page.content, 'html.parser')
    print(soup)

    global expfile
    global ParameterNames


    expsheet = expfile.active
    expsheet.title = "Приход"
    #               1        2             3           4        5        6    
    urls = []       
    nextpaging = True
    counter1 = 0
    #while(nextpaging):
    links = soup.find_all('div', {'class' : 'js-product'})
        #js-product t-store__card t-store__stretch-col t-store__stretch-col_25 t-align_left t-item
        #print(soup)
    
    #print(links)
    #for link in links:
        #link = l.find('a',{'data-test':'product-image'})
        #print(link['href'])
        #urls.append(link.find('a')['href'])
        #print(url)
        #print(counter1)
        #nextpageli = soup.find('li',{'class':''})
       # try:
       #     nextpage=nextpageli.find('a')
       #except:
       #     nextpaging=False
        #if nextpage and counter1<200:
         #   print(nextpage['href'])
         #   page = requests.get(nextpage['href'])
         #   #<li class="last-arrow arrow-active"><a href="/catalog_2341_page_2.htm">&nbsp;</a></li>
         #   counter1=counter1+1
        #    soup = BeautifulSoup(page.content, 'html.parser')
        #    lastnextpage= nextpage
            #nextpaging=False
        #else:           
            #URLText.delete('1.0', END)
            #URLText.insert(END,lastnextpage['href'])
            #URLText.update()
        #nextpaging = False
    #<a class="pagination__next" href="/catalog/nizkvoltnoe-oborudovanie/avtomaticheskie-vyklyuchateli/&amp;PAGEN_1=2">След.</a>
    #print(urls)
    if counter == 2:
        fieldnames = ['Описание', 'Фото', 'Артикул','Наименование']
        expsheet.append(fieldnames)

    
    #for url in urls:
        #time.sleep(1)
        #page = requests.get(url)
        

        #soup = BeautifulSoup(page.content, 'html.parser')
        #articlespan = soup.find_all('span', {'class' : 'changeArticle'})
        #for data in articlespan:
        #    article = data.text
        #print(soup)
       #     n=data.text
            
        #print(n)

    counter=counter+1
        #expsheet.cell(counter,1).value = n
        #print(url)


    try:

        descrч = soup.find('div',{'class' : 'js-store-prod-all-charcs'})
        #print(descrч)
        #descr = descrч.find_all('p')
        #print(descr)
        #for n in nomenkls:
            #nomenkl = n.find('span')
            #expsheet.cell(counter,3+nomenkls.index(n)).value = descr.text
        descr1 = descrч.get_text()   #print()
        
        #for d in descr:
        #    descr1 = descr1+str(descr.text)
        print(descr1)
        expsheet.cell(counter,1).value = descr1
        
    except:
        print("FAILURE")
    
    try:
        #<span class="js-store-prod-sku js-product-sku" itemprop="sku">
        article = soup.find('span',{'class' : 'js-store-prod-sku', 'itemprop' : 'sku'})
        print(article)
        #for n in nomenkls:
            #nomenkl = n.find('span')
            #expsheet.cell(counter,3+nomenkls.index(n)).value = descr.text
            #print()
        expsheet.cell(counter,3).value = article.text.strip()
    except:
        print("FAILURE")    
    
    
    #class="box-catalog_detail-top js-shopitem"
    
    try:

        nomenkl = soup.find('meta',{'property' : 'og:title'})
        #print(nomenkls)
        #for n in nomenkls:
            #nomenkl = n.find('span')
            #expsheet.cell(counter,3+nomenkls.index(n)).value = descr.text
            #print()
        expsheet.cell(counter,4).value = nomenkl['content']
    except:
        print("FAILURE")    
    try:
        nomenkl = soup.find('meta',{'itemprop' : 'name'})
        #print(nomenkls)
        #for n in nomenkls:
            #nomenkl = n.find('span')
            #expsheet.cell(counter,3+nomenkls.index(n)).value = descr.text
            #print()
        expsheet.cell(counter,5).value = url
    except:
        print("FAILURE")   



    counterX = 2
    try:
    
        images = soup.find_all('meta', {'itemprop' : 'image'})
        #images = imagesx.find_all('a')
        control = []
        for image in images:
            #img = image.find('img')
            #image = i.find('a')
            if not image in control:
                control.append(image)
                expsheet.cell(counter,counterX).value = str(expsheet.cell(counter,counterX).value)+";"+image['content']
            #counterX=counterX+1
            #print(img['src'])
    except:
        print('fail')
        

    #<div class="box-catalog_detail-tech-wrap"><div class="box-catalog_detail-info-table"><div><div>Цвет</div><div>Прозрачный (светопрониц.)</div></div><div><div>Длина</div><div>1 м</div></div><div><div>Модель/исполнение</div><div>Тонкостенная</div></div><div><div>Материал</div><div>Полиолефин (PEX)</div></div><div><div>Номин. поперечное сечение (диапазон)</div><div>78,50...19,63 мм кв.</div></div><div><div>Не содержит (без) галогенов</div><div>Нет</div></div><div><div>Тип</div><div>Термоусаживаемая (-ый)</div></div><div><div>Рабочая температура</div><div>-40...125 °C</div></div></div><div class="box-catalog_detail-info-table"><div><div>Соотв. стандарту UL (Underwriters Laboratories , США)</div><div>Да</div></div><div><div>Возможность нанесения печатной маркировки</div><div>Да</div></div><div><div>С внутр. клеевым слоем (клеевая)</div><div>Нет</div></div><div><div>Коэффициент усадки</div><div>2:1</div></div><div><div>Толщина стенки после усадки</div><div>0.6 мм</div></div><div><div>Внутр. диаметр до термоусадки</div><div>10 мм</div></div><div><div>Внутр. диаметр после термоусадки</div><div>5 мм</div></div><div><div>Номин. диаметр в дюймах</div><div>Прочее</div></div></div></div>    

    try:#ul = soup.find('ul', {'class' : 'product-properties-list'})    
        characteristicsbox= soup.find('div', {'class' : 'box-catalog_detail-tech-wrap'})
        characteristicsx = characteristicsbox.find('div', {'class' : 'box-catalog_detail-info-table'})
        characteristics = characteristicsx.find_all('div',recursive=False)
        for characteristic in characteristics:
            #print(characteristic)
            name = characteristic.find_all('div')[0]
            value = characteristic.find_all('div')[1]
            #value = values[1]
            #valuea = value.find('a')
            #print(name.text+value.text)
            Nametext = name.text.strip()
            if not name.text.strip() in ParameterNames:
                ParameterNames.append(Nametext)
                expsheet.cell(1,ParameterNames.index(Nametext)+9).value = Nametext
            
            expsheet.cell(counter,ParameterNames.index(Nametext)+9).value = value.text.strip()
    except:
        print('Fail')
    print(counter)
    
    expfile.save(save_path)
    print("The End")

if __name__ == '__main__':
    main()
